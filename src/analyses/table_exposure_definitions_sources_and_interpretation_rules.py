"""Build the exposure definitions, sources, and interpretation rules table."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = (
    ROOT
    / "data/exp/internal_output_archive/tables/Table_exposure_definitions_sources_and_interpretation_rules.xlsx"
)


ROWS = [
    (
        "Historical conflict exposure",
        "Source: Yale CGEO U.S. bombing records. Unique bombing coordinates are collapsed on a 10 m projected grid, counted within the linked survey geography, expressed per 100 km², and transformed as log(1 + density). A structural zero is assigned only after geography linkage is valid.",
    ),
    (
        "Annual extreme-wet rainfall",
        "Source: CHIRPS v2. Annual rainfall is classified as extreme wet when it reaches or exceeds the geography-specific 90th percentile of the fixed 1991–2020 normal. This is a rainfall shock—not a direct measure of flooding.",
    ),
    (
        "Interview-month SPI-12",
        "Source: CHIRPS v2. The 12-month rainfall total ending in the interview month is standardized by geography and end month against 1991–2020 using a fitted gamma distribution and standard-normal transformation. The 2019 interview month is unavailable, so SPI-12 is missing rather than zero.",
    ),
    (
        "Wholesale rice-price shock",
        "Source: WFP market-level food prices. Province-month median low-quality rice log price is measured relative to the same-month national median; the shock is its exact 12-month change. Absent or noncontiguous observations remain missing.",
    ),
    (
        "Broad retail food-price shock",
        "Source: WFP market-level food prices. The measure averages local relative log prices across at least two observed commodities and takes the exact 12-month change. It is a later-wave robustness measure rather than the primary price exposure.",
    ),
    (
        "Survey-year satellite inundation",
        "Source: Global Flood Database v1.4. The measure is the maximum event-level flooded share during the survey year, excluding permanent water. The event product ends in 2018, so 2019 and 2021 remain missing. Use only as secondary flood validation.",
    ),
    (
        "Preceding-12-month satellite inundation",
        "Source: Global Flood Database v1.4. The measure is the maximum event-level flooded share in the 12 months before interview, excluding permanent water. Incomplete windows remain missing; 2007 is unavailable. Use only as secondary flood validation.",
    ),
    (
        "Released analytical samples",
        "Source: Cambodia Socio-Economic Survey (CSES), 2007–2021. The released data contain 62,920 household records and 268,485 person records from repeated cross-sections; households and persons are not followed as panels.",
    ),
    (
        "SPI-12 linked sample",
        "Use outcome-specific observations with nonmissing interview-aligned SPI-12. Do not impose a global complete-case sample across unrelated shocks or outcomes.",
    ),
    (
        "Wholesale-12-month linked sample",
        "Use outcome-specific observations with a nonmissing exact 12-month wholesale rice-price shock. Its narrower temporal and geographic support must be reported separately.",
    ),
    (
        "Coverage statistics",
        "Unweighted coverage is the fraction of observations with a nonmissing linkage. Survey-weighted coverage is the positive survey-weight average of that indicator. Zero-percent coverage means no observations were linked; it does not mean zero exposure.",
    ),
    (
        "Geography counts and linkage",
        "Count PSUs and provinces from household records. Primary specifications use commune linkage where available; district- and province-level fallbacks are retained for sensitivity analysis and must be identified as such.",
    ),
    (
        "Interpretation boundary",
        "Coverage establishes whether a model can be estimated; it does not establish exposure validity or causal identification. Historical conflict is observational, satellite inundation is secondary validation, and missing exposure values must never be recoded as zero.",
    ),
]


def build_frame() -> pd.DataFrame:
    frame = pd.DataFrame(
        ROWS,
        columns=["Item", "Auditable definition / source / rule"],
    )
    assert frame.shape == (13, 2)
    assert frame["Item"].is_unique
    joined = " ".join(frame.iloc[:, 1].tolist())
    for source in ("Yale CGEO", "CHIRPS", "WFP", "Global Flood Database", "CSES"):
        assert source in joined
    assert "must never be recoded as zero" in frame.iloc[-1, 1]
    return frame


def write_workbook(frame: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Definitions"

    for column_index, heading in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=heading)
    for row_index, row in enumerate(frame.itertuples(index=False), start=2):
        sheet.cell(row=row_index, column=1, value=row[0])
        sheet.cell(row=row_index, column=2, value=row[1])

    navy = "17365D"
    pale_blue = "EAF2F8"
    pale_gold = "FFF2CC"
    white = "FFFFFF"
    grey = "D9E2F3"
    thin_grey = Side(style="thin", color="B7C9DC")
    medium_blue = Side(style="medium", color=navy)

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True, size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium_blue)
    sheet.row_dimensions[1].height = 30

    for row_index in range(2, 15):
        item_cell = sheet.cell(row=row_index, column=1)
        rule_cell = sheet.cell(row=row_index, column=2)
        for cell in (item_cell, rule_cell):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_grey)
            cell.font = Font(size=10)
        item_cell.font = Font(size=10, bold=True, color=navy)
        if row_index % 2 == 0:
            item_cell.fill = PatternFill("solid", fgColor="F4F7FA")
            rule_cell.fill = PatternFill("solid", fgColor="F4F7FA")
        sheet.row_dimensions[row_index].height = 52

    # Explicitly mark the transition from exposure definitions to sample/rule rows.
    for cell in sheet[9]:
        cell.border = Border(top=medium_blue, bottom=thin_grey)
        cell.fill = PatternFill("solid", fgColor=pale_blue)
    # Highlight the final interpretation boundary without relying on colour alone.
    for cell in sheet[14]:
        cell.fill = PatternFill("solid", fgColor=pale_gold)
        cell.border = Border(top=medium_blue, bottom=medium_blue)

    sheet.column_dimensions["A"].width = 37
    sheet.column_dimensions["B"].width = 112
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B14"

    table = Table(displayName="ExposureDefinitionsRulesTable", ref="A1:B14")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    sheet.sheet_view.showGridLines = False
    sheet.print_title_rows = "1:1"
    sheet.print_area = "A1:B14"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.1, footer=0.1
    )
    sheet.oddFooter.center.text = ""

    workbook.save(OUTPUT)


def validate_workbook() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == ["Definitions"]
    sheet = workbook["Definitions"]
    assert sheet.max_row == 14
    assert sheet.max_column == 2
    assert list(sheet.tables) == ["ExposureDefinitionsRulesTable"]
    assert sheet.tables["ExposureDefinitionsRulesTable"].ref == "A1:B14"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?"))


def main() -> None:
    frame = build_frame()
    write_workbook(frame)
    validate_workbook()
    print(f"Wrote {OUTPUT}")
    print(f"Dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")


if __name__ == "__main__":
    main()
