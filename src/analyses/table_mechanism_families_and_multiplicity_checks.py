#!/usr/bin/env python3
"""Mechanism Families and Multiplicity Checks.

Plan: Report six activated infrastructure and agricultural-capacity tests,
two village-weighting sensitivities, and one transparent deferral row.
Framework: AnaSOP Sections 5.6 and 6.11. Village persistence models absorb
province-by-wave effects; village and household drought-response models absorb
historical-geography and province-by-wave effects. The six primary tests use
Holm adjustment and are interpreted as channel-consistent associations.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
VILLAGE_PATH = ROOT / "data/processed/direction3_village_mechanisms_preprocessed.parquet"
HOUSEHOLD_PATH = ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
PRIOR_INFERENCE_PATH = ROOT / "data/results/tables/Table_inference_and_multiplicity_checks.xlsx"
OUTPUT_PATH = ROOT / "data/results/tables/Table_mechanism_families_and_multiplicity_checks.xlsx"

YEAR = "Survey Year"
PSU = "PSU"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
PROVINCE = "Province Code Component"
WEIGHT = "Household Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"

VILLAGE_IRRIGATION = "Village Irrigated Agricultural Land Share"
MARKET = "Permanent Market Access"
PARCEL_IRRIGATION = "Irrigable Parcel Share"
DIVERSITY = "Crop Diversity Count"
INPUT_COST = "Real 2021 Agricultural Input Cost Riels"

COLUMNS = [
    "Family",
    "Variable",
    "Waves",
    "Timing",
    "Model",
    "Estimate",
    "95% CI",
    "Raw p-value",
    "Holm p-value",
    "Availability",
    "Channel interpretation",
]


@dataclass(frozen=True)
class Estimate:
    key: str
    coefficient: float
    standard_error: float
    p_value: float
    sample_size: int
    clusters: int
    waves: tuple[int, ...]

    @property
    def lower(self) -> float:
        return self.coefficient - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.coefficient + 1.96 * self.standard_error


def standardize(values: np.ndarray, weights: np.ndarray | None = None) -> np.ndarray:
    if weights is None:
        mean = float(np.mean(values))
        variance = float(np.mean((values - mean) ** 2))
    else:
        mean = float(np.average(values, weights=weights))
        variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def fit_absorbed(
    sample: pd.DataFrame,
    outcome: np.ndarray,
    exogenous: pd.DataFrame,
    absorbed: pd.DataFrame,
    target: str,
    weights: np.ndarray | None = None,
) -> Estimate:
    model = AbsorbingLS(
        dependent=pd.Series(outcome, index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=(
            pd.Series(weights, index=sample.index, name="weight")
            if weights is not None
            else None
        ),
        drop_absorbed=True,
    )
    fitted = model.fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample[GEOGRAPHY]}, index=sample.index),
        debiased=True,
    )
    if target not in fitted.params.index:
        raise ValueError(f"Target parameter was absorbed: {target}")
    return Estimate(
        key="",
        coefficient=float(fitted.params[target]),
        standard_error=float(fitted.std_errors[target]),
        p_value=float(fitted.pvalues[target]),
        sample_size=len(sample),
        clusters=int(sample[GEOGRAPHY].nunique()),
        waves=tuple(sorted(sample[YEAR].astype(int).unique())),
    )


def with_key(result: Estimate, key: str) -> Estimate:
    return Estimate(
        key=key,
        coefficient=result.coefficient,
        standard_error=result.standard_error,
        p_value=result.p_value,
        sample_size=result.sample_size,
        clusters=result.clusters,
        waves=result.waves,
    )


def collapse_household_spine(households: pd.DataFrame) -> pd.DataFrame:
    keys = [YEAR, PSU]
    invariant = [RESOLUTION, GEOGRAPHY, PROVINCE, CONFLICT, SPI12]
    for column in invariant:
        if households.groupby(keys, observed=True)[column].nunique(dropna=False).max() > 1:
            raise ValueError(f"Household exposure field varies within PSU-year: {column}")
    collapsed = households.groupby(keys, as_index=False, observed=True).agg(
        **{column: (column, "first") for column in invariant},
        **{"PSU Household Weight Sum": (WEIGHT, "sum")},
    )
    return collapsed


def prepare_village_panel(
    villages: pd.DataFrame, households: pd.DataFrame
) -> pd.DataFrame:
    spine = collapse_household_spine(households)
    panel = villages.merge(spine, on=[YEAR, PSU], how="left", validate="one_to_one")
    panel[GEOGRAPHY] = panel[GEOGRAPHY].astype("string").str.zfill(6)
    return panel


def fit_village_persistence(
    panel: pd.DataFrame,
    outcome: str,
    key: str,
    use_population_weights: bool = False,
) -> Estimate:
    required = [outcome, CONFLICT, GEOGRAPHY, PROVINCE, YEAR]
    if use_population_weights:
        required.append("PSU Household Weight Sum")
    sample = panel.loc[panel[RESOLUTION].eq("commune"), required].dropna().copy()
    if use_population_weights:
        sample = sample.loc[sample["PSU Household Weight Sum"].gt(0)].copy()
    conflict_z = standardize(sample[CONFLICT].to_numpy(dtype=float))
    exogenous = pd.DataFrame({"conflict": conflict_z}, index=sample.index)
    province_wave = (
        sample[PROVINCE].astype(str) + "_" + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {"province_wave": province_wave.astype("category")}, index=sample.index
    )
    weights = (
        sample["PSU Household Weight Sum"].to_numpy(dtype=float)
        if use_population_weights
        else None
    )
    result = fit_absorbed(
        sample,
        100.0 * sample[outcome].to_numpy(dtype=float),
        exogenous,
        absorbed,
        "conflict",
        weights,
    )
    return with_key(result, key)


def fit_village_drought(
    panel: pd.DataFrame,
    key: str,
    use_population_weights: bool = False,
) -> Estimate:
    required = [
        VILLAGE_IRRIGATION,
        CONFLICT,
        SPI12,
        GEOGRAPHY,
        PROVINCE,
        YEAR,
    ]
    if use_population_weights:
        required.append("PSU Household Weight Sum")
    sample = panel.loc[panel[RESOLUTION].eq("commune"), required].dropna().copy()
    if use_population_weights:
        sample = sample.loc[sample["PSU Household Weight Sum"].gt(0)].copy()
    conflict_z = standardize(sample[CONFLICT].to_numpy(dtype=float))
    drought = -standardize(sample[SPI12].to_numpy(dtype=float))
    exogenous = pd.DataFrame(
        {
            "drought": drought,
            "conflict_x_drought": conflict_z * drought,
        },
        index=sample.index,
    )
    province_wave = (
        sample[PROVINCE].astype(str) + "_" + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {
            "geography": sample[GEOGRAPHY].astype("category"),
            "province_wave": province_wave.astype("category"),
        },
        index=sample.index,
    )
    weights = (
        sample["PSU Household Weight Sum"].to_numpy(dtype=float)
        if use_population_weights
        else None
    )
    result = fit_absorbed(
        sample,
        100.0 * sample[VILLAGE_IRRIGATION].to_numpy(dtype=float),
        exogenous,
        absorbed,
        "conflict_x_drought",
        weights,
    )
    return with_key(result, key)


def fit_household_drought(
    households: pd.DataFrame,
    outcome: str,
    transform: str,
    key: str,
) -> Estimate:
    required = [
        outcome,
        SPI12,
        CONFLICT,
        WEIGHT,
        HOUSEHOLD_SIZE,
        GEOGRAPHY,
        PROVINCE,
        YEAR,
        AGRICULTURAL_HOUSEHOLD,
    ]
    sample = households.loc[
        households[RESOLUTION].eq("commune"), required
    ].dropna().copy()
    sample = sample.loc[
        sample[WEIGHT].gt(0) & sample[AGRICULTURAL_HOUSEHOLD].eq(1)
    ].copy()
    sample[GEOGRAPHY] = sample[GEOGRAPHY].astype(str).str.zfill(6)
    weights = sample[WEIGHT].to_numpy(dtype=float)
    conflict_z = standardize(sample[CONFLICT].to_numpy(dtype=float), weights)
    drought = -standardize(sample[SPI12].to_numpy(dtype=float), weights)
    raw_outcome = sample[outcome].to_numpy(dtype=float)
    if transform == "percentage_points":
        if np.nanmin(raw_outcome) < 0 or np.nanmax(raw_outcome) > 1:
            raise ValueError(f"Share outcome falls outside [0,1]: {outcome}")
        model_outcome = 100.0 * raw_outcome
    elif transform == "asinh":
        if np.nanmin(raw_outcome) < 0:
            raise ValueError(f"Asinh outcome is negative: {outcome}")
        model_outcome = np.arcsinh(raw_outcome)
    elif transform == "level":
        model_outcome = raw_outcome
    else:
        raise ValueError(f"Unknown transform: {transform}")
    exogenous = pd.DataFrame(
        {
            "drought": drought,
            "conflict_x_drought": conflict_z * drought,
            "household_size": sample[HOUSEHOLD_SIZE].to_numpy(dtype=float),
        },
        index=sample.index,
    )
    province_wave = (
        sample[PROVINCE].astype(str) + "_" + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {
            "geography": sample[GEOGRAPHY].astype("category"),
            "province_wave": province_wave.astype("category"),
        },
        index=sample.index,
    )
    result = fit_absorbed(
        sample,
        model_outcome,
        exogenous,
        absorbed,
        "conflict_x_drought",
        weights,
    )
    return with_key(result, key)


def holm_adjust(p_values: list[float]) -> list[float]:
    values = np.asarray(p_values, dtype=float)
    order = np.argsort(values)
    adjusted = np.empty_like(values)
    running_maximum = 0.0
    total = len(values)
    for rank, position in enumerate(order):
        candidate = min(1.0, (total - rank) * values[position])
        running_maximum = max(running_maximum, candidate)
        adjusted[position] = running_maximum
    return adjusted.tolist()


def wave_label(result: Estimate) -> str:
    return ", ".join(map(str, result.waves))


def ci_label(result: Estimate) -> str:
    return f"[{result.lower:.4f}, {result.upper:.4f}]"


def availability(result: Estimate) -> str:
    return f"N={result.sample_size:,}; {result.clusters:,} geography clusters"


def build_table(
    villages: pd.DataFrame, households: pd.DataFrame
) -> tuple[pd.DataFrame, dict[str, Estimate]]:
    panel = prepare_village_panel(villages, households)
    primary = [
        fit_village_persistence(panel, VILLAGE_IRRIGATION, "village_irrigation_gradient"),
        fit_village_persistence(panel, MARKET, "market_gradient"),
        fit_village_drought(panel, "village_irrigation_drought"),
        fit_household_drought(
            households,
            PARCEL_IRRIGATION,
            "percentage_points",
            "parcel_irrigation_drought",
        ),
        fit_household_drought(
            households, DIVERSITY, "level", "crop_diversity_drought"
        ),
        fit_household_drought(
            households, INPUT_COST, "asinh", "input_cost_drought"
        ),
    ]
    sensitivities = [
        fit_village_persistence(
            panel,
            VILLAGE_IRRIGATION,
            "village_irrigation_gradient_weighted",
            use_population_weights=True,
        ),
        fit_village_drought(
            panel,
            "village_irrigation_drought_weighted",
            use_population_weights=True,
        ),
    ]
    results = {result.key: result for result in [*primary, *sensitivities]}
    adjusted = dict(
        zip(
            [result.key for result in primary],
            holm_adjust([result.p_value for result in primary]),
            strict=True,
        )
    )

    specifications = [
        (
            "Village irrigation capacity",
            "village_irrigation_gradient",
            VILLAGE_IRRIGATION,
            "Persistent level",
            "Province-by-wave FE; unweighted",
            "Negative values indicate less village irrigation at greater historical conflict.",
        ),
        (
            "Village market infrastructure",
            "market_gradient",
            MARKET,
            "Persistent level; two waves",
            "Province-by-wave FE; unweighted",
            "Negative values indicate less permanent-market access at greater historical conflict.",
        ),
        (
            "Village irrigation capacity",
            "village_irrigation_drought",
            VILLAGE_IRRIGATION,
            "Interview-window drought response",
            "Geography and province-by-wave FE; unweighted",
            "Negative values align with a larger irrigation decline under drought at greater conflict.",
        ),
        (
            "Household agricultural capacity",
            "parcel_irrigation_drought",
            PARCEL_IRRIGATION,
            "Interview-window drought response",
            "Geography and province-by-wave FE; household weighted",
            "Negative values align with weaker parcel irrigation under drought at greater conflict.",
        ),
        (
            "Household agricultural capacity",
            "crop_diversity_drought",
            DIVERSITY,
            "Interview-window drought response",
            "Geography and province-by-wave FE; household weighted",
            "Negative values align with weaker crop diversification under drought at greater conflict.",
        ),
        (
            "Household agricultural capacity",
            "input_cost_drought",
            INPUT_COST,
            "Interview-window drought response",
            "Geography and province-by-wave FE; household weighted",
            "Negative values align with weaker agricultural investment under drought at greater conflict.",
        ),
        (
            "Village irrigation sensitivity",
            "village_irrigation_gradient_weighted",
            VILLAGE_IRRIGATION,
            "Persistent level",
            "Province-by-wave FE; household-population exposure weighted",
            "Weighting sensitivity only; not an additional multiplicity-family member.",
        ),
        (
            "Village irrigation sensitivity",
            "village_irrigation_drought_weighted",
            VILLAGE_IRRIGATION,
            "Interview-window drought response",
            "Geography and province-by-wave FE; household-population exposure weighted",
            "Weighting sensitivity only; not an additional multiplicity-family member.",
        ),
    ]

    rows: list[dict[str, object]] = []
    for family, key, variable, timing, model, interpretation in specifications:
        result = results[key]
        rows.append(
            {
                "Family": family,
                "Variable": variable,
                "Waves": wave_label(result),
                "Timing": timing,
                "Model": model,
                "Estimate": result.coefficient,
                "95% CI": ci_label(result),
                "Raw p-value": result.p_value,
                "Holm p-value": adjusted.get(key, np.nan),
                "Availability": availability(result),
                "Channel interpretation": interpretation,
            }
        )
    rows.append(
        {
            "Family": "Deferred candidates",
            "Variable": "All-weather road access and five costly-coping indicators",
            "Waves": "Conditional road fields; coping in 2021 only",
            "Timing": "Not activated",
            "Model": "Not estimated",
            "Estimate": np.nan,
            "95% CI": "Not estimable under approved contract",
            "Raw p-value": np.nan,
            "Holm p-value": np.nan,
            "Availability": "Road missingness is routed; coping positives range from 5 to 146 households",
            "Channel interpretation": "Deferred transparently; non-estimation is not evidence of no mechanism.",
        }
    )
    return pd.DataFrame(rows, columns=COLUMNS), results


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mechanism Checks"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "F2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    navy = "17365D"
    blue = "D9EAF7"
    green = "E2F0D9"
    amber = "FFF2CC"
    grey = "E7E6E6"
    white = "FFFFFF"
    text = "1F2937"
    thin = Side(style="thin", color="C7CFD9")
    medium = Side(style="medium", color=navy)

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Arial", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium)
    sheet.row_dimensions[1].height = 42

    family_fills = {
        "Village irrigation capacity": blue,
        "Village market infrastructure": blue,
        "Household agricultural capacity": green,
        "Village irrigation sensitivity": amber,
        "Deferred candidates": grey,
    }
    for row_index in range(2, len(frame) + 2):
        family = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, len(COLUMNS) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = Font(name="Arial", size=9, color=text)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {6, 8, 9} else "left",
                vertical="center",
                wrap_text=column_index not in {6, 8, 9},
            )
            cell.border = Border(bottom=thin)
        sheet.cell(row=row_index, column=1).fill = PatternFill(
            "solid", fgColor=family_fills[family]
        )
        sheet.cell(row=row_index, column=1).font = Font(
            name="Arial", size=9, bold=True, color=navy
        )
        sheet.row_dimensions[row_index].height = 56 if family == "Deferred candidates" else 46

    for column in ("F", "H", "I"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.0000"

    widths = [24, 39, 25, 28, 44, 13, 23, 14, 14, 33, 52]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    last_row = len(frame) + 1
    table = Table(displayName="MechanismMultiplicityTable", ref=f"A1:K{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.auto_filter.ref = f"A1:K{last_row}"
    sheet.sheet_view.zoomScale = 65
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:K{last_row}"
    sheet.page_margins.left = 0.12
    sheet.page_margins.right = 0.12
    sheet.page_margins.top = 0.18
    sheet.page_margins.bottom = 0.18

    workbook.properties.title = "Mechanism Families and Multiplicity Checks"
    workbook.properties.subject = "Expanded infrastructure and agricultural-capacity mechanism family"
    workbook.properties.creator = "Mike Li"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame, results: dict[str, Estimate]) -> None:
    assert frame.shape == (9, 11), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame.iloc[:6]["Holm p-value"].notna().all()
    assert frame.iloc[6:]["Holm p-value"].isna().all()
    assert frame.iloc[:8]["Raw p-value"].between(0, 1).all()
    assert frame.iloc[:6]["Holm p-value"].between(0, 1).all()
    assert frame.iloc[:8]["Estimate"].notna().all()

    prior = pd.read_excel(PRIOR_INFERENCE_PATH, sheet_name="Inference and Multiplicity")
    prior = prior.loc[prior["Domain"].eq("Mechanism")].reset_index(drop=True)
    keys = [
        "parcel_irrigation_drought",
        "crop_diversity_drought",
        "input_cost_drought",
    ]
    assert len(prior) == len(keys)
    for row, key in zip(prior.itertuples(index=False), keys, strict=True):
        assert np.isclose(float(row.Estimate), results[key].coefficient, atol=1e-10)
        assert np.isclose(
            float(getattr(row, "_4")), results[key].p_value, atol=1e-10
        )

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Mechanism Checks"]
    sheet = workbook["Mechanism Checks"]
    assert sheet.max_row == 10
    assert sheet.max_column == 11
    assert list(sheet.tables) == ["MechanismMultiplicityTable"]
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    village_columns = [YEAR, PSU, VILLAGE_IRRIGATION, MARKET]
    household_columns = [
        YEAR,
        PSU,
        RESOLUTION,
        GEOGRAPHY,
        PROVINCE,
        WEIGHT,
        HOUSEHOLD_SIZE,
        AGRICULTURAL_HOUSEHOLD,
        CONFLICT,
        SPI12,
        PARCEL_IRRIGATION,
        DIVERSITY,
        INPUT_COST,
    ]
    villages = pd.read_parquet(VILLAGE_PATH, columns=village_columns)
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    frame, results = build_table(villages, households)
    write_workbook(frame)
    validate_output(frame, results)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(frame.to_string(index=False))


if __name__ == "__main__":
    main()
