#!/usr/bin/env python3
"""Mechanisms Heterogeneity and Influence Checks.

Plan: Report 18 mechanism, subgroup, leave-one-province, leave-one-wave, and
outlier-influence diagnostics for the conflict-conditioned shock design.
Framework: AnaSOP Sections 5.1-5.2, 6.2-6.7, and the mechanism and influence
workflow steps in Section 7. Mechanism estimates are supporting associations,
not causal mediation effects; outlier exclusions are sensitivity checks only.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
INPUT_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT
    / "data/results/tables/Table_mechanisms_heterogeneity_and_influence_checks.xlsx"
)

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
PROVINCE = "Province Code Component"
WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
EXTREME_WET = "Annual Rainfall Extreme Wet Shock"
PRICE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
URBAN_RURAL = "Urban Rural"
URBAN_BINARY = "Urban Household"
EDUCATION_PROVINCE = "Province Code"

IRRIGATION = "Irrigable Parcel Share"
DIVERSITY = "Crop Diversity Count"
INPUT_COST = "Real 2021 Agricultural Input Cost Riels"
CROP_YIELD = "Crop Yield kg per ha"
CROP_VALUE = "Real 2021 Crop Production Value Riels"
LOSS_SHARE = "Post Harvest Loss Share"
FOOD_CONSUMPTION = "Real 2021 Food Consumption Value per Household Member Riels"
FOOD_ITEMS = "Food Items with Positive Consumption Count"
SEVERE_FOOD_INSECURITY = "Any Severe Food Insecurity Experience"
ATTENDANCE = "Currently Attending School"

COLUMNS = [
    "Diagnostic family",
    "Outcome or mechanism (model scale)",
    "Shock",
    "Estimand or diagnostic",
    "Estimate",
    "95% CI",
    "N",
    "Subgroup or exclusion",
    "Diagnostic result",
    "Interpretation",
]


@dataclass(frozen=True)
class OutcomeSpec:
    variable: str
    label: str
    transform: str
    agriculture_only: bool = False


@dataclass(frozen=True)
class CoreSpec:
    outcome: OutcomeSpec
    shock_kind: str
    shock_label: str


@dataclass(frozen=True)
class ModelResult:
    estimate: float
    standard_error: float
    sample_size: int
    cluster_count: int

    @property
    def lower(self) -> float:
        return self.estimate - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.estimate + 1.96 * self.standard_error


MECHANISMS = [
    OutcomeSpec(IRRIGATION, "Irrigable parcel share (percentage points)", "percentage_points", True),
    OutcomeSpec(DIVERSITY, "Crop diversity (count)", "level", True),
    OutcomeSpec(INPUT_COST, "Agricultural input cost (asinh, 2021 riels)", "asinh", True),
]

AGRICULTURAL_HETEROGENEITY = [
    OutcomeSpec(FOOD_CONSUMPTION, "Food consumption per member (asinh)", "asinh"),
    OutcomeSpec(FOOD_ITEMS, "Food items with positive consumption (count)", "level"),
    OutcomeSpec(SEVERE_FOOD_INSECURITY, "Severe food insecurity (percentage points)", "percentage_points"),
]

PLACE_HETEROGENEITY = [
    OutcomeSpec(FOOD_CONSUMPTION, "Food consumption per member (asinh)", "asinh"),
    OutcomeSpec(SEVERE_FOOD_INSECURITY, "Severe food insecurity (percentage points)", "percentage_points"),
]

INFLUENCE_SPECS = [
    CoreSpec(
        OutcomeSpec(CROP_YIELD, "Crop yield (asinh)", "asinh", True),
        "drought",
        "Drought severity (−SPI-12; 1 SD)",
    ),
    CoreSpec(
        OutcomeSpec(CROP_YIELD, "Crop yield (asinh)", "asinh", True),
        "wet",
        "Annual extreme wet (0→1)",
    ),
    CoreSpec(
        OutcomeSpec(FOOD_CONSUMPTION, "Food consumption per member (asinh)", "asinh"),
        "price",
        "Local rice-price change (1 SD)",
    ),
    CoreSpec(
        OutcomeSpec(
            SEVERE_FOOD_INSECURITY,
            "Severe food insecurity (percentage points)",
            "percentage_points",
        ),
        "price",
        "Local rice-price change (1 SD)",
    ),
]

OUTLIER_SPECS = [
    CoreSpec(
        OutcomeSpec(CROP_VALUE, "Crop production value (asinh, 2021 riels)", "asinh", True),
        "drought",
        "Drought severity (−SPI-12; 1 SD)",
    ),
    CoreSpec(
        OutcomeSpec(FOOD_CONSUMPTION, "Food consumption per member (asinh)", "asinh"),
        "price",
        "Local rice-price change (1 SD)",
    ),
]


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def transform_outcome(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "asinh":
        if np.nanmin(values) < 0:
            raise ValueError("asinh outcome unexpectedly contains negative values")
        return np.arcsinh(values)
    if transform == "percentage_points":
        if np.nanmin(values) < 0 or np.nanmax(values) > 1:
            raise ValueError("Share or binary outcome falls outside [0, 1]")
        return 100.0 * values
    if transform == "level":
        return values
    raise ValueError(f"Unknown outcome transform: {transform}")


def shock_variable(shock_kind: str) -> str:
    return {"drought": SPI12, "wet": EXTREME_WET, "price": PRICE_12M}[shock_kind]


def build_model_sample(
    data: pd.DataFrame,
    outcome: OutcomeSpec,
    shock_kind: str,
    extra_columns: list[str] | None = None,
) -> pd.DataFrame:
    shock = shock_variable(shock_kind)
    required = [
        outcome.variable,
        shock,
        CONFLICT,
        WEIGHT,
        HOUSEHOLD_SIZE,
        RESOLUTION,
        GEOGRAPHY,
        PROVINCE,
        YEAR,
        *(extra_columns or []),
    ]
    if outcome.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    required = list(dict.fromkeys(required))
    sample = data.loc[data[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[sample[WEIGHT].gt(0)].copy()
    if outcome.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)].copy()
    if sample[GEOGRAPHY].nunique() < 50:
        raise ValueError(f"Too few geography clusters for {outcome.label}")

    weights = sample[WEIGHT].to_numpy(dtype=float)
    sample["_outcome_raw"] = sample[outcome.variable].to_numpy(dtype=float)
    sample["_outcome"] = transform_outcome(sample["_outcome_raw"].to_numpy(), outcome.transform)
    sample["_conflict_z"] = weighted_standardize(
        sample[CONFLICT].to_numpy(dtype=float), weights
    )
    raw_shock = sample[shock].to_numpy(dtype=float)
    if shock_kind == "drought":
        sample["_shock"] = -weighted_standardize(raw_shock, weights)
    elif shock_kind == "price":
        sample["_shock"] = weighted_standardize(raw_shock, weights)
    else:
        if not pd.Series(raw_shock).isin([0, 1]).all():
            raise ValueError("Extreme-wet shock is not binary")
        sample["_shock"] = raw_shock
    sample["_geography"] = sample[GEOGRAPHY].astype(str)
    sample["_province_wave"] = (
        sample[PROVINCE].astype(str) + "_" + sample[YEAR].astype(int).astype(str)
    )
    return sample


def fit_absorbed(
    sample: pd.DataFrame,
    exogenous: pd.DataFrame,
    target: str,
) -> ModelResult:
    absorbed = pd.DataFrame(
        {
            "geography": sample["_geography"].astype("category"),
            "province_wave": sample["_province_wave"].astype("category"),
        },
        index=sample.index,
    )
    fitted = AbsorbingLS(
        dependent=pd.Series(sample["_outcome"].to_numpy(), index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(sample[WEIGHT].to_numpy(dtype=float), index=sample.index),
        drop_absorbed=True,
    ).fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample["_geography"]}, index=sample.index),
        debiased=True,
    )
    if target not in fitted.params.index:
        raise ValueError(f"Target parameter was absorbed: {target}")
    return ModelResult(
        estimate=float(fitted.params[target]),
        standard_error=float(fitted.std_errors[target]),
        sample_size=len(sample),
        cluster_count=int(sample["_geography"].nunique()),
    )


def fit_interaction(sample: pd.DataFrame) -> ModelResult:
    exogenous = pd.DataFrame(
        {
            "shock": sample["_shock"].to_numpy(dtype=float),
            "conflict_x_shock": (
                sample["_conflict_z"].to_numpy(dtype=float)
                * sample["_shock"].to_numpy(dtype=float)
            ),
            "household_size": sample[HOUSEHOLD_SIZE].to_numpy(dtype=float),
        },
        index=sample.index,
    )
    return fit_absorbed(sample, exogenous, "conflict_x_shock")


def build_education_price_sample(data: pd.DataFrame) -> pd.DataFrame:
    required = [
        ATTENDANCE,
        PRICE_12M,
        CONFLICT,
        PERSON_WEIGHT,
        HOUSEHOLD_SIZE,
        AGE,
        FEMALE,
        RESOLUTION,
        GEOGRAPHY,
        EDUCATION_PROVINCE,
        YEAR,
    ]
    sample = data.loc[data[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[
        sample[PERSON_WEIGHT].gt(0) & sample[AGE].between(6, 17)
    ].copy()
    if not sample[ATTENDANCE].isin([0, 1, False, True]).all():
        raise ValueError("School-attendance outcome is not binary")
    weights = sample[PERSON_WEIGHT].to_numpy(dtype=float)
    sample["_outcome"] = 100.0 * sample[ATTENDANCE].to_numpy(dtype=float)
    sample["_conflict_z"] = weighted_standardize(
        sample[CONFLICT].to_numpy(dtype=float), weights
    )
    sample["_shock"] = weighted_standardize(
        sample[PRICE_12M].to_numpy(dtype=float), weights
    )
    sample["_geography"] = sample[GEOGRAPHY].astype(str)
    sample["_province_wave"] = (
        sample[EDUCATION_PROVINCE].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    return sample


def fit_education_price_interaction(sample: pd.DataFrame) -> ModelResult:
    exogenous = pd.DataFrame(
        {
            "shock": sample["_shock"].to_numpy(dtype=float),
            "conflict_x_shock": (
                sample["_conflict_z"].to_numpy(dtype=float)
                * sample["_shock"].to_numpy(dtype=float)
            ),
            "household_size": sample[HOUSEHOLD_SIZE].to_numpy(dtype=float),
            "age": sample[AGE].to_numpy(dtype=float),
            "female": sample[FEMALE].to_numpy(dtype=float),
        },
        index=sample.index,
    )
    absorbed = pd.DataFrame(
        {
            "geography": sample["_geography"].astype("category"),
            "province_wave": sample["_province_wave"].astype("category"),
        },
        index=sample.index,
    )
    fitted = AbsorbingLS(
        dependent=pd.Series(sample["_outcome"].to_numpy(), index=sample.index),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(sample[PERSON_WEIGHT].to_numpy(dtype=float), index=sample.index),
        drop_absorbed=True,
    ).fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample["_geography"]}, index=sample.index),
        debiased=True,
    )
    target = "conflict_x_shock"
    return ModelResult(
        estimate=float(fitted.params[target]),
        standard_error=float(fitted.std_errors[target]),
        sample_size=len(sample),
        cluster_count=int(sample["_geography"].nunique()),
    )


def fit_three_way(sample: pd.DataFrame, subgroup: str) -> ModelResult:
    h = sample[subgroup].to_numpy(dtype=float)
    conflict = sample["_conflict_z"].to_numpy(dtype=float)
    shock = sample["_shock"].to_numpy(dtype=float)
    exogenous = pd.DataFrame(
        {
            "shock": shock,
            "subgroup": h,
            "conflict_x_shock": conflict * shock,
            "shock_x_subgroup": shock * h,
            "conflict_x_subgroup": conflict * h,
            "conflict_x_shock_x_subgroup": conflict * shock * h,
            "household_size": sample[HOUSEHOLD_SIZE].to_numpy(dtype=float),
        },
        index=sample.index,
    )
    return fit_absorbed(sample, exogenous, "conflict_x_shock_x_subgroup")


def format_ci(result: ModelResult) -> str:
    return f"[{result.lower:.3f}, {result.upper:.3f}]"


def precision_text(result: ModelResult) -> str:
    if result.lower > 0 or result.upper < 0:
        return "95% CI excludes zero"
    return "95% CI includes zero"


def base_row(
    family: str,
    outcome_label: str,
    shock_label: str,
    estimand: str,
    result: ModelResult,
    subgroup: str,
    diagnostic: str,
    interpretation: str,
) -> dict[str, object]:
    return {
        "Diagnostic family": family,
        "Outcome or mechanism (model scale)": outcome_label,
        "Shock": shock_label,
        "Estimand or diagnostic": estimand,
        "Estimate": result.estimate,
        "95% CI": format_ci(result),
        "N": result.sample_size,
        "Subgroup or exclusion": subgroup,
        "Diagnostic result": diagnostic,
        "Interpretation": interpretation,
    }


def mechanism_rows(data: pd.DataFrame) -> list[dict[str, object]]:
    rows = []
    for outcome in MECHANISMS:
        sample = build_model_sample(data, outcome, "drought", [AGRICULTURAL_HOUSEHOLD])
        result = fit_interaction(sample)
        rows.append(
            base_row(
                "Adaptive-capacity mechanism",
                outcome.label,
                "Drought severity (−SPI-12; 1 SD)",
                "Conflict × drought",
                result,
                "Agricultural households",
                precision_text(result),
                "Supporting channel association; not causal mediation",
            )
        )
    return rows


def heterogeneity_rows(data: pd.DataFrame) -> list[dict[str, object]]:
    rows = []
    for outcome in AGRICULTURAL_HETEROGENEITY:
        sample = build_model_sample(data, outcome, "drought", [AGRICULTURAL_HOUSEHOLD])
        result = fit_three_way(sample, AGRICULTURAL_HOUSEHOLD)
        rows.append(
            base_row(
                "Agricultural-dependence heterogeneity",
                outcome.label,
                "Drought severity (−SPI-12; 1 SD)",
                "Conflict × drought × agricultural household",
                result,
                "Agricultural vs non-agricultural; final preprocessed indicator",
                precision_text(result),
                "Difference in amplification by agricultural dependence",
            )
        )
    for outcome in PLACE_HETEROGENEITY:
        sample = build_model_sample(data, outcome, "drought", [URBAN_BINARY])
        result = fit_three_way(sample, URBAN_BINARY)
        rows.append(
            base_row(
                "Place heterogeneity",
                outcome.label,
                "Drought severity (−SPI-12; 1 SD)",
                "Conflict × drought × urban household",
                result,
                "Urban (code 1) vs rural (code 2)",
                precision_text(result),
                "Place-type difference; not an early-life exposure test",
            )
        )
    return rows


def leave_one_out_row(
    data: pd.DataFrame,
    spec: CoreSpec,
    exclusion_variable: str,
    family: str,
) -> dict[str, object]:
    sample = build_model_sample(data, spec.outcome, spec.shock_kind)
    full = fit_interaction(sample)
    estimates: list[float] = []
    values = sorted(sample[exclusion_variable].dropna().unique().tolist())
    for value in values:
        reduced = sample.loc[~sample[exclusion_variable].eq(value)].copy()
        estimates.append(fit_interaction(reduced).estimate)
    minimum = float(min(estimates))
    maximum = float(max(estimates))
    full_sign = np.sign(full.estimate)
    signs_retained = int(sum(np.sign(value) == full_sign for value in estimates))
    max_shift = float(max(abs(value - full.estimate) for value in estimates))
    unit = "province" if exclusion_variable == PROVINCE else "survey wave"
    diagnostic = (
        f"Range [{minimum:.3f}, {maximum:.3f}]; sign retained "
        f"{signs_retained}/{len(values)}; max |Δ|={max_shift:.3f}"
    )
    interpretation = (
        f"No single {unit} changes the sign"
        if signs_retained == len(values)
        else f"Sign changes under {len(values) - signs_retained} {unit} exclusion(s)"
    )
    return base_row(
        family,
        spec.outcome.label,
        spec.shock_label,
        "Full-sample conflict × shock; leave-one-out range",
        full,
        f"One of {len(values)} {unit}s omitted per re-estimation",
        diagnostic,
        interpretation,
    )


def education_leave_one_out_row(
    data: pd.DataFrame,
    exclusion_variable: str,
    family: str,
) -> dict[str, object]:
    sample = build_education_price_sample(data)
    full = fit_education_price_interaction(sample)
    estimates: list[float] = []
    values = sorted(sample[exclusion_variable].dropna().unique().tolist())
    for value in values:
        reduced = sample.loc[~sample[exclusion_variable].eq(value)].copy()
        estimates.append(fit_education_price_interaction(reduced).estimate)
    minimum = float(min(estimates))
    maximum = float(max(estimates))
    full_sign = np.sign(full.estimate)
    signs_retained = int(sum(np.sign(value) == full_sign for value in estimates))
    max_shift = float(max(abs(value - full.estimate) for value in estimates))
    unit = "province" if exclusion_variable == EDUCATION_PROVINCE else "survey wave"
    diagnostic = (
        f"Range [{minimum:.3f}, {maximum:.3f}]; sign retained "
        f"{signs_retained}/{len(values)}; max |Δ|={max_shift:.3f}"
    )
    interpretation = (
        f"No single {unit} changes the sign"
        if signs_retained == len(values)
        else f"Sign changes under {len(values) - signs_retained} {unit} exclusion(s)"
    )
    return base_row(
        family,
        "School attendance, ages 6–17 (percentage points)",
        "Local rice-price change (1 SD)",
        "Full-sample conflict × shock; leave-one-out range",
        full,
        f"One of {len(values)} {unit}s omitted per re-estimation",
        diagnostic,
        interpretation,
    )


def outlier_row(data: pd.DataFrame, spec: CoreSpec) -> dict[str, object]:
    sample = build_model_sample(data, spec.outcome, spec.shock_kind)
    full = fit_interaction(sample)
    threshold = float(sample["_outcome_raw"].quantile(0.99))
    trimmed = sample.loc[sample["_outcome_raw"].le(threshold)].copy()
    result = fit_interaction(trimmed)
    shift = result.estimate - full.estimate
    sign_retained = np.sign(result.estimate) == np.sign(full.estimate)
    return base_row(
        "Influential observations",
        spec.outcome.label,
        spec.shock_label,
        "Conflict × shock after upper-tail exclusion",
        result,
        "Exclude raw outcome above empirical p99; sensitivity only",
        (
            f"Full {full.estimate:.3f} → sensitivity {result.estimate:.3f}; "
            f"Δ={shift:.3f}; sign {'retained' if sign_retained else 'changed'}"
        ),
        "Primary sample remains untrimmed",
    )


def build_table(data: pd.DataFrame, education: pd.DataFrame) -> pd.DataFrame:
    rows = mechanism_rows(data)
    rows.extend(heterogeneity_rows(data))
    rows.extend(
        leave_one_out_row(data, spec, PROVINCE, "Leave-one-province-out")
        for spec in INFLUENCE_SPECS
    )
    rows.append(
        education_leave_one_out_row(
            education, EDUCATION_PROVINCE, "Leave-one-province-out"
        )
    )
    rows.extend(
        leave_one_out_row(data, spec, YEAR, "Leave-one-wave-out")
        for spec in INFLUENCE_SPECS
    )
    rows.append(
        education_leave_one_out_row(education, YEAR, "Leave-one-wave-out")
    )
    rows.extend(outlier_row(data, spec) for spec in OUTLIER_SPECS)
    return pd.DataFrame(rows, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mechanisms and Influence"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "E2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="MechanismsInfluenceTable", ref=f"A1:J{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    family_fills = {
        "Adaptive-capacity mechanism": PatternFill("solid", fgColor="EAF2F8"),
        "Agricultural-dependence heterogeneity": PatternFill("solid", fgColor="E2F0D9"),
        "Place heterogeneity": PatternFill("solid", fgColor="FFF2CC"),
        "Leave-one-province-out": PatternFill("solid", fgColor="F2F2F2"),
        "Leave-one-wave-out": PatternFill("solid", fgColor="DDEBF7"),
        "Influential observations": PatternFill("solid", fgColor="E4DFEC"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=8.5)
    navy_bold = Font(color="1F4E78", bold=True, size=8.5)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")
    section_start_rows = {2, 5, 8, 10, 15, 20}

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 50

    for row_index in range(2, last_row + 1):
        family = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 11):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {5, 7} else "left",
                vertical="center",
                wrap_text=column_index not in {5, 7},
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = family_fills[family]
        sheet.cell(row=row_index, column=1).font = navy_bold
        sheet.row_dimensions[row_index].height = 46

    for cell in sheet["E"][1:]:
        cell.number_format = "0.000"
    for cell in sheet["G"][1:]:
        cell.number_format = "#,##0"

    widths = [31, 39, 29, 41, 15, 20, 13, 48, 49, 42]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:J{last_row}"
    sheet.sheet_view.zoomScale = 65
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:J{last_row}"
    sheet.page_margins.left = 0.12
    sheet.page_margins.right = 0.12
    sheet.page_margins.top = 0.16
    sheet.page_margins.bottom = 0.16

    workbook.properties.title = "Mechanisms Heterogeneity and Influence Checks"
    workbook.properties.subject = "Direction 3 mechanism, subgroup, and influence diagnostics"
    workbook.properties.creator = "Mike Li"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (20, 10), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["N"].gt(0).all()
    assert np.isfinite(frame["Estimate"]).all()
    assert frame["95% CI"].str.match(r"^\[-?\d+\.\d{3}, -?\d+\.\d{3}\]$").all()
    assert frame.groupby("Diagnostic family").size().to_dict() == {
        "Adaptive-capacity mechanism": 3,
        "Agricultural-dependence heterogeneity": 3,
        "Influential observations": 2,
        "Leave-one-province-out": 5,
        "Leave-one-wave-out": 5,
        "Place heterogeneity": 2,
    }

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Mechanisms and Influence"]
    sheet = workbook["Mechanisms and Influence"]
    assert sheet.max_row == 21
    assert sheet.max_column == 10
    assert list(sheet.tables) == ["MechanismsInfluenceTable"]
    assert sheet.tables["MechanismsInfluenceTable"].ref == "A1:J21"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            PROVINCE,
            WEIGHT,
            HOUSEHOLD_SIZE,
            CONFLICT,
            SPI12,
            EXTREME_WET,
            PRICE_12M,
            AGRICULTURAL_HOUSEHOLD,
            URBAN_RURAL,
            *[spec.variable for spec in MECHANISMS],
            *[spec.variable for spec in AGRICULTURAL_HETEROGENEITY],
            *[spec.variable for spec in PLACE_HETEROGENEITY],
            *[spec.outcome.variable for spec in INFLUENCE_SPECS],
            *[spec.outcome.variable for spec in OUTLIER_SPECS],
        }
    )
    data = pd.read_parquet(INPUT_PATH, columns=columns)
    if (
        data[AGRICULTURAL_HOUSEHOLD].isna().any()
        or not data[AGRICULTURAL_HOUSEHOLD].isin([False, True]).all()
    ):
        raise ValueError("Agricultural Household must be a complete binary indicator")
    urban_rural_numeric = pd.to_numeric(data[URBAN_RURAL], errors="coerce")
    if urban_rural_numeric.isna().any() or not urban_rural_numeric.isin([1, 2]).all():
        raise ValueError("Urban/rural codes fall outside expected 1/2 coding")
    data[URBAN_BINARY] = urban_rural_numeric.eq(1).astype(float)

    education_columns = [
        YEAR,
        RESOLUTION,
        GEOGRAPHY,
        EDUCATION_PROVINCE,
        PERSON_WEIGHT,
        HOUSEHOLD_SIZE,
        AGE,
        FEMALE,
        CONFLICT,
        PRICE_12M,
        ATTENDANCE,
    ]
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)

    table = build_table(data, education)
    write_workbook(table)
    validate_output(table)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {table.shape[0]} rows x {table.shape[1]} columns")
    print(table.groupby("Diagnostic family").size().to_string())


if __name__ == "__main__":
    main()
