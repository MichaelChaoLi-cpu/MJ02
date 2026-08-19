#!/usr/bin/env python3
"""Survey-Weighted Linkage Coverage.

Plan: Test whether exposure availability differs materially after applying
household and person survey weights across survey waves and restricted samples.
Framework: Supports AnaSOP Sections 5.1, 6.5, and the sample-audit step in
Section 7 by comparing survey-weighted shock-specific coverage denominators.
Missing exposure is never recoded as zero.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = ROOT / "data/results/tables/Table_survey_weighted_linkage_coverage.xlsx"

YEAR = "Survey Year"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
WHOLESALE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"
FLOOD_YEAR = "Survey Year Maximum Flooded Geography Share"
EXPECTED_WAVES = [2007, 2009, 2011, 2013, 2014, 2016, 2017, 2019, 2021]

COLUMNS = [
    "Sample / survey wave",
    "Household weighted: conflict (%)",
    "Household weighted: SPI-12 (%)",
    "Household weighted: wholesale 12m (%)",
    "Household weighted: satellite flood (%)",
    "Person weighted: conflict (%)",
    "Person weighted: SPI-12 (%)",
    "Person weighted: wholesale 12m (%)",
    "Person weighted: satellite flood (%)",
]


def coverage(frame: pd.DataFrame, variable: str) -> pd.Series:
    """Return nonmissing exposure coverage without converting missing to zero."""
    return frame[variable].notna()


def weighted_share(indicator: pd.Series, weights: pd.Series) -> float:
    valid = indicator.notna() & weights.notna() & weights.gt(0)
    if not valid.any():
        return np.nan
    return float(
        np.average(indicator.loc[valid].astype(float), weights=weights.loc[valid])
    )


def weighted_row(
    label: str, households: pd.DataFrame, people: pd.DataFrame
) -> dict[str, object]:
    return {
        "Sample / survey wave": label,
        "Household weighted: conflict (%)": weighted_share(
            coverage(households, CONFLICT), households[HOUSEHOLD_WEIGHT]
        ),
        "Household weighted: SPI-12 (%)": weighted_share(
            coverage(households, SPI12), households[HOUSEHOLD_WEIGHT]
        ),
        "Household weighted: wholesale 12m (%)": weighted_share(
            coverage(households, WHOLESALE_12M), households[HOUSEHOLD_WEIGHT]
        ),
        "Household weighted: satellite flood (%)": weighted_share(
            coverage(households, FLOOD_YEAR), households[HOUSEHOLD_WEIGHT]
        ),
        "Person weighted: conflict (%)": weighted_share(
            coverage(people, CONFLICT), people[PERSON_WEIGHT]
        ),
        "Person weighted: SPI-12 (%)": weighted_share(
            coverage(people, SPI12), people[PERSON_WEIGHT]
        ),
        "Person weighted: wholesale 12m (%)": weighted_share(
            coverage(people, WHOLESALE_12M), people[PERSON_WEIGHT]
        ),
        "Person weighted: satellite flood (%)": weighted_share(
            coverage(people, FLOOD_YEAR), people[PERSON_WEIGHT]
        ),
    }


def build_sample_rows(
    households: pd.DataFrame, people: pd.DataFrame
) -> list[tuple[str, pd.DataFrame, pd.DataFrame]]:
    rows = [
        (
            str(year),
            households.loc[households[YEAR].eq(year)],
            people.loc[people[YEAR].eq(year)],
        )
        for year in EXPECTED_WAVES
    ]
    rows.extend(
        [
            ("All survey waves", households, people),
            (
                "SPI-12 linked sample",
                households.loc[coverage(households, SPI12)],
                people.loc[coverage(people, SPI12)],
            ),
            (
                "Wholesale 12m linked sample",
                households.loc[coverage(households, WHOLESALE_12M)],
                people.loc[coverage(people, WHOLESALE_12M)],
            ),
        ]
    )
    return rows


def build_table(households: pd.DataFrame, people: pd.DataFrame) -> pd.DataFrame:
    rows = build_sample_rows(households, people)
    return pd.DataFrame(
        [
            weighted_row(label, household_sample, person_sample)
            for label, household_sample, person_sample in rows
        ],
        columns=COLUMNS,
    )


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Weighted Coverage"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="SurveyWeightedCoverageTable", ref=f"A1:I{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    white_bold = Font(color="FFFFFF", bold=True, size=9)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    restricted_fill = PatternFill("solid", fgColor="FFF2CC")
    light_rule = Side(style="thin", color="B7C9D6")
    section_rule = Side(style="medium", color="7F9DB9")
    group_rule = Side(style="medium", color="A6BFD3")
    for column_index, cell in enumerate(worksheet[1], start=1):
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=light_rule,
            left=group_rule if column_index == 6 else None,
        )
    worksheet.row_dimensions[1].height = 48

    for row_index in range(2, last_row + 1):
        label = str(worksheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 10):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="left" if column_index == 1 else "right",
                vertical="center",
                wrap_text=column_index == 1,
                indent=1 if column_index == 1 else 0,
            )
            if column_index == 6:
                cell.border = Border(left=group_rule)
        if label == "All survey waves":
            for column_index, cell in enumerate(worksheet[row_index], start=1):
                cell.fill = total_fill
                cell.font = Font(bold=True, size=9)
                cell.border = Border(
                    top=section_rule,
                    left=group_rule if column_index == 6 else None,
                )
        elif label in {"SPI-12 linked sample", "Wholesale 12m linked sample"}:
            for cell in worksheet[row_index]:
                cell.fill = restricted_fill
        worksheet.row_dimensions[row_index].height = 27

    for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=9):
        for cell in row:
            cell.number_format = "0.0%"

    color_scale = ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color="F8696B",
        mid_type="num",
        mid_value=0.75,
        mid_color="FFEB84",
        end_type="num",
        end_value=1,
        end_color="63BE7B",
    )
    worksheet.conditional_formatting.add(f"B2:I{last_row}", color_scale)

    widths = [30, 25, 25, 29, 28, 25, 25, 29, 28]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:I{last_row}"
    worksheet.sheet_view.zoomScale = 75
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:I{last_row}"
    worksheet.page_margins.left = 0.18
    worksheet.page_margins.right = 0.18
    worksheet.page_margins.top = 0.20
    worksheet.page_margins.bottom = 0.20

    workbook.properties.title = "Survey-Weighted Linkage Coverage"
    workbook.properties.subject = "Direction 3 survey-weighted exposure coverage audit"
    workbook.properties.creator = "Mike Li"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (12, 9), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["Sample / survey wave"].tolist() == [
        *[str(year) for year in EXPECTED_WAVES],
        "All survey waves",
        "SPI-12 linked sample",
        "Wholesale 12m linked sample",
    ]
    rates = frame.iloc[:, 1:].to_numpy(dtype=float)
    assert np.isfinite(rates).all()
    assert ((rates >= 0) & (rates <= 1)).all()
    full = frame.loc[frame["Sample / survey wave"].eq("All survey waves")].iloc[0]
    assert float(full["Household weighted: conflict (%)"]) == 1.0
    assert float(full["Person weighted: conflict (%)"]) == 1.0
    spi_restricted = frame.loc[frame["Sample / survey wave"].eq("SPI-12 linked sample")].iloc[0]
    assert float(spi_restricted["Household weighted: SPI-12 (%)"]) == 1.0
    assert float(spi_restricted["Person weighted: SPI-12 (%)"]) == 1.0
    price_restricted = frame.loc[
        frame["Sample / survey wave"].eq("Wholesale 12m linked sample")
    ].iloc[0]
    assert float(price_restricted["Household weighted: wholesale 12m (%)"]) == 1.0
    assert float(price_restricted["Person weighted: wholesale 12m (%)"]) == 1.0

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Weighted Coverage"], workbook.sheetnames
    worksheet = workbook["Weighted Coverage"]
    assert worksheet.max_row == 13 and worksheet.max_column == 9
    assert worksheet.tables.keys() == {"SurveyWeightedCoverageTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"))


def main() -> None:
    household_columns = [
        YEAR,
        HOUSEHOLD_WEIGHT,
        CONFLICT,
        SPI12,
        WHOLESALE_12M,
        FLOOD_YEAR,
    ]
    person_columns = [
        YEAR,
        PERSON_WEIGHT,
        CONFLICT,
        SPI12,
        WHOLESALE_12M,
        FLOOD_YEAR,
    ]
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    people = pd.read_parquet(EDUCATION_PATH, columns=person_columns)

    observed_waves = sorted(int(value) for value in households[YEAR].dropna().unique())
    assert observed_waves == EXPECTED_WAVES, (observed_waves, EXPECTED_WAVES)
    assert len(households) == 62_920
    assert len(people) == 268_485

    frame = build_table(households, people)
    write_workbook(frame)
    validate_output(frame)

    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Weighted Coverage)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(frame.to_string(index=False))


if __name__ == "__main__":
    main()
