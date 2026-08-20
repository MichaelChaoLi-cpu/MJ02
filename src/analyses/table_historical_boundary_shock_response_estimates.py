#!/usr/bin/env python3
"""Historical Boundary Shock-Response Estimates.

Plan: Report twelve frozen annual land-NPP repression-by-rainfall specifications:
primary and mandatory confirmation models in natural and standardized units, annual-rainfall
alternative, five fixed bandwidths, river-distance adjustment, and NPP-quality adjustment.
Framework: AnaSOP Sections 5.3-5.4, 6.8-6.9, and the annual land-productivity workflow
in Section 7. Triangular-weight sensitivity is displayed in the companion figure.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from figure_historical_repression_and_contemporary_shock_sensitivity import (
    ALTERNATIVE_SHOCK,
    PRIMARY_OUTCOME,
    PRIMARY_SHOCK,
    QUALITY,
    RIVER,
    STANDARDIZED_OUTCOME,
    Estimate,
    fit_model,
    prepared_panel,
)


ROOT = Path(__file__).resolve().parents[2]
POWER = (
    ROOT
    / "data/exp/feasibility-check/historical-boundary-annual-spatial/annual_spatial_blinded_power.csv"
)
OUTPUT = (
    ROOT
    / "data/results/tables/Table_historical_boundary_shock_response_estimates.xlsx"
)

SESOI = 0.20
COLUMNS = [
    "Outcome",
    "Shock",
    "Sample",
    "Bandwidth (km)",
    "Interaction estimate",
    "Scale",
    "95% CI",
    "SESOI comparison",
    "Effective units",
    "Fixed effects",
    "Inference",
    "Interpretation",
]


@dataclass(frozen=True)
class SpecificationResult:
    natural: Estimate | None
    standardized: Estimate
    label: str
    sample_label: str
    fixed_effects: str
    effective_units: str


def power_lookup() -> dict[tuple[int, str], pd.Series]:
    power = pd.read_csv(POWER)
    strong = power.loc[power["dependence_scenario"].eq("strong clustered dependence")]
    return {
        (int(row.bandwidth_km), str(row.shock)): row
        for row in strong.itertuples(index=False)
        if row.shock in {"May-October rainfall anomaly", "Annual rainfall anomaly"}
    }


def ci_text(estimate: Estimate) -> str:
    precision = 4 if estimate.outcome == STANDARDIZED_OUTCOME else 6
    return f"[{estimate.ci_low:.{precision}f}, {estimate.ci_high:.{precision}f}]"


def sesoi_text(standardized: Estimate) -> tuple[str, str]:
    if standardized.ci_low >= -SESOI and standardized.ci_high <= SESOI:
        return (
            f"Paired z CI {ci_text(standardized)} inside ±0.20",
            "Substantively precise null",
        )
    if standardized.ci_low <= -SESOI and standardized.ci_high >= SESOI:
        return (
            f"Paired z CI {ci_text(standardized)} crosses both bounds",
            "Inconclusive relative to SESOI",
        )
    return (
        f"Paired z CI {ci_text(standardized)} not fully inside ±0.20",
        "Bounded but not equivalent",
    )


def effective_units_text(
    estimate: Estimate,
    power: dict[tuple[int, str], pd.Series],
    *,
    confirmation: bool = False,
) -> str:
    if confirmation:
        return (
            f"{estimate.villages} villages; {estimate.district_year_clusters} district-years; "
            "9 cross-side communes"
        )
    shock_label = (
        "Annual rainfall anomaly"
        if estimate.shock == ALTERNATIVE_SHOCK
        else "May-October rainfall anomaly"
    )
    row = power[(estimate.bandwidth_km, shock_label)]
    return (
        f"{estimate.villages} villages; {estimate.district_year_clusters} district-years; "
        f"{int(round(row.iid_equivalent_village_years)):,} design-equivalent village-years"
    )


def make_row(
    estimate: Estimate,
    standardized: Estimate,
    *,
    sample_label: str,
    fixed_effects: str,
    effective_units: str,
) -> dict[str, object]:
    sesoi_comparison, interpretation = sesoi_text(standardized)
    displayed_standardized = estimate.outcome == STANDARDIZED_OUTCOME
    if displayed_standardized:
        sesoi_comparison = (
            "95% CI inside ±0.20"
            if estimate.ci_low >= -SESOI and estimate.ci_high <= SESOI
            else sesoi_comparison
        )
    return {
        "Outcome": (
            "Annual land NPP anomaly (standardized)"
            if displayed_standardized
            else "Annual land NPP anomaly"
        ),
        "Shock": (
            "Annual rainfall anomaly (1 SD)"
            if estimate.shock == ALTERNATIVE_SHOCK
            else "May–October rainfall anomaly (1 SD)"
        ),
        "Sample": sample_label,
        "Bandwidth (km)": estimate.bandwidth_km,
        "Interaction estimate": estimate.estimate,
        "Scale": (
            "Outcome SD per 1-SD shock"
            if displayed_standardized
            else "kg C m⁻² per 1-SD shock"
        ),
        "95% CI": ci_text(estimate),
        "SESOI comparison": sesoi_comparison,
        "Effective units": effective_units,
        "Fixed effects": fixed_effects,
        "Inference": "Village + district-by-year two-way clustered; equal village-year weights",
        "Interpretation": interpretation,
    }


def build_table() -> pd.DataFrame:
    panel = prepared_panel()
    power = power_lookup()
    primary_natural = fit_model(panel, label="Primary 5 km")
    primary_z = fit_model(panel, label="Primary 5 km", outcome=STANDARDIZED_OUTCOME)
    confirmation_natural = fit_model(panel, label="Confirmation 5 km", confirmation=True)
    confirmation_z = fit_model(
        panel,
        label="Confirmation 5 km",
        outcome=STANDARDIZED_OUTCOME,
        confirmation=True,
    )
    annual_natural = fit_model(panel, label="Annual rainfall", shock=ALTERNATIVE_SHOCK)
    annual_z = fit_model(
        panel,
        label="Annual rainfall",
        outcome=STANDARDIZED_OUTCOME,
        shock=ALTERNATIVE_SHOCK,
    )
    river_natural = fit_model(panel, label="River-distance adjusted", modifier=RIVER)
    river_z = fit_model(
        panel,
        label="River-distance adjusted",
        outcome=STANDARDIZED_OUTCOME,
        modifier=RIVER,
    )
    quality_natural = fit_model(panel, label="NPP-quality adjusted", modifier=QUALITY)
    quality_z = fit_model(
        panel,
        label="NPP-quality adjusted",
        outcome=STANDARDIZED_OUTCOME,
        modifier=QUALITY,
    )
    bandwidth_models = {
        bandwidth: (
            fit_model(panel, label=f"{bandwidth} km", bandwidth_km=bandwidth),
            fit_model(
                panel,
                label=f"{bandwidth} km",
                outcome=STANDARDIZED_OUTCOME,
                bandwidth_km=bandwidth,
            ),
        )
        for bandwidth in [2, 10, 15, 20, 30]
    }

    primary_fe = "Village; boundary-segment-by-year"
    confirmation_fe = (
        "Village; boundary-segment-by-year; climate-commune-by-year"
    )
    primary_units = effective_units_text(primary_natural.estimate, power)
    confirmation_units = effective_units_text(
        confirmation_natural.estimate, power, confirmation=True
    )
    records = [
        make_row(
            primary_natural.estimate,
            primary_z.estimate,
            sample_label="All eligible villages, 2001–2021",
            fixed_effects=primary_fe,
            effective_units=primary_units,
        ),
        make_row(
            confirmation_natural.estimate,
            confirmation_z.estimate,
            sample_label="9 cross-side modern climate communes, 2001–2021",
            fixed_effects=confirmation_fe,
            effective_units=confirmation_units,
        ),
        make_row(
            primary_z.estimate,
            primary_z.estimate,
            sample_label="All eligible villages, 2001–2021",
            fixed_effects=primary_fe,
            effective_units=primary_units,
        ),
        make_row(
            confirmation_z.estimate,
            confirmation_z.estimate,
            sample_label="9 cross-side modern climate communes, 2001–2021",
            fixed_effects=confirmation_fe,
            effective_units=confirmation_units,
        ),
        make_row(
            annual_natural.estimate,
            annual_z.estimate,
            sample_label="All eligible villages, 2001–2021",
            fixed_effects=primary_fe,
            effective_units=effective_units_text(annual_natural.estimate, power),
        ),
    ]
    for bandwidth in [2, 10, 15, 20, 30]:
        natural, standardized = bandwidth_models[bandwidth]
        records.append(
            make_row(
                natural.estimate,
                standardized.estimate,
                sample_label=f"All eligible villages within {bandwidth} km, 2001–2021",
                fixed_effects=primary_fe,
                effective_units=effective_units_text(natural.estimate, power),
            )
        )
    records.extend(
        [
            make_row(
                river_natural.estimate,
                river_z.estimate,
                sample_label="All eligible villages; rainfall×river-distance hierarchy",
                fixed_effects=primary_fe,
                effective_units=primary_units,
            ),
            make_row(
                quality_natural.estimate,
                quality_z.estimate,
                sample_label="All eligible villages; NPP-quality interaction hierarchy",
                fixed_effects=primary_fe,
                effective_units=primary_units,
            ),
        ]
    )
    return pd.DataFrame.from_records(records, columns=COLUMNS)


def excel_value(value: object) -> object:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Shock Response"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="HistoricalBoundaryShockResponseTable", ref=f"A1:L{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    rule = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=rule)
    worksheet.row_dimensions[1].height = 40

    precise_fill = PatternFill("solid", fgColor="E2F0D9")
    for row_index in range(2, last_row + 1):
        for column_index in range(1, 13):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(size=8.3)
            cell.alignment = Alignment(
                horizontal="center" if column_index in {4, 5} else "left",
                vertical="center",
                wrap_text=True,
            )
        worksheet.cell(row=row_index, column=12).fill = precise_fill
        worksheet.row_dimensions[row_index].height = 43
        worksheet.cell(row=row_index, column=4).number_format = "0"
        worksheet.cell(row=row_index, column=5).number_format = (
            "0.0000"
            if worksheet.cell(row=row_index, column=6).value == "Outcome SD per 1-SD shock"
            else "0.000000"
        )

    widths = [32, 32, 44, 14, 19, 27, 24, 42, 43, 43, 47, 28]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:L{last_row}"
    worksheet.sheet_view.zoomScale = 70
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:L{last_row}"
    worksheet.page_margins.left = 0.15
    worksheet.page_margins.right = 0.15
    worksheet.page_margins.top = 0.18
    worksheet.page_margins.bottom = 0.18

    workbook.properties.title = "Historical Boundary Shock-Response Estimates"
    workbook.properties.subject = "Frozen annual land-NPP repression-by-rainfall estimates"
    workbook.properties.creator = "Mike Li"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (12, 12), frame.shape
    assert list(frame.columns) == COLUMNS
    primary = frame.iloc[0]
    confirmation = frame.iloc[1]
    primary_z = frame.iloc[2]
    confirmation_z = frame.iloc[3]
    # Validate the published precision while allowing harmless solver-level floating-point
    # variation across repeated fits and supported linear-algebra backends.
    assert np.isclose(primary["Interaction estimate"], 0.001091, rtol=0, atol=5e-7)
    assert np.isclose(confirmation["Interaction estimate"], 0.000414, rtol=0, atol=5e-7)
    assert np.isclose(primary_z["Interaction estimate"], 0.025749, rtol=0, atol=5e-7)
    assert np.isclose(confirmation_z["Interaction estimate"], 0.018640, rtol=0, atol=5e-7)
    assert frame["Interpretation"].eq("Substantively precise null").all()
    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == ["Shock Response"]
    worksheet = workbook["Shock Response"]
    assert worksheet.max_row == 13 and worksheet.max_column == 12
    assert set(worksheet.tables.keys()) == {"HistoricalBoundaryShockResponseTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?"))


def main() -> None:
    frame = build_table()
    write_workbook(frame)
    validate_output(frame)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Shock Response)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(
        frame[["Sample", "Bandwidth (km)", "Interaction estimate", "95% CI", "Interpretation"]]
        .to_string(index=False)
    )


if __name__ == "__main__":
    main()
