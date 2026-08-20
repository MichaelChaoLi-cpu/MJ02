#!/usr/bin/env python3
"""Nighttime Activity Independent Validation Estimates.

Plan: Report the twelve frozen observed-VIIRS repression-by-rainfall specifications:
the primary and mandatory within-modern-commune models, rainfall and outcome
alternatives, coverage and weighting checks, and fixed bandwidths.
Framework: AnaSOP Sections 5.5, 6.10, and the nighttime-activity independent-
validation workflow in Section 7.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from experiment_viirs_historical_boundary_shock_response import (
    prepare_panel,
    run_experiment,
)


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = (
    ROOT
    / "data/results/tables/Table_nighttime_activity_independent_validation_estimates.xlsx"
)

SESOI = 0.20
COLUMNS = [
    "Outcome",
    "Shock",
    "Sample",
    "Bandwidth (km)",
    "Interaction estimate",
    "Standardized scale",
    "95% CI",
    "SESOI comparison",
    "Effective units",
    "Fixed effects",
    "Inference",
    "Interpretation",
]

SPECIFICATION_ORDER = [
    "Primary 5 km",
    "Within-commune confirmation 5 km",
    "Annual rainfall alternative 5 km",
    "Annual median radiance 5 km",
    "Any nonzero radiance 5 km",
    "At least 40 cloud-free observations 5 km",
    "Triangular distance weights 5 km",
    "Fixed bandwidth 2 km",
    "Fixed bandwidth 10 km",
    "Fixed bandwidth 15 km",
    "Fixed bandwidth 20 km",
    "Fixed bandwidth 30 km",
]


def outcome_label(value: str) -> str:
    labels = {
        "Asinh Annual Mean Radiance": "Asinh annual mean radiance",
        "Asinh Annual Median Radiance": "Asinh annual median radiance",
        "Any Nonzero Annual Mean Radiance": "Any nonzero annual mean radiance",
    }
    return labels[value]


def shock_label(value: str) -> str:
    if value == "Annual Rainfall Anomaly Z (1991-2020)":
        return "Annual rainfall anomaly (1 SD)"
    return "May–October rainfall anomaly (1 SD)"


def sample_label(row: pd.Series) -> str:
    specification = str(row["specification"])
    if bool(row["confirmation_model"]):
        return "Confirmation: 10 cross-side climate communes, 2013–2021"
    if bool(row["quality_restricted"]):
        return "Grid-cell years with ≥40 cloud-free observations, 2013–2021"
    if specification == "Primary 5 km":
        return "Primary: all eligible grid cells, 2013–2021"
    return "All eligible grid cells, 2013–2021"


def fixed_effects_label(row: pd.Series) -> str:
    base = "Grid cell; boundary-segment-by-year"
    if bool(row["confirmation_model"]):
        return base + "; climate-commune-by-year"
    return base


def inference_label(row: pd.Series) -> str:
    weights = (
        "triangular distance weights"
        if bool(row["triangular_weights"])
        else "equal grid-cell-year weights"
    )
    return f"Grid cell + district-by-year two-way clustered; {weights}"


def effective_units_label(row: pd.Series) -> str:
    units = (
        f"{int(row['observations']):,} cell-years; "
        f"{int(row['grid_cells']):,} grid cells; "
        f"{int(row['district_year_clusters'])} district-years"
    )
    if bool(row["confirmation_model"]):
        units += f"; {int(row['cross_side_communes'])} cross-side communes"
    return units


def ci_label(row: pd.Series) -> str:
    return f"[{row['standardized_ci_low']:.4f}, {row['standardized_ci_high']:.4f}]"


def build_table() -> pd.DataFrame:
    estimates = run_experiment(prepare_panel())
    indexed = estimates.set_index("specification", drop=False)
    missing = [name for name in SPECIFICATION_ORDER if name not in indexed.index]
    if missing:
        raise ValueError(f"Missing frozen specifications: {missing}")
    ordered = indexed.loc[SPECIFICATION_ORDER].reset_index(drop=True)

    records: list[dict[str, object]] = []
    for _, row in ordered.iterrows():
        inside_sesoi = (
            row["standardized_ci_low"] >= -SESOI
            and row["standardized_ci_high"] <= SESOI
        )
        records.append(
            {
                "Outcome": outcome_label(str(row["outcome"])),
                "Shock": shock_label(str(row["shock"])),
                "Sample": sample_label(row),
                "Bandwidth (km)": int(row["bandwidth_km"]),
                "Interaction estimate": float(row["standardized_estimate"]),
                "Standardized scale": "Within-cell outcome SD per 1-SD shock",
                "95% CI": ci_label(row),
                "SESOI comparison": (
                    "Entire 95% CI inside ±0.20 SD"
                    if inside_sesoi
                    else "95% CI not fully inside ±0.20 SD"
                ),
                "Effective units": effective_units_label(row),
                "Fixed effects": fixed_effects_label(row),
                "Inference": inference_label(row),
                "Interpretation": (
                    "Substantively precise null"
                    if inside_sesoi
                    else "Not equivalent at the pre-specified SESOI"
                ),
            }
        )
    return pd.DataFrame.from_records(records, columns=COLUMNS)


def excel_value(value: object) -> object:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NTL Validation"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(
        displayName="NighttimeActivityValidationTable",
        ref=f"A1:L{last_row}",
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    bottom_rule = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=bottom_rule)
    worksheet.row_dimensions[1].height = 40

    precise_fill = PatternFill("solid", fgColor="E2F0D9")
    primary_fill = PatternFill("solid", fgColor="FCE4D6")
    confirmation_fill = PatternFill("solid", fgColor="E7E6E6")
    for row_index in range(2, last_row + 1):
        for column_index in range(1, 13):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(size=8.3)
            cell.alignment = Alignment(
                horizontal="center" if column_index in {4, 5} else "left",
                vertical="center",
                wrap_text=True,
            )
        worksheet.cell(row=row_index, column=12).fill = precise_fill
        worksheet.row_dimensions[row_index].height = 43
        worksheet.cell(row=row_index, column=4).number_format = "0"
        worksheet.cell(row=row_index, column=5).number_format = "0.0000"

    for column_index in range(1, 12):
        worksheet.cell(row=2, column=column_index).fill = primary_fill
        worksheet.cell(row=2, column=column_index).font = Font(size=8.3, bold=True)
        worksheet.cell(row=3, column=column_index).fill = confirmation_fill
    worksheet.cell(row=3, column=5).font = Font(size=8.3, bold=True)

    widths = [29, 31, 43, 14, 19, 29, 23, 31, 39, 39, 47, 28]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:L{last_row}"
    worksheet.sheet_view.zoomScale = 70
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:L{last_row}"
    worksheet.page_margins.left = 0.15
    worksheet.page_margins.right = 0.15
    worksheet.page_margins.top = 0.18
    worksheet.page_margins.bottom = 0.18

    workbook.properties.title = "Nighttime Activity Independent Validation Estimates"
    workbook.properties.subject = "Observed-VIIRS historical-boundary rainfall-response estimates"
    workbook.properties.creator = "Mike Li"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (12, 12), frame.shape
    assert list(frame.columns) == COLUMNS
    assert np.isclose(frame.iloc[0]["Interaction estimate"], 0.04105468, atol=1e-7)
    assert np.isclose(frame.iloc[1]["Interaction estimate"], 0.00278940, atol=1e-7)
    assert frame["Interpretation"].eq("Substantively precise null").all()
    assert frame["SESOI comparison"].eq("Entire 95% CI inside ±0.20 SD").all()

    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == ["NTL Validation"]
    worksheet = workbook["NTL Validation"]
    assert worksheet.max_row == 13 and worksheet.max_column == 12
    assert set(worksheet.tables.keys()) == {"NighttimeActivityValidationTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    frame = build_table()
    write_workbook(frame)
    validate_output(frame)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print("Workbook sheets: 1 (NTL Validation)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(
        frame[
            [
                "Outcome",
                "Shock",
                "Bandwidth (km)",
                "Interaction estimate",
                "95% CI",
                "Interpretation",
            ]
        ].to_string(index=False)
    )


if __name__ == "__main__":
    main()
