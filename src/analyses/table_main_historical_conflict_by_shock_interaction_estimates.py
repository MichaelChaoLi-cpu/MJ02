#!/usr/bin/env python3
"""Main Historical Conflict by Shock Interaction Estimates.

Plan: Report 24 central outcome-by-shock interaction estimates for drought,
extreme-wet rainfall, and local rice-price shocks across outcome domains.
Framework: AnaSOP Sections 5.1-5.2, 6.2-6.6, and the three central shock
workflow steps in Section 7. Estimates use outcome-specific samples, survey
weights, historical-geography and province-by-wave fixed effects, and
historical-geography clustered standard errors.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT
    / "data/results/tables/Table_main_historical_conflict_by_shock_interaction_estimates.xlsx"
)

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
HOUSEHOLD_PROVINCE = "Province Code Component"
EDUCATION_PROVINCE = "Province Code"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
EXTREME_WET = "Annual Rainfall Extreme Wet Shock"
PRICE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"

CROP_YIELD = "Crop Yield kg per ha"
LOSS_SHARE = "Post Harvest Loss Share"
CROP_VALUE = "Real 2021 Crop Production Value Riels"
FOOD_CONSUMPTION = "Real 2021 Food Consumption Value per Household Member Riels"
SEVERE_FOOD_INSECURITY = "Any Severe Food Insecurity Experience"
FOOD_INSECURITY_SEVERITY = "Food Insecurity Severity Sum"
ATTENDANCE = "Currently Attending School"
EDUCATION_EXPENDITURE = "Real 2021 Education Expenditure Riels"

SHOCK_LABELS = {
    SPI12: "SPI-12 (1 SD; higher = wetter)",
    EXTREME_WET: "Annual extreme wet (0→1)",
    PRICE_12M: "Local rice-price change (1 SD)",
}

COLUMNS = [
    "Domain",
    "Outcome (model scale)",
    "Contemporary shock",
    "Estimate",
    "95% CI",
    "N",
    "Fixed effects",
    "Controls",
    "Weights",
    "Uncertainty",
]


@dataclass(frozen=True)
class OutcomeSpec:
    domain: str
    outcome: str
    label: str
    source: str
    transform: str
    agriculture_only: bool = False
    school_age_only: bool = False


@dataclass(frozen=True)
class EstimateSpec:
    outcome: OutcomeSpec
    shock: str


OUTCOME_SPECS = [
    OutcomeSpec("Agriculture", CROP_YIELD, "Crop yield (asinh)", "household", "asinh", True),
    OutcomeSpec("Agriculture", LOSS_SHARE, "Post-harvest loss (percentage points)", "household", "percentage_points", True),
    OutcomeSpec("Agriculture", CROP_VALUE, "Crop production value (asinh, 2021 riels)", "household", "asinh", True),
    OutcomeSpec("Consumption", FOOD_CONSUMPTION, "Food consumption per member (asinh, 2021 riels)", "household", "asinh"),
    OutcomeSpec("Food security", SEVERE_FOOD_INSECURITY, "Severe food insecurity (percentage points)", "household", "percentage_points"),
    OutcomeSpec("Food security", FOOD_INSECURITY_SEVERITY, "Food insecurity severity (index points)", "household", "level"),
    OutcomeSpec("Education", ATTENDANCE, "School attendance, ages 6–17 (percentage points)", "education", "percentage_points", school_age_only=True),
    OutcomeSpec("Education", EDUCATION_EXPENDITURE, "Education expenditure (asinh, 2021 riels)", "education", "asinh"),
]
ESTIMATE_SPECS = [
    EstimateSpec(outcome, shock)
    for outcome in OUTCOME_SPECS
    for shock in [SPI12, EXTREME_WET, PRICE_12M]
]


@dataclass
class Estimate:
    spec: EstimateSpec
    coefficient: float
    standard_error: float
    sample_size: int
    cluster_count: int

    @property
    def lower(self) -> float:
        return self.coefficient - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.coefficient + 1.96 * self.standard_error


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def transform_outcome(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "asinh":
        if np.nanmin(values) < 0:
            raise ValueError("asinh outcome unexpectedly contains negative values")
        return np.arcsinh(values)
    if transform == "percentage_points":
        if np.nanmin(values) < 0 or np.nanmax(values) > 1:
            raise ValueError("Share or binary outcome falls outside [0, 1]")
        return 100.0 * values
    if transform == "level":
        return values
    raise ValueError(f"Unknown outcome transform: {transform}")


def prepare_sample(
    households: pd.DataFrame, education: pd.DataFrame, spec: EstimateSpec
) -> tuple[pd.DataFrame, str, str, list[str]]:
    outcome = spec.outcome
    if outcome.source == "household":
        source = households
        weight = HOUSEHOLD_WEIGHT
        province = HOUSEHOLD_PROVINCE
        controls = [HOUSEHOLD_SIZE]
    else:
        source = education
        weight = PERSON_WEIGHT
        province = EDUCATION_PROVINCE
        controls = [HOUSEHOLD_SIZE, AGE, FEMALE]

    required = [
        outcome.outcome,
        spec.shock,
        CONFLICT,
        weight,
        GEOGRAPHY,
        province,
        YEAR,
        *controls,
    ]
    if outcome.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    sample = source.loc[source[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    if outcome.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].astype(bool)].copy()
    if outcome.school_age_only:
        sample = sample.loc[sample[AGE].between(6, 17)].copy()
    if spec.shock == EXTREME_WET and not sample[spec.shock].isin([0, 1, False, True]).all():
        raise ValueError("Extreme-wet shock is not binary")
    if outcome.transform == "percentage_points" and not sample[outcome.outcome].between(0, 1).all():
        raise ValueError(f"Outcome is outside [0, 1]: {outcome.label}")
    if sample[GEOGRAPHY].nunique() < 50:
        raise ValueError(f"Too few clusters for {outcome.label} × {SHOCK_LABELS[spec.shock]}")
    return sample, weight, province, controls


def fit_estimate(
    households: pd.DataFrame, education: pd.DataFrame, spec: EstimateSpec
) -> Estimate:
    sample, weight_column, province_column, controls = prepare_sample(
        households, education, spec
    )
    weights = sample[weight_column].to_numpy(dtype=float)
    conflict_z = weighted_standardize(sample[CONFLICT].to_numpy(dtype=float), weights)
    raw_shock = sample[spec.shock].to_numpy(dtype=float)
    shock_model = (
        raw_shock
        if spec.shock == EXTREME_WET
        else weighted_standardize(raw_shock, weights)
    )
    outcome = transform_outcome(
        sample[spec.outcome.outcome].to_numpy(dtype=float), spec.outcome.transform
    )

    exogenous = pd.DataFrame(
        {
            "shock": shock_model,
            "conflict_x_shock": conflict_z * shock_model,
        },
        index=sample.index,
    )
    for control in controls:
        exogenous[control] = sample[control].to_numpy(dtype=float)
    province_wave = (
        sample[province_column].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {
            "geography": sample[GEOGRAPHY].astype("category"),
            "province_wave": province_wave.astype("category"),
        },
        index=sample.index,
    )
    fitted = AbsorbingLS(
        dependent=pd.Series(outcome, index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(weights, index=sample.index),
        drop_absorbed=True,
    ).fit(
        cov_type="clustered",
        clusters=sample[[GEOGRAPHY]],
        debiased=True,
    )
    if "conflict_x_shock" not in fitted.params.index:
        raise ValueError(
            f"Interaction absorbed for {spec.outcome.label} × {SHOCK_LABELS[spec.shock]}"
        )
    return Estimate(
        spec=spec,
        coefficient=float(fitted.params["conflict_x_shock"]),
        standard_error=float(fitted.std_errors["conflict_x_shock"]),
        sample_size=len(sample),
        cluster_count=int(sample[GEOGRAPHY].nunique()),
    )


def build_table(estimates: list[Estimate]) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for estimate in estimates:
        outcome = estimate.spec.outcome
        controls = (
            "Household size"
            if outcome.source == "household"
            else "Household size; age; female"
        )
        uncertainty = "Historical-geography clustered SE"
        records.append(
            {
                "Domain": outcome.domain,
                "Outcome (model scale)": outcome.label,
                "Contemporary shock": SHOCK_LABELS[estimate.spec.shock],
                "Estimate": estimate.coefficient,
                "95% CI": f"[{estimate.lower:.3f}, {estimate.upper:.3f}]",
                "N": estimate.sample_size,
                "Fixed effects": "Geography + province×wave",
                "Controls": controls,
                "Weights": "Household" if outcome.source == "household" else "Person",
                "Uncertainty": uncertainty,
            }
        )
    return pd.DataFrame.from_records(records, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main Estimates"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="MainInteractionEstimatesTable", ref=f"A1:J{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    white_bold = Font(color="FFFFFF", bold=True, size=8.5)
    header_rule = Side(style="thin", color="B7C9D6")
    section_rule = Side(style="medium", color="7F9DB9")
    column_rule = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=header_rule)
    worksheet.row_dimensions[1].height = 40

    prior_domain: str | None = None
    for row_index in range(2, last_row + 1):
        domain = str(worksheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 11):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(name="Aptos", size=8)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {4, 6} else "left",
                vertical="center",
                wrap_text=column_index not in {4, 6},
                indent=1 if column_index not in {4, 6} else 0,
            )
            if column_index in {4, 5, 6, 7, 10}:
                cell.border = Border(left=column_rule)
            if prior_domain is not None and domain != prior_domain:
                cell.border = Border(
                    top=section_rule,
                    left=column_rule if column_index in {4, 5, 6, 7, 10} else None,
                )
        worksheet.row_dimensions[row_index].height = 28
        prior_domain = domain

    for row_index in range(2, last_row + 1):
        worksheet.cell(row=row_index, column=4).number_format = "0.000"
        worksheet.cell(row=row_index, column=6).number_format = "#,##0"

    widths = [14, 32, 27, 12, 18, 10, 25, 21, 14, 34]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:J{last_row}"
    worksheet.sheet_view.zoomScale = 75
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:J{last_row}"
    worksheet.page_margins.left = 0.18
    worksheet.page_margins.right = 0.18
    worksheet.page_margins.top = 0.20
    worksheet.page_margins.bottom = 0.20

    workbook.properties.title = "Main Historical Conflict by Shock Interaction Estimates"
    workbook.properties.subject = "Direction 3 central weighted fixed-effects estimates"
    workbook.properties.creator = "Mike Li"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame, estimates: list[Estimate]) -> None:
    assert frame.shape == (24, 10), frame.shape
    assert list(frame.columns) == COLUMNS
    assert len(estimates) == 24
    assert frame[["Domain", "Outcome (model scale)", "Contemporary shock"]].duplicated().sum() == 0
    assert np.isfinite(frame["Estimate"].to_numpy(dtype=float)).all()
    assert frame["N"].gt(0).all()
    assert {estimate.spec.shock for estimate in estimates} == {SPI12, EXTREME_WET, PRICE_12M}

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Main Estimates"], workbook.sheetnames
    worksheet = workbook["Main Estimates"]
    assert worksheet.max_row == 25 and worksheet.max_column == 10
    assert worksheet.tables.keys() == {"MainInteractionEstimatesTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"))


def main() -> None:
    household_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            HOUSEHOLD_PROVINCE,
            HOUSEHOLD_WEIGHT,
            HOUSEHOLD_SIZE,
            AGRICULTURAL_HOUSEHOLD,
            CONFLICT,
            SPI12,
            EXTREME_WET,
            PRICE_12M,
            CROP_YIELD,
            LOSS_SHARE,
            CROP_VALUE,
            FOOD_CONSUMPTION,
            SEVERE_FOOD_INSECURITY,
            FOOD_INSECURITY_SEVERITY,
        }
    )
    education_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            EDUCATION_PROVINCE,
            PERSON_WEIGHT,
            HOUSEHOLD_SIZE,
            AGE,
            FEMALE,
            CONFLICT,
            SPI12,
            EXTREME_WET,
            PRICE_12M,
            ATTENDANCE,
            EDUCATION_EXPENDITURE,
        }
    )
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    assert len(households) == 62_920
    assert len(education) == 268_485

    estimates = [fit_estimate(households, education, spec) for spec in ESTIMATE_SPECS]
    frame = build_table(estimates)
    write_workbook(frame)
    validate_output(frame, estimates)

    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Main Estimates)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(frame[["Domain", "Outcome (model scale)", "Contemporary shock", "Estimate", "95% CI", "N"]].to_string(index=False))


if __name__ == "__main__":
    main()
