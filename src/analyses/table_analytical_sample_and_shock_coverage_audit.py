#!/usr/bin/env python3
"""Analytical Sample and Shock Coverage Audit.

Plan: Document the estimable sample and unequal temporal/geographic coverage.
Framework: Supports AnaSOP Sections 5.1, 6.5, and the first analysis block
in Section 7 by reporting shock-specific denominators before estimation.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT / "data/results/tables/Table_analytical_sample_and_shock_coverage_audit.xlsx"
)

YEAR = "Survey Year"
PSU = "PSU"
PROVINCE = "Province Code Component"
HH_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
ANNUAL_WET = "Annual Rainfall Extreme Wet Shock"
SPI12 = "Interview Month SPI 12 Month"
WHOLESALE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"
RETAIL_12M = "12 Month Change in Broad Retail Food Local Relative Log Price"
FLOOD_YEAR = "Survey Year Maximum Flooded Geography Share"
FLOOD_12M = "Preceding 12 Month Maximum Flooded Geography Share"
EXPECTED_WAVES = [2007, 2009, 2011, 2013, 2014, 2016, 2017, 2019, 2021]

PRIMARY_COLUMNS = [
    "Sample / survey wave",
    "Household observations",
    "Person observations",
    "PSUs",
    "Provinces",
    "Conflict linked (HH %)",
    "SPI-12 linked (HH %)",
    "Wholesale 12m shock linked (HH %)",
    "Satellite flood linked (HH %)",
]


def coverage(frame: pd.DataFrame, variable: str) -> pd.Series:
    """Return nonmissing exposure coverage without converting missing shocks to zero."""
    return frame[variable].notna()


def share(indicator: pd.Series) -> float:
    return np.nan if indicator.empty else float(indicator.mean())


def weighted_share(indicator: pd.Series, weights: pd.Series) -> float:
    valid = indicator.notna() & weights.notna() & (weights > 0)
    if not valid.any():
        return np.nan
    return float(np.average(indicator.loc[valid].astype(float), weights=weights.loc[valid]))


def row_metrics(label: str, hh: pd.DataFrame, people: pd.DataFrame) -> dict[str, object]:
    return {
        "Sample / survey wave": label,
        "Household observations": len(hh),
        "Person observations": len(people),
        "PSUs": hh[PSU].nunique(dropna=True),
        "Provinces": hh[PROVINCE].nunique(dropna=True),
        "Conflict linked (HH %)": share(coverage(hh, CONFLICT)),
        "SPI-12 linked (HH %)": share(coverage(hh, SPI12)),
        "Wholesale 12m shock linked (HH %)": share(coverage(hh, WHOLESALE_12M)),
        "Satellite flood linked (HH %)": share(coverage(hh, FLOOD_YEAR)),
    }


def weighted_row(label: str, hh: pd.DataFrame, people: pd.DataFrame) -> dict[str, object]:
    return {
        "Sample / survey wave": label,
        "Conflict linked (HH weighted %)": weighted_share(coverage(hh, CONFLICT), hh[HH_WEIGHT]),
        "SPI-12 linked (HH weighted %)": weighted_share(coverage(hh, SPI12), hh[HH_WEIGHT]),
        "Wholesale 12m linked (HH weighted %)": weighted_share(
            coverage(hh, WHOLESALE_12M), hh[HH_WEIGHT]
        ),
        "Satellite flood linked (HH weighted %)": weighted_share(
            coverage(hh, FLOOD_YEAR), hh[HH_WEIGHT]
        ),
        "Conflict linked (person weighted %)": weighted_share(
            coverage(people, CONFLICT), people[PERSON_WEIGHT]
        ),
        "SPI-12 linked (person weighted %)": weighted_share(
            coverage(people, SPI12), people[PERSON_WEIGHT]
        ),
        "Wholesale 12m linked (person weighted %)": weighted_share(
            coverage(people, WHOLESALE_12M), people[PERSON_WEIGHT]
        ),
        "Satellite flood linked (person weighted %)": weighted_share(
            coverage(people, FLOOD_YEAR), people[PERSON_WEIGHT]
        ),
    }


def build_sample_rows(
    hh: pd.DataFrame, people: pd.DataFrame
) -> list[tuple[str, pd.DataFrame, pd.DataFrame]]:
    rows: list[tuple[str, pd.DataFrame, pd.DataFrame]] = []
    for year in EXPECTED_WAVES:
        rows.append(
            (
                str(year),
                hh.loc[hh[YEAR].eq(year)],
                people.loc[people[YEAR].eq(year)],
            )
        )
    rows.extend(
        [
            ("All survey waves", hh, people),
            (
                "SPI-12 linked sample",
                hh.loc[coverage(hh, SPI12)],
                people.loc[coverage(people, SPI12)],
            ),
            (
                "Wholesale 12m linked sample",
                hh.loc[coverage(hh, WHOLESALE_12M)],
                people.loc[coverage(people, WHOLESALE_12M)],
            ),
        ]
    )
    return rows


def build_shock_samples(hh: pd.DataFrame, people: pd.DataFrame) -> pd.DataFrame:
    shocks = [
        ("Historical conflict", CONFLICT, "Predetermined exposure"),
        ("Annual extreme-wet rainfall", ANNUAL_WET, "Primary weather shock"),
        ("Interview-month SPI-12", SPI12, "Primary drought shock"),
        ("Wholesale rice price, 12m change", WHOLESALE_12M, "Primary price shock"),
        ("Broad retail food price, 12m change", RETAIL_12M, "Price robustness"),
        ("Survey-year satellite inundation", FLOOD_YEAR, "Secondary validation"),
        ("Preceding-12m satellite inundation", FLOOD_12M, "Secondary validation"),
    ]
    records: list[dict[str, object]] = []
    for label, variable, role in shocks:
        hh_mask = coverage(hh, variable)
        person_mask = coverage(people, variable)
        hh_supported = hh.loc[hh_mask]
        wave_values = sorted(int(value) for value in hh_supported[YEAR].dropna().unique())
        records.append(
            {
                "Exposure / shock": label,
                "Analytical role": role,
                "Household N": int(hh_mask.sum()),
                "Household coverage (%)": share(hh_mask),
                "Person N": int(person_mask.sum()),
                "Person coverage (%)": share(person_mask),
                "PSUs": hh_supported[PSU].nunique(dropna=True),
                "Provinces": hh_supported[PROVINCE].nunique(dropna=True),
                "Survey waves with support": ", ".join(map(str, wave_values)),
            }
        )
    return pd.DataFrame.from_records(records)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_frame(ws, frame: pd.DataFrame, header_row: int) -> tuple[int, int]:
    for col_idx, column in enumerate(frame.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=column)
    for row_idx, values in enumerate(
        frame.itertuples(index=False, name=None), start=header_row + 1
    ):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=excel_value(value))
    return header_row, header_row + len(frame)


def add_excel_table(ws, name: str, header_row: int, last_row: int) -> None:
    table = Table(displayName=name, ref=f"A{header_row}:I{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_color_scale(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.75,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )


def write_integrated_workbook(
    path: Path,
    primary: pd.DataFrame,
    weighted: pd.DataFrame,
    shock_samples: pd.DataFrame,
    definitions: pd.DataFrame,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Integrated Audit"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    title_fill = PatternFill("solid", fgColor="17365D")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    band_fill = PatternFill("solid", fgColor="5B9BD5")
    band_font = Font(color="FFFFFF", bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    restricted_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_gray = Side(style="thin", color="B7C9D6")

    ws.merge_cells("A1:I1")
    ws["A1"] = "Analytical Sample and Shock Coverage Audit"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:I2")
    ws["A2"] = (
        "Direction 3: historical conflict and contemporary climate, food-price, "
        "and satellite-inundation exposure support"
    )
    ws["A2"].font = Font(color="44546A", italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    cursor = 4
    ranges: dict[str, tuple[int, int]] = {}
    sections = [
        ("A. Survey-wave sample and unweighted linkage coverage", primary, "CoverageAuditTable"),
        ("B. Shock-specific analytical support", shock_samples, "ShockSamplesTable"),
        ("C. Survey-weighted linkage coverage", weighted, "WeightedCoverageTable"),
    ]
    for title, frame, table_name in sections:
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=9)
        cell = ws.cell(row=cursor, column=1, value=title)
        cell.fill = band_fill
        cell.font = band_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cursor].height = 22

        header_row, last_row = write_frame(ws, frame, cursor + 1)
        ranges[table_name] = (header_row, last_row)
        add_excel_table(ws, table_name, header_row, last_row)
        ws.row_dimensions[header_row].height = 34
        for header_cell in ws[header_row]:
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            header_cell.border = Border(bottom=thin_gray)
        cursor = last_row + 2

    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=9)
    cell = ws.cell(row=cursor, column=1, value="D. Definitions, sources, and interpretation boundary")
    cell.fill = band_fill
    cell.font = band_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cursor].height = 22

    definition_header = cursor + 1
    ws.cell(row=definition_header, column=1, value="Item")
    ws.merge_cells(
        start_row=definition_header, start_column=2, end_row=definition_header, end_column=9
    )
    ws.cell(row=definition_header, column=2, value="Definition / rule")
    for header_cell in ws[definition_header]:
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_cell.border = Border(bottom=thin_gray)
    ws.row_dimensions[definition_header].height = 24

    for row_idx, values in enumerate(
        definitions.itertuples(index=False, name=None), start=definition_header + 1
    ):
        ws.cell(row=row_idx, column=1, value=excel_value(values[0]))
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=9)
        ws.cell(row=row_idx, column=2, value=excel_value(values[1]))
        ws.cell(row=row_idx, column=1).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws.cell(row=row_idx, column=2).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws.row_dimensions[row_idx].height = 28

    primary_header, primary_last = ranges["CoverageAuditTable"]
    shock_header, shock_last = ranges["ShockSamplesTable"]
    weighted_header, weighted_last = ranges["WeightedCoverageTable"]

    for header_row, last_row in [
        (primary_header, primary_last),
        (weighted_header, weighted_last),
    ]:
        for row_idx in range(header_row + 1, last_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if label == "All survey waves":
                for data_cell in ws[row_idx]:
                    data_cell.fill = total_fill
                    data_cell.font = Font(bold=True)
            elif label in {"SPI-12 linked sample", "Wholesale 12m linked sample"}:
                for data_cell in ws[row_idx]:
                    data_cell.fill = restricted_fill

    for row in ws.iter_rows(
        min_row=primary_header + 1, max_row=primary_last, min_col=2, max_col=5
    ):
        for data_cell in row:
            data_cell.number_format = "#,##0"
    for row in ws.iter_rows(
        min_row=primary_header + 1, max_row=primary_last, min_col=6, max_col=9
    ):
        for data_cell in row:
            data_cell.number_format = "0.0%"
    for row in ws.iter_rows(
        min_row=weighted_header + 1, max_row=weighted_last, min_col=2, max_col=9
    ):
        for data_cell in row:
            data_cell.number_format = "0.0%"
    for row in ws.iter_rows(
        min_row=shock_header + 1, max_row=shock_last, min_col=3, max_col=8
    ):
        for data_cell in row:
            data_cell.number_format = "0.0%" if data_cell.column in {4, 6} else "#,##0"

    add_color_scale(ws, f"F{primary_header + 1}:I{primary_last}")
    add_color_scale(ws, f"B{weighted_header + 1}:I{weighted_last}")
    add_color_scale(ws, f"D{shock_header + 1}:D{shock_last}")
    add_color_scale(ws, f"F{shock_header + 1}:F{shock_last}")

    for row_idx in list(range(primary_header + 1, primary_last + 1)) + list(
        range(weighted_header + 1, weighted_last + 1)
    ):
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left" if col_idx == 1 else "right", vertical="center"
            )
    for row_idx in range(shock_header + 1, shock_last + 1):
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left" if col_idx in {1, 2, 9} else "right",
                vertical="center",
                wrap_text=col_idx in {1, 2, 9},
            )
        ws.row_dimensions[row_idx].height = 22

    widths = [36, 31, 22, 25, 20, 25, 25, 34, 48]
    for col_idx, width in enumerate(widths, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 70
    ws.freeze_panes = "A6"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = "8"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:I{ws.max_row}"
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.oddFooter.center.text = "Table 1 | Direction 3 analytical coverage audit"
    ws.oddFooter.center.size = 8

    wb.properties.title = "Analytical Sample and Shock Coverage Audit"
    wb.properties.subject = "Direction 3 sample and exposure support"
    wb.properties.creator = "Mike Li"
    wb.save(path)


def main() -> None:
    hh_columns = [
        YEAR,
        PSU,
        PROVINCE,
        HH_WEIGHT,
        CONFLICT,
        ANNUAL_WET,
        SPI12,
        WHOLESALE_12M,
        RETAIL_12M,
        FLOOD_YEAR,
        FLOOD_12M,
    ]
    person_columns = [
        YEAR,
        PERSON_WEIGHT,
        CONFLICT,
        ANNUAL_WET,
        SPI12,
        WHOLESALE_12M,
        RETAIL_12M,
        FLOOD_YEAR,
        FLOOD_12M,
    ]
    hh = pd.read_parquet(HOUSEHOLD_PATH, columns=hh_columns)
    people = pd.read_parquet(EDUCATION_PATH, columns=person_columns)

    observed_waves = sorted(int(value) for value in hh[YEAR].dropna().unique())
    assert observed_waves == EXPECTED_WAVES, (observed_waves, EXPECTED_WAVES)
    assert len(hh) == 62_920
    assert len(people) == 268_485

    sample_rows = build_sample_rows(hh, people)
    primary = pd.DataFrame(
        [row_metrics(label, hh_i, people_i) for label, hh_i, people_i in sample_rows]
    )
    weighted = pd.DataFrame(
        [weighted_row(label, hh_i, people_i) for label, hh_i, people_i in sample_rows]
    )
    shock_samples = build_shock_samples(hh, people)
    definitions = pd.DataFrame(
        [
            (
                "Primary table dimensions",
                "The main audit block contains exactly 12 data rows and 9 columns, as specified in AnaSOP Section 8.",
            ),
            (
                "Survey-wave rows",
                "One row for each georeferenced survey wave: 2007, 2009, 2011, 2013, 2014, 2016, 2017, 2019, and 2021.",
            ),
            (
                "All survey waves",
                "All released household-wave and person-wave observations before outcome-specific complete-case restrictions.",
            ),
            (
                "SPI-12 linked sample",
                "Records with a nonmissing interview-month 12-month Standardized Precipitation Index.",
            ),
            (
                "Wholesale 12m linked sample",
                "Records with a nonmissing exact 12-month change in local relative log wholesale rice price.",
            ),
            (
                "Conflict linked",
                "Nonmissing log bombing unique-location density per 100 square kilometres; valid mapped structural zeros count as linked.",
            ),
            (
                "Satellite flood linked",
                "Nonmissing survey-year maximum flooded-geography share; post-2018 noncoverage remains missing and is not recoded as no flood.",
            ),
            (
                "Unweighted percentages",
                "Share of household observations in the displayed row with a nonmissing exposure.",
            ),
            (
                "Weighted percentages",
                "Coverage share using positive, nonmissing household or person survey weights within the displayed row.",
            ),
            (
                "PSUs and provinces",
                "Distinct PSUs and province codes among household observations in the displayed row.",
            ),
            ("Household source", str(HOUSEHOLD_PATH.relative_to(ROOT))),
            ("Person source", str(EDUCATION_PATH.relative_to(ROOT))),
            (
                "Interpretation boundary",
                "Coverage describes estimability and does not establish causal identification or exposure validity.",
            ),
        ],
        columns=["Item", "Definition / rule"],
    )

    assert primary.shape == (12, 9), primary.shape
    assert list(primary.columns) == PRIMARY_COLUMNS
    assert weighted.shape == (12, 9), weighted.shape

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    write_integrated_workbook(OUTPUT_PATH, primary, weighted, shock_samples, definitions)

    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Integrated Audit)")
    print(f"Primary table: {primary.shape[0]} rows x {primary.shape[1]} columns")
    print(f"Household observations: {len(hh):,}")
    print(f"Person observations: {len(people):,}")


if __name__ == "__main__":
    main()
