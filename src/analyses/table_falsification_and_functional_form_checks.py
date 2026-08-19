#!/usr/bin/env python3
"""Falsification and Functional Form Checks.

Plan: Report 20 future-shock placebo, nonlinear binary/count, and alternative
fixed-effect diagnostics for representative conflict-conditioned shock models.
Framework: AnaSOP Sections 5.1-5.2, 6.2-6.7, and the robustness workflow step
in Section 7. Future shocks are exact 12-month leads for monthly SPI and price
measures and one-calendar-year leads for annual extreme-wet rainfall.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy.stats import norm


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
CLIMATE_MONTH_PATH = (
    ROOT / "data/processed/chirps_long_baseline_commune_month_preprocessed.parquet"
)
CLIMATE_YEAR_PATH = (
    ROOT / "data/processed/chirps_long_baseline_commune_year_preprocessed.parquet"
)
PRICE_PATH = (
    ROOT / "data/processed/wfp_wholesale_rice_province_month_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT / "data/results/tables/Table_falsification_and_functional_form_checks.xlsx"
)

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
HOUSEHOLD_PROVINCE = "Province Code Component"
EDUCATION_PROVINCE = "Province Code"
MATCHED_PROVINCE = "Matched Climate Province Name"
PRICE_DATE = "Price Exposure Date"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
EXTREME_WET = "Annual Rainfall Extreme Wet Shock"
PRICE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"
FUTURE_SPI12 = "Future 12 Month SPI 12 Month"
FUTURE_EXTREME_WET = "Future Annual Extreme Wet Shock"
FUTURE_PRICE_12M = "Future 12 Month Rice Price Change"

CROP_YIELD = "Crop Yield kg per ha"
FOOD_CONSUMPTION = "Real 2021 Food Consumption Value per Household Member Riels"
FOOD_ITEMS = "Food Items with Positive Consumption Count"
SEVERE_FOOD_INSECURITY = "Any Severe Food Insecurity Experience"
FOOD_INSECURITY_SEVERITY = "Food Insecurity Severity Sum"
ATTENDANCE = "Currently Attending School"
EDUCATION_EXPENDITURE = "Real 2021 Education Expenditure Riels"

SHOCK_LABELS = {
    SPI12: "SPI-12 (1 SD; higher = wetter)",
    EXTREME_WET: "Annual extreme wet (0→1)",
    PRICE_12M: "Local rice-price change (1 SD)",
}
FUTURE_SHOCKS = {
    SPI12: FUTURE_SPI12,
    EXTREME_WET: FUTURE_EXTREME_WET,
    PRICE_12M: FUTURE_PRICE_12M,
}

COLUMNS = [
    "Check family",
    "Outcome and shock",
    "Core estimate",
    "Alternative estimate",
    "Alternative 95% CI or p-value",
    "N",
    "Estimator",
    "Fixed effects",
    "Comparison rule",
    "Diagnostic conclusion",
]


@dataclass(frozen=True)
class OutcomeSpec:
    domain: str
    variable: str
    label: str
    source: str
    linear_transform: str
    agriculture_only: bool = False
    school_age_only: bool = False


@dataclass(frozen=True)
class ModelSpec:
    outcome: OutcomeSpec
    shock: str


@dataclass(frozen=True)
class LinearResult:
    estimate: float
    standard_error: float
    p_value: float
    sample_size: int

    @property
    def lower(self) -> float:
        return self.estimate - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.estimate + 1.96 * self.standard_error


@dataclass(frozen=True)
class NonlinearResult:
    average_cross_partial: float
    interaction_p_value: float
    sample_size: int
    iterations: int


CROP_YIELD_SPEC = OutcomeSpec(
    "Agriculture", CROP_YIELD, "Crop yield (asinh)", "household", "asinh", True
)
FOOD_CONSUMPTION_SPEC = OutcomeSpec(
    "Consumption",
    FOOD_CONSUMPTION,
    "Food consumption per member (asinh, 2021 riels)",
    "household",
    "asinh",
)
FOOD_ITEMS_SPEC = OutcomeSpec(
    "Consumption", FOOD_ITEMS, "Food items with positive consumption (count)", "household", "level"
)
SEVERE_FOOD_SPEC = OutcomeSpec(
    "Food security",
    SEVERE_FOOD_INSECURITY,
    "Severe food insecurity (percentage points)",
    "household",
    "percentage_points",
)
FOOD_SEVERITY_SPEC = OutcomeSpec(
    "Food security",
    FOOD_INSECURITY_SEVERITY,
    "Food insecurity severity (index points)",
    "household",
    "level",
)
ATTENDANCE_SPEC = OutcomeSpec(
    "Education",
    ATTENDANCE,
    "School attendance, ages 6–17 (percentage points)",
    "education",
    "percentage_points",
    school_age_only=True,
)
EDUCATION_EXPENDITURE_SPEC = OutcomeSpec(
    "Education",
    EDUCATION_EXPENDITURE,
    "Education expenditure (asinh, 2021 riels)",
    "education",
    "asinh",
)

PLACEBO_SPECS = [
    ModelSpec(CROP_YIELD_SPEC, SPI12),
    ModelSpec(CROP_YIELD_SPEC, EXTREME_WET),
    ModelSpec(CROP_YIELD_SPEC, PRICE_12M),
    ModelSpec(FOOD_CONSUMPTION_SPEC, SPI12),
    ModelSpec(FOOD_CONSUMPTION_SPEC, EXTREME_WET),
    ModelSpec(FOOD_CONSUMPTION_SPEC, PRICE_12M),
    ModelSpec(ATTENDANCE_SPEC, PRICE_12M),
    ModelSpec(EDUCATION_EXPENDITURE_SPEC, PRICE_12M),
]

NONLINEAR_SPECS = [
    (ModelSpec(SEVERE_FOOD_SPEC, SPI12), "logit"),
    (ModelSpec(SEVERE_FOOD_SPEC, PRICE_12M), "logit"),
    (ModelSpec(ATTENDANCE_SPEC, PRICE_12M), "logit"),
    (ModelSpec(FOOD_SEVERITY_SPEC, SPI12), "poisson"),
    (ModelSpec(FOOD_SEVERITY_SPEC, PRICE_12M), "poisson"),
    (ModelSpec(FOOD_ITEMS_SPEC, SPI12), "poisson"),
]

FIXED_EFFECT_REPRESENTATIVES = [
    ModelSpec(CROP_YIELD_SPEC, SPI12),
    ModelSpec(FOOD_CONSUMPTION_SPEC, EXTREME_WET),
    ModelSpec(ATTENDANCE_SPEC, PRICE_12M),
]


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def transform_outcome(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "asinh":
        if np.nanmin(values) < 0:
            raise ValueError("asinh outcome unexpectedly contains negative values")
        return np.arcsinh(values)
    if transform == "percentage_points":
        if np.nanmin(values) < 0 or np.nanmax(values) > 1:
            raise ValueError("Share or binary outcome falls outside [0, 1]")
        return 100.0 * values
    if transform == "level":
        return values
    raise ValueError(f"Unknown outcome transform: {transform}")


def attach_future_shocks(data: pd.DataFrame) -> pd.DataFrame:
    result = data.copy()
    result[GEOGRAPHY] = result[GEOGRAPHY].astype(str).str.zfill(6)
    exposure_date = pd.to_datetime(result[PRICE_DATE], errors="coerce")
    future_month_start = (
        exposure_date.dt.to_period("M").dt.to_timestamp() + pd.DateOffset(months=12)
    )
    future_price_date = exposure_date + pd.DateOffset(months=12)

    climate_month = pd.read_parquet(
        CLIMATE_MONTH_PATH, columns=[GEOGRAPHY, "Date", "SPI 12 Month"]
    )
    climate_month[GEOGRAPHY] = climate_month[GEOGRAPHY].astype(str).str.zfill(6)
    climate_month["Date"] = pd.to_datetime(climate_month["Date"])
    climate_month = climate_month.drop_duplicates([GEOGRAPHY, "Date"])
    spi_lookup = climate_month.set_index([GEOGRAPHY, "Date"])["SPI 12 Month"]
    spi_keys = pd.MultiIndex.from_arrays(
        [result[GEOGRAPHY].to_numpy(), future_month_start.to_numpy()],
        names=[GEOGRAPHY, "Date"],
    )
    result[FUTURE_SPI12] = spi_lookup.reindex(spi_keys).to_numpy()

    climate_year = pd.read_parquet(
        CLIMATE_YEAR_PATH,
        columns=[GEOGRAPHY, "Year", "Annual Rainfall Extreme Wet Shock"],
    )
    climate_year[GEOGRAPHY] = climate_year[GEOGRAPHY].astype(str).str.zfill(6)
    climate_year = climate_year.drop_duplicates([GEOGRAPHY, "Year"])
    wet_lookup = climate_year.set_index([GEOGRAPHY, "Year"])[
        "Annual Rainfall Extreme Wet Shock"
    ]
    wet_keys = pd.MultiIndex.from_arrays(
        [result[GEOGRAPHY].to_numpy(), (result[YEAR] + 1).to_numpy()],
        names=[GEOGRAPHY, "Year"],
    )
    result[FUTURE_EXTREME_WET] = wet_lookup.reindex(wet_keys).to_numpy()

    prices = pd.read_parquet(
        PRICE_PATH,
        columns=[
            "Province Name",
            "Date",
            "12 Month Change in Local Relative Log Wholesale Rice Price",
        ],
    )
    prices["Date"] = pd.to_datetime(prices["Date"])
    prices = prices.drop_duplicates(["Province Name", "Date"])
    price_lookup = prices.set_index(["Province Name", "Date"])[
        "12 Month Change in Local Relative Log Wholesale Rice Price"
    ]
    price_keys = pd.MultiIndex.from_arrays(
        [result[MATCHED_PROVINCE].to_numpy(), future_price_date.to_numpy()],
        names=["Province Name", "Date"],
    )
    result[FUTURE_PRICE_12M] = price_lookup.reindex(price_keys).to_numpy()
    return result


def prepare_sample(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: ModelSpec,
    required_shocks: list[str] | None = None,
) -> tuple[pd.DataFrame, str, str, list[str]]:
    outcome = spec.outcome
    if outcome.source == "household":
        source = households
        weight = HOUSEHOLD_WEIGHT
        province = HOUSEHOLD_PROVINCE
        controls = [HOUSEHOLD_SIZE]
    else:
        source = education
        weight = PERSON_WEIGHT
        province = EDUCATION_PROVINCE
        controls = [HOUSEHOLD_SIZE, AGE, FEMALE]
    shocks = required_shocks or [spec.shock]
    required = [
        outcome.variable,
        *shocks,
        CONFLICT,
        weight,
        GEOGRAPHY,
        province,
        YEAR,
        *controls,
    ]
    if outcome.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    required = list(dict.fromkeys(required))
    sample = source.loc[source[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    if outcome.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)].copy()
    if outcome.school_age_only:
        sample = sample.loc[sample[AGE].between(6, 17)].copy()
    sample[GEOGRAPHY] = sample[GEOGRAPHY].astype(str).str.zfill(6)
    if sample[GEOGRAPHY].nunique() < 50:
        raise ValueError(f"Too few geography clusters for {outcome.label}")
    return sample, weight, province, controls


def model_arrays(
    sample: pd.DataFrame,
    spec: ModelSpec,
    shock_column: str,
    weight_column: str,
    controls: list[str],
) -> tuple[pd.DataFrame, np.ndarray, np.ndarray, np.ndarray]:
    weights = sample[weight_column].to_numpy(dtype=float)
    conflict_z = weighted_standardize(sample[CONFLICT].to_numpy(dtype=float), weights)
    raw_shock = sample[shock_column].to_numpy(dtype=float)
    shock_model = (
        raw_shock
        if shock_column in {EXTREME_WET, FUTURE_EXTREME_WET}
        else weighted_standardize(raw_shock, weights)
    )
    exogenous = pd.DataFrame(
        {
            "shock": shock_model,
            "conflict_x_shock": conflict_z * shock_model,
        },
        index=sample.index,
    )
    for control in controls:
        exogenous[control] = sample[control].to_numpy(dtype=float)
    outcome = transform_outcome(
        sample[spec.outcome.variable].to_numpy(dtype=float),
        spec.outcome.linear_transform,
    )
    return exogenous, outcome, conflict_z, shock_model


def fixed_effect_frame(
    sample: pd.DataFrame,
    province_column: str,
    mode: str,
) -> pd.DataFrame:
    province_wave = (
        sample[province_column].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    frame = {"geography": sample[GEOGRAPHY].astype("category")}
    if mode == "core":
        frame["province_wave"] = province_wave.astype("category")
    elif mode in {"wave", "province_trend"}:
        frame["survey_wave"] = sample[YEAR].astype(str).astype("category")
    else:
        raise ValueError(f"Unknown fixed-effect mode: {mode}")
    return pd.DataFrame(frame, index=sample.index)


def fit_linear_on_sample(
    sample: pd.DataFrame,
    spec: ModelSpec,
    shock_column: str,
    weight_column: str,
    province_column: str,
    controls: list[str],
    fixed_effect_mode: str = "core",
) -> LinearResult:
    exogenous, outcome, _, _ = model_arrays(
        sample, spec, shock_column, weight_column, controls
    )
    if fixed_effect_mode == "province_trend":
        centered_year = sample[YEAR].to_numpy(dtype=float) - float(sample[YEAR].mean())
        province_dummies = pd.get_dummies(
            sample[province_column].astype(str), prefix="province_trend", drop_first=True, dtype=float
        )
        province_dummies = province_dummies.mul(centered_year, axis=0)
        province_dummies.index = sample.index
        exogenous = pd.concat([exogenous, province_dummies], axis=1)
    absorbed = fixed_effect_frame(sample, province_column, fixed_effect_mode)
    fitted = AbsorbingLS(
        dependent=pd.Series(outcome, index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(sample[weight_column].to_numpy(dtype=float), index=sample.index),
        drop_absorbed=True,
    ).fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample[GEOGRAPHY]}, index=sample.index),
        debiased=True,
    )
    target = "conflict_x_shock"
    if target not in fitted.params.index:
        raise ValueError("Target interaction was absorbed")
    return LinearResult(
        estimate=float(fitted.params[target]),
        standard_error=float(fitted.std_errors[target]),
        p_value=float(fitted.pvalues[target]),
        sample_size=len(sample),
    )


def clustered_nonlinear_pvalue(
    model: AbsorbingLS,
    sample: pd.DataFrame,
    survey_weights: np.ndarray,
    outcome: np.ndarray,
    mean: np.ndarray,
    target: str,
) -> float:
    x_weighted = model.absorbed_exog.to_numpy(dtype=float)
    model_weights = model.weights.ndarray.reshape(-1)
    x_residualized = x_weighted / np.sqrt(model_weights)[:, None]
    working_variance = np.maximum(mean * (1.0 - mean), 1e-8)
    if np.nanmax(outcome) > 1:
        working_variance = np.maximum(mean, 1e-8)
    working_weight = survey_weights * working_variance
    scale = float(np.mean(working_weight))
    score = x_residualized * (
        survey_weights * (outcome - mean) / scale
    )[:, None]
    cluster_codes, inverse = np.unique(
        sample[GEOGRAPHY].astype(str).to_numpy(), return_inverse=True
    )
    cluster_scores = np.zeros((len(cluster_codes), score.shape[1]), dtype=float)
    np.add.at(cluster_scores, inverse, score)
    bread = np.linalg.inv(x_weighted.T @ x_weighted)
    adjustment = (len(cluster_codes) / (len(cluster_codes) - 1.0)) * (
        (len(sample) - 1.0) / (len(sample) - x_weighted.shape[1])
    )
    covariance = adjustment * bread @ (cluster_scores.T @ cluster_scores) @ bread
    target_index = list(model.absorbed_exog.columns).index(target)
    standard_error = float(np.sqrt(covariance[target_index, target_index]))
    coefficient = float(np.linalg.lstsq(x_weighted, model.absorbed_dependent.to_numpy(), rcond=None)[0][target_index, 0])
    return float(2.0 * norm.sf(abs(coefficient / standard_error)))


def fit_nonlinear_on_sample(
    sample: pd.DataFrame,
    spec: ModelSpec,
    family: str,
    weight_column: str,
    province_column: str,
    controls: list[str],
    max_iterations: int = 120,
) -> NonlinearResult:
    exogenous, _, conflict_z, shock_model = model_arrays(
        sample, spec, spec.shock, weight_column, controls
    )
    outcome = sample[spec.outcome.variable].to_numpy(dtype=float)
    if family == "logit" and not pd.Series(outcome).isin([0, 1]).all():
        raise ValueError("Logit outcome is not binary")
    if family == "poisson" and np.nanmin(outcome) < 0:
        raise ValueError("Poisson outcome contains negative values")
    survey_weights = sample[weight_column].to_numpy(dtype=float)
    survey_weights = survey_weights / float(np.mean(survey_weights))
    absorbed = fixed_effect_frame(sample, province_column, "core")

    weighted_mean = float(np.average(outcome, weights=survey_weights))
    if family == "logit":
        probability = np.clip(weighted_mean, 1e-4, 1 - 1e-4)
        eta = np.full(len(sample), np.log(probability / (1.0 - probability)))
    else:
        eta = np.full(len(sample), np.log(max(weighted_mean, 1e-4)))

    previous_parameters: np.ndarray | None = None
    converged = False
    model: AbsorbingLS | None = None
    fitted = None
    for iteration in range(1, max_iterations + 1):
        if family == "logit":
            mean = 1.0 / (1.0 + np.exp(-np.clip(eta, -20, 20)))
            variance = np.maximum(mean * (1.0 - mean), 1e-8)
        else:
            mean = np.exp(np.clip(eta, -15, 15))
            variance = np.maximum(mean, 1e-8)
        pseudo_outcome = eta + (outcome - mean) / variance
        working_weight = survey_weights * variance
        model = AbsorbingLS(
            dependent=pd.Series(pseudo_outcome, index=sample.index, name="pseudo_outcome"),
            exog=exogenous,
            absorb=absorbed,
            weights=pd.Series(working_weight, index=sample.index),
            drop_absorbed=True,
        )
        fitted = model.fit(cov_type="unadjusted")
        parameters = fitted.params.to_numpy(dtype=float)
        eta_candidate = pseudo_outcome - fitted.resids.to_numpy(dtype=float)
        eta = np.clip(
            eta_candidate,
            -20 if family == "logit" else -15,
            20 if family == "logit" else 15,
        )
        if previous_parameters is not None and np.max(
            np.abs(parameters - previous_parameters)
        ) < 1e-8:
            converged = True
            break
        previous_parameters = parameters
    if not converged or model is None or fitted is None:
        raise ValueError(f"{family} HDFE IRLS did not converge")

    if family == "logit":
        mean = 1.0 / (1.0 + np.exp(-eta))
    else:
        mean = np.exp(eta)
    target = "conflict_x_shock"
    beta = float(fitted.params[target])
    theta = float(fitted.params["shock"])
    if family == "logit":
        first = mean * (1.0 - mean)
        second = first * (1.0 - 2.0 * mean)
        cross_partial = first * beta + second * (
            beta * shock_model
        ) * (theta + beta * conflict_z)
        average_cross_partial = 100.0 * float(
            np.average(cross_partial, weights=survey_weights)
        )
    else:
        cross_partial = mean * (
            beta + (beta * shock_model) * (theta + beta * conflict_z)
        )
        average_cross_partial = float(
            np.average(cross_partial, weights=survey_weights)
        )
    interaction_p = clustered_nonlinear_pvalue(
        model,
        sample,
        survey_weights,
        outcome,
        mean,
        target,
    )
    return NonlinearResult(
        average_cross_partial=average_cross_partial,
        interaction_p_value=interaction_p,
        sample_size=len(sample),
        iterations=iteration,
    )


def ci_text(result: LinearResult) -> str:
    return f"[{result.lower:.3f}, {result.upper:.3f}]"


def precision_conclusion(result: LinearResult, prefix: str) -> str:
    if result.lower <= 0 <= result.upper:
        return f"{prefix}: alternative CI includes zero"
    return f"{prefix}: alternative CI excludes zero"


def placebo_rows(
    households: pd.DataFrame,
    education: pd.DataFrame,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for spec in PLACEBO_SPECS:
        future_shock = FUTURE_SHOCKS[spec.shock]
        sample, weight, province, controls = prepare_sample(
            households,
            education,
            spec,
            [spec.shock, future_shock],
        )
        core = fit_linear_on_sample(sample, spec, spec.shock, weight, province, controls)
        placebo = fit_linear_on_sample(
            sample, spec, future_shock, weight, province, controls
        )
        rows.append(
            {
                "Check family": "Future-shock placebo",
                "Outcome and shock": f"{spec.outcome.label} × {SHOCK_LABELS[spec.shock]}",
                "Core estimate": core.estimate,
                "Alternative estimate": placebo.estimate,
                "Alternative 95% CI or p-value": ci_text(placebo),
                "N": len(sample),
                "Estimator": "Survey-weighted linear model; one-year lead shock",
                "Fixed effects": "Geography + province × wave",
                "Comparison rule": "Future shock should not predict the current outcome",
                "Diagnostic conclusion": precision_conclusion(placebo, "Pass" if placebo.lower <= 0 <= placebo.upper else "Flag"),
            }
        )
    return rows


def nonlinear_rows(
    households: pd.DataFrame,
    education: pd.DataFrame,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for spec, family in NONLINEAR_SPECS:
        sample, weight, province, controls = prepare_sample(
            households, education, spec
        )
        core = fit_linear_on_sample(sample, spec, spec.shock, weight, province, controls)
        try:
            alternative = fit_nonlinear_on_sample(
                sample, spec, family, weight, province, controls
            )
            same_sign = np.sign(core.estimate) == np.sign(alternative.average_cross_partial)
            conclusion = (
                "Sign retained; nonlinear interaction p<0.05"
                if same_sign and alternative.interaction_p_value < 0.05
                else "Sign retained; nonlinear interaction imprecise"
                if same_sign
                else "Sign differs under nonlinear model"
            )
            alternative_estimate: float | None = alternative.average_cross_partial
            inference = f"Interaction p={alternative.interaction_p_value:.3f}"
            estimator = (
                "HDFE logit; average cross-partial (percentage points)"
                if family == "logit"
                else "HDFE PPML; average cross-partial (outcome units)"
            )
            sample_size = alternative.sample_size
        except Exception as error:
            alternative_estimate = None
            inference = "Not estimable"
            estimator = "HDFE logit" if family == "logit" else "HDFE PPML"
            conclusion = f"Not estimable: {error}"
            sample_size = len(sample)
        rows.append(
            {
                "Check family": "Nonlinear functional form",
                "Outcome and shock": f"{spec.outcome.label} × {SHOCK_LABELS[spec.shock]}",
                "Core estimate": core.estimate,
                "Alternative estimate": alternative_estimate,
                "Alternative 95% CI or p-value": inference,
                "N": sample_size,
                "Estimator": estimator,
                "Fixed effects": "Geography + province × wave",
                "Comparison rule": "Compare sign and nonlinear interaction inference",
                "Diagnostic conclusion": conclusion,
            }
        )
    return rows


def fixed_effect_rows(
    households: pd.DataFrame,
    education: pd.DataFrame,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    alternatives = [
        ("wave", "Geography + survey wave"),
        ("province_trend", "Geography + survey wave + province trends"),
    ]
    for mode, fixed_effect_label in alternatives:
        for spec in FIXED_EFFECT_REPRESENTATIVES:
            sample, weight, province, controls = prepare_sample(
                households, education, spec
            )
            core = fit_linear_on_sample(
                sample, spec, spec.shock, weight, province, controls, "core"
            )
            alternative = fit_linear_on_sample(
                sample, spec, spec.shock, weight, province, controls, mode
            )
            same_sign = np.sign(core.estimate) == np.sign(alternative.estimate)
            rows.append(
                {
                    "Check family": "Alternative fixed effects",
                    "Outcome and shock": f"{spec.outcome.label} × {SHOCK_LABELS[spec.shock]}",
                    "Core estimate": core.estimate,
                    "Alternative estimate": alternative.estimate,
                    "Alternative 95% CI or p-value": ci_text(alternative),
                    "N": len(sample),
                    "Estimator": "Survey-weighted linear model",
                    "Fixed effects": fixed_effect_label,
                    "Comparison rule": "Alternative structure must preserve identifying variation",
                    "Diagnostic conclusion": precision_conclusion(
                        alternative, "Sign retained" if same_sign else "Sign changed"
                    ),
                }
            )
    return rows


def build_table(
    households: pd.DataFrame,
    education: pd.DataFrame,
) -> pd.DataFrame:
    rows = placebo_rows(households, education)
    print(f"Completed future-shock placebos: {len(rows)}/8")
    nonlinear = nonlinear_rows(households, education)
    rows.extend(nonlinear)
    print(f"Completed nonlinear checks: {len(nonlinear)}/6")
    fixed_effects = fixed_effect_rows(households, education)
    rows.extend(fixed_effects)
    print(f"Completed alternative fixed-effect checks: {len(fixed_effects)}/6")
    return pd.DataFrame(rows, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Falsification and Form"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="FalsificationFunctionalFormTable", ref=f"A1:J{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    family_fills = {
        "Future-shock placebo": PatternFill("solid", fgColor="EAF2F8"),
        "Nonlinear functional form": PatternFill("solid", fgColor="E2F0D9"),
        "Alternative fixed effects": PatternFill("solid", fgColor="FFF2CC"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=8.5)
    navy_bold = Font(color="1F4E78", bold=True, size=8.5)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")
    section_start_rows = {2, 10, 16}

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 52

    for row_index in range(2, last_row + 1):
        family = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 11):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {3, 4, 6} else "left",
                vertical="center",
                wrap_text=column_index not in {3, 4, 6},
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = family_fills[family]
        sheet.cell(row=row_index, column=1).font = navy_bold
        sheet.row_dimensions[row_index].height = 48

    for column in ("C", "D"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.0000"
    for cell in sheet["F"][1:]:
        cell.number_format = "#,##0"

    widths = [28, 49, 16, 18, 28, 13, 45, 38, 45, 43]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:J{last_row}"
    sheet.sheet_view.zoomScale = 62
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:J{last_row}"
    sheet.page_margins.left = 0.10
    sheet.page_margins.right = 0.10
    sheet.page_margins.top = 0.14
    sheet.page_margins.bottom = 0.14

    workbook.properties.title = "Falsification and Functional Form Checks"
    workbook.properties.subject = "Future-shock placebo, nonlinear-model, and alternative-fixed-effect diagnostics"
    workbook.properties.creator = "Mike Li"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (20, 10), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["N"].gt(0).all()
    assert np.isfinite(frame["Core estimate"]).all()
    estimable = ~frame["Diagnostic conclusion"].str.startswith("Not estimable")
    assert np.isfinite(frame.loc[estimable, "Alternative estimate"]).all()
    assert frame.groupby("Check family").size().to_dict() == {
        "Alternative fixed effects": 6,
        "Future-shock placebo": 8,
        "Nonlinear functional form": 6,
    }

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Falsification and Form"]
    sheet = workbook["Falsification and Form"]
    assert sheet.max_row == 21
    assert sheet.max_column == 10
    assert list(sheet.tables) == ["FalsificationFunctionalFormTable"]
    assert sheet.tables["FalsificationFunctionalFormTable"].ref == "A1:J21"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    outcomes = {
        spec.outcome
        for spec in [*PLACEBO_SPECS, *(item[0] for item in NONLINEAR_SPECS), *FIXED_EFFECT_REPRESENTATIVES]
    }
    shocks = {spec.shock for spec in PLACEBO_SPECS}
    shocks.update(spec.shock for spec, _ in NONLINEAR_SPECS)
    shocks.update(spec.shock for spec in FIXED_EFFECT_REPRESENTATIVES)
    household_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            HOUSEHOLD_PROVINCE,
            MATCHED_PROVINCE,
            PRICE_DATE,
            HOUSEHOLD_WEIGHT,
            HOUSEHOLD_SIZE,
            AGRICULTURAL_HOUSEHOLD,
            CONFLICT,
            *shocks,
            *[outcome.variable for outcome in outcomes if outcome.source == "household"],
        }
    )
    education_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            EDUCATION_PROVINCE,
            MATCHED_PROVINCE,
            PRICE_DATE,
            PERSON_WEIGHT,
            HOUSEHOLD_SIZE,
            AGE,
            FEMALE,
            CONFLICT,
            *shocks,
            *[outcome.variable for outcome in outcomes if outcome.source == "education"],
        }
    )
    households = attach_future_shocks(
        pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    )
    education = attach_future_shocks(
        pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    )
    table = build_table(households, education)
    write_workbook(table)
    validate_output(table)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {table.shape[0]} rows x {table.shape[1]} columns")
    print("Not estimable rows:", int(table["Diagnostic conclusion"].str.startswith("Not estimable").sum()))


if __name__ == "__main__":
    main()
