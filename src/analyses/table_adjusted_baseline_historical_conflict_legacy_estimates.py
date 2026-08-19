#!/usr/bin/env python3
"""Adjusted Baseline Historical Conflict Legacy Estimates.

Plan: Report 18 conditional associations between standardized historical
conflict exposure and current outcomes, mechanisms, and demographic profiles.
Framework: AnaSOP Sections 5.1, 6.1, 6.3, 6.5-6.6, and the baseline-legacy
workflow step in Section 7. Models absorb province-by-wave fixed effects, use
survey weights, and cluster uncertainty by linked historical geography. They
are descriptive legacy associations rather than causal bombing effects.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT
    / "data/results/tables/Table_adjusted_baseline_historical_conflict_legacy_estimates.xlsx"
)

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
HOUSEHOLD_PROVINCE = "Province Code Component"
EDUCATION_PROVINCE = "Province Code"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
ATTENDANCE_ELIGIBLE = "School Attendance Outcome Eligible"
CONFLICT = "Log Bombing Unique Locations per 100 km2"

COLUMNS = [
    "Domain",
    "Outcome (model scale)",
    "Adjusted conflict coefficient",
    "95% CI",
    "N",
    "Fixed effects",
    "Controls",
    "Weights",
    "Interpretation boundary",
]


@dataclass(frozen=True)
class OutcomeSpec:
    domain: str
    variable: str
    label: str
    source: str
    transform: str
    agriculture_only: bool = False
    attendance_only: bool = False


OUTCOMES = [
    OutcomeSpec("Agriculture", "Cultivated Crop Area m2", "Cultivated crop area (asinh m²)", "household", "asinh", True),
    OutcomeSpec("Agriculture", "Crop Production Quantity kg", "Crop production quantity (asinh kg)", "household", "asinh", True),
    OutcomeSpec("Agriculture", "Crop Yield kg per ha", "Crop yield (asinh kg/ha)", "household", "asinh", True),
    OutcomeSpec("Agriculture", "Post Harvest Loss Share", "Post-harvest loss (percentage points)", "household", "percentage_points", True),
    OutcomeSpec("Agriculture", "Crop Diversity Count", "Crop diversity (count)", "household", "level", True),
    OutcomeSpec("Agriculture", "Irrigable Parcel Share", "Irrigable parcel share (percentage points)", "household", "percentage_points", True),
    OutcomeSpec("Agriculture", "Real 2021 Crop Production Value Riels", "Crop production value (asinh, 2021 riels)", "household", "asinh", True),
    OutcomeSpec("Agriculture", "Real 2021 Agricultural Input Cost Riels", "Agricultural input cost (asinh, 2021 riels)", "household", "asinh", True),
    OutcomeSpec("Consumption and food security", "Real 2021 Food Consumption Value per Household Member Riels", "Food consumption per member (asinh, 2021 riels)", "household", "asinh"),
    OutcomeSpec("Consumption and food security", "Food Items with Positive Consumption Count", "Food items with positive consumption (count)", "household", "level"),
    OutcomeSpec("Consumption and food security", "Any Severe Food Insecurity Experience", "Severe food insecurity (percentage points)", "household", "percentage_points"),
    OutcomeSpec("Consumption and food security", "Food Insecurity Severity Sum", "Food insecurity severity (index points)", "household", "level"),
    OutcomeSpec("Education", "Currently Attending School", "School attendance, ages 6–17 (percentage points)", "education", "percentage_points", attendance_only=True),
    OutcomeSpec("Education", "Years Attended School", "Years attended school", "education", "level"),
    OutcomeSpec("Education", "Real 2021 Education Expenditure Riels", "Education expenditure (asinh, 2021 riels)", "education", "asinh"),
    OutcomeSpec("Demographic profile", HOUSEHOLD_SIZE, "Household size", "household", "level"),
    OutcomeSpec("Demographic profile", AGE, "Age (years)", "education", "level"),
    OutcomeSpec("Demographic profile", FEMALE, "Female (percentage points)", "education", "percentage_points"),
]


@dataclass(frozen=True)
class Estimate:
    coefficient: float
    standard_error: float
    sample_size: int
    cluster_count: int
    controls: tuple[str, ...]

    @property
    def lower(self) -> float:
        return self.coefficient - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.coefficient + 1.96 * self.standard_error


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize conflict exposure with invalid variance")
    return (values - mean) / standard_deviation


def transform_outcome(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "asinh":
        if np.nanmin(values) < 0:
            raise ValueError("asinh outcome unexpectedly contains negative values")
        return np.arcsinh(values)
    if transform == "percentage_points":
        if np.nanmin(values) < 0 or np.nanmax(values) > 1:
            raise ValueError("Share or binary outcome falls outside [0, 1]")
        return 100.0 * values
    if transform == "level":
        return values
    raise ValueError(f"Unknown outcome transform: {transform}")


def prepare_sample(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: OutcomeSpec,
) -> tuple[pd.DataFrame, str, str, list[str]]:
    if spec.source == "household":
        source = households
        weight = HOUSEHOLD_WEIGHT
        province = HOUSEHOLD_PROVINCE
        candidate_controls = [HOUSEHOLD_SIZE]
    else:
        source = education
        weight = PERSON_WEIGHT
        province = EDUCATION_PROVINCE
        candidate_controls = [HOUSEHOLD_SIZE, AGE, FEMALE]
    controls = [control for control in candidate_controls if control != spec.variable]

    required = [
        spec.variable,
        CONFLICT,
        weight,
        GEOGRAPHY,
        province,
        YEAR,
        *controls,
    ]
    if spec.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    if spec.attendance_only:
        required.extend([AGE, ATTENDANCE_ELIGIBLE])
    required = list(dict.fromkeys(required))

    sample = source.loc[source[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    if spec.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)].copy()
    if spec.attendance_only:
        sample = sample.loc[
            sample[AGE].between(6, 17) & sample[ATTENDANCE_ELIGIBLE].eq(1)
        ].copy()
    if sample[GEOGRAPHY].nunique() < 50:
        raise ValueError(f"Too few geography clusters for {spec.label}")
    return sample, weight, province, controls


def fit_estimate(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: OutcomeSpec,
) -> Estimate:
    sample, weight_column, province_column, controls = prepare_sample(
        households, education, spec
    )
    weights = sample[weight_column].to_numpy(dtype=float)
    conflict_z = weighted_standardize(sample[CONFLICT].to_numpy(dtype=float), weights)
    outcome = transform_outcome(
        sample[spec.variable].to_numpy(dtype=float), spec.transform
    )

    exogenous = pd.DataFrame({"conflict_z": conflict_z}, index=sample.index)
    for control in controls:
        exogenous[control] = sample[control].to_numpy(dtype=float)
    province_wave = (
        sample[province_column].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {"province_wave": province_wave.astype("category")}, index=sample.index
    )
    fitted = AbsorbingLS(
        dependent=pd.Series(outcome, index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(weights, index=sample.index),
        drop_absorbed=True,
    ).fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample[GEOGRAPHY]}, index=sample.index),
        debiased=True,
    )
    if "conflict_z" not in fitted.params.index:
        raise ValueError(f"Conflict exposure was absorbed for {spec.label}")
    return Estimate(
        coefficient=float(fitted.params["conflict_z"]),
        standard_error=float(fitted.std_errors["conflict_z"]),
        sample_size=len(sample),
        cluster_count=int(sample[GEOGRAPHY].nunique()),
        controls=tuple(controls),
    )


def control_label(controls: tuple[str, ...]) -> str:
    labels = {
        HOUSEHOLD_SIZE: "household size",
        AGE: "age",
        FEMALE: "female",
    }
    return "; ".join(labels[control] for control in controls) or "None; outcome is the demographic control"


def build_table(households: pd.DataFrame, education: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for spec in OUTCOMES:
        result = fit_estimate(households, education, spec)
        rows.append(
            {
                "Domain": spec.domain,
                "Outcome (model scale)": spec.label,
                "Adjusted conflict coefficient": result.coefficient,
                "95% CI": f"[{result.lower:.3f}, {result.upper:.3f}]",
                "N": result.sample_size,
                "Fixed effects": "Province × survey wave",
                "Controls": control_label(result.controls),
                "Weights": "Household survey weight" if spec.source == "household" else "Person survey weight",
                "Interpretation boundary": "Conditional legacy association; not a causal effect of historical conflict",
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Adjusted Baseline Legacy"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="AdjustedBaselineLegacyTable", ref=f"A1:I{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    domain_fills = {
        "Agriculture": PatternFill("solid", fgColor="EAF2F8"),
        "Consumption and food security": PatternFill("solid", fgColor="E2F0D9"),
        "Education": PatternFill("solid", fgColor="FFF2CC"),
        "Demographic profile": PatternFill("solid", fgColor="E4DFEC"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=9)
    navy_bold = Font(color="1F4E78", bold=True, size=9)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")
    section_start_rows = {2, 10, 14, 17}

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 46

    for row_index in range(2, last_row + 1):
        domain = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 10):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {3, 5} else "left",
                vertical="center",
                wrap_text=column_index not in {3, 5},
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = domain_fills[domain]
        sheet.cell(row=row_index, column=1).font = navy_bold
        sheet.row_dimensions[row_index].height = 42

    for cell in sheet["C"][1:]:
        cell.number_format = "0.000"
    for cell in sheet["E"][1:]:
        cell.number_format = "#,##0"

    widths = [27, 47, 22, 20, 13, 29, 38, 27, 49]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:I{last_row}"
    sheet.sheet_view.zoomScale = 70
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:I{last_row}"
    sheet.page_margins.left = 0.12
    sheet.page_margins.right = 0.12
    sheet.page_margins.top = 0.16
    sheet.page_margins.bottom = 0.16

    workbook.properties.title = "Adjusted Baseline Historical Conflict Legacy Estimates"
    workbook.properties.subject = "Conditional legacy associations under the AnaSOP Section 6.1 model"
    workbook.properties.creator = "Mike Li"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (18, 9), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["N"].gt(0).all()
    assert np.isfinite(frame["Adjusted conflict coefficient"]).all()
    assert frame["95% CI"].str.match(r"^\[-?\d+\.\d{3}, -?\d+\.\d{3}\]$").all()
    assert frame.groupby("Domain").size().to_dict() == {
        "Agriculture": 8,
        "Consumption and food security": 4,
        "Demographic profile": 3,
        "Education": 3,
    }

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Adjusted Baseline Legacy"]
    sheet = workbook["Adjusted Baseline Legacy"]
    assert sheet.max_row == 19
    assert sheet.max_column == 9
    assert list(sheet.tables) == ["AdjustedBaselineLegacyTable"]
    assert sheet.tables["AdjustedBaselineLegacyTable"].ref == "A1:I19"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    household_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            HOUSEHOLD_PROVINCE,
            HOUSEHOLD_WEIGHT,
            HOUSEHOLD_SIZE,
            AGRICULTURAL_HOUSEHOLD,
            CONFLICT,
            *[spec.variable for spec in OUTCOMES if spec.source == "household"],
        }
    )
    education_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            EDUCATION_PROVINCE,
            PERSON_WEIGHT,
            HOUSEHOLD_SIZE,
            AGE,
            FEMALE,
            ATTENDANCE_ELIGIBLE,
            CONFLICT,
            *[spec.variable for spec in OUTCOMES if spec.source == "education"],
        }
    )
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    table = build_table(households, education)
    write_workbook(table)
    validate_output(table)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {table.shape[0]} rows x {table.shape[1]} columns")
    print(table.groupby("Domain").size().to_string())


if __name__ == "__main__":
    main()
