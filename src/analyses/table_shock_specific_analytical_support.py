#!/usr/bin/env python3
"""Generate the Shock-Specific Analytical Support table.

This table reports the estimable household and person samples for each
historical-conflict or contemporary-shock exposure in Direction 3. Missing
shock values remain missing and are never recoded as zero.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = ROOT / "data/results/tables/Table_shock_specific_analytical_support.xlsx"

YEAR = "Survey Year"
PSU = "PSU"
PROVINCE = "Province Code Component"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
ANNUAL_WET = "Annual Rainfall Extreme Wet Shock"
SPI12 = "Interview Month SPI 12 Month"
WHOLESALE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"
RETAIL_12M = "12 Month Change in Broad Retail Food Local Relative Log Price"
FLOOD_YEAR = "Survey Year Maximum Flooded Geography Share"
FLOOD_12M = "Preceding 12 Month Maximum Flooded Geography Share"
EXPECTED_WAVES = [2007, 2009, 2011, 2013, 2014, 2016, 2017, 2019, 2021]

COLUMNS = [
    "Exposure / shock",
    "Analytical role",
    "Household N",
    "Household coverage (%)",
    "Person N",
    "Person coverage (%)",
    "PSUs",
    "Provinces",
    "Survey waves with support",
]


def coverage(frame: pd.DataFrame, variable: str) -> pd.Series:
    """Return a nonmissing exposure indicator without imputing structural zeros."""
    return frame[variable].notna()


def share(indicator: pd.Series) -> float:
    return np.nan if indicator.empty else float(indicator.mean())


def build_table(households: pd.DataFrame, people: pd.DataFrame) -> pd.DataFrame:
    shocks = [
        ("Historical conflict", CONFLICT, "Predetermined exposure"),
        ("Annual extreme-wet rainfall", ANNUAL_WET, "Primary weather shock"),
        ("Interview-month SPI-12", SPI12, "Primary drought shock"),
        ("Wholesale rice price, 12m change", WHOLESALE_12M, "Primary price shock"),
        ("Broad retail food price, 12m change", RETAIL_12M, "Price robustness"),
        ("Survey-year satellite inundation", FLOOD_YEAR, "Secondary validation"),
        ("Preceding-12m satellite inundation", FLOOD_12M, "Secondary validation"),
    ]
    records: list[dict[str, object]] = []
    for label, variable, role in shocks:
        household_mask = coverage(households, variable)
        person_mask = coverage(people, variable)
        supported_households = households.loc[household_mask]
        supported_waves = sorted(
            int(value) for value in supported_households[YEAR].dropna().unique()
        )
        records.append(
            {
                "Exposure / shock": label,
                "Analytical role": role,
                "Household N": int(household_mask.sum()),
                "Household coverage (%)": share(household_mask),
                "Person N": int(person_mask.sum()),
                "Person coverage (%)": share(person_mask),
                "PSUs": int(supported_households[PSU].nunique(dropna=True)),
                "Provinces": int(supported_households[PROVINCE].nunique(dropna=True)),
                "Survey waves with support": ", ".join(map(str, supported_waves)),
            }
        )
    return pd.DataFrame.from_records(records, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Shock Support"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="ShockSpecificSupportTable", ref=f"A1:I{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    white_bold = Font(color="FFFFFF", bold=True, size=9)
    light_rule = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=light_rule)
    worksheet.row_dimensions[1].height = 40

    for row_index in range(2, last_row + 1):
        for column_index in range(1, 10):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="left" if column_index in {1, 2, 9} else "right",
                vertical="center",
                wrap_text=column_index in {1, 2, 9},
            )
        worksheet.row_dimensions[row_index].height = 30

    for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=8):
        for cell in row:
            cell.number_format = "0.0%" if cell.column in {4, 6} else "#,##0"

    color_scale = ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color="F8696B",
        mid_type="num",
        mid_value=0.75,
        mid_color="FFEB84",
        end_type="num",
        end_value=1,
        end_color="63BE7B",
    )
    worksheet.conditional_formatting.add(f"D2:D{last_row}", color_scale)
    worksheet.conditional_formatting.add(f"F2:F{last_row}", color_scale)

    widths = [34, 24, 16, 21, 16, 19, 12, 13, 47]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:I{last_row}"
    worksheet.sheet_view.zoomScale = 80
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:I{last_row}"
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.25
    worksheet.page_margins.bottom = 0.25
    worksheet.oddFooter.center.text = "Shock-Specific Analytical Support | Direction 3"
    worksheet.oddFooter.center.size = 8

    workbook.properties.title = "Shock-Specific Analytical Support"
    workbook.properties.subject = "Direction 3 exposure support and estimable samples"
    workbook.properties.creator = "Mike Li"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    expected_counts = {
        "Historical conflict": (62_920, 268_485, 1_220, 25),
        "Annual extreme-wet rainfall": (62_920, 268_485, 1_220, 25),
        "Interview-month SPI-12": (52_845, 226_177, 1_215, 25),
        "Wholesale rice price, 12m change": (31_493, 134_308, 974, 22),
        "Broad retail food price, 12m change": (11_310, 46_715, 869, 22),
        "Survey-year satellite inundation": (42_765, 184_718, 1_054, 24),
        "Preceding-12m satellite inundation": (39_172, 168_929, 1_044, 24),
    }
    assert frame.shape == (7, 9), frame.shape
    assert list(frame.columns) == COLUMNS
    for label, expected in expected_counts.items():
        row = frame.loc[frame["Exposure / shock"].eq(label)].iloc[0]
        observed = tuple(int(row[column]) for column in ["Household N", "Person N", "PSUs", "Provinces"])
        assert observed == expected, (label, observed, expected)

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Shock Support"], workbook.sheetnames
    worksheet = workbook["Shock Support"]
    assert worksheet.max_row == 8 and worksheet.max_column == 9
    assert worksheet.tables.keys() == {"ShockSpecificSupportTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?"))


def main() -> None:
    household_columns = [YEAR, PSU, PROVINCE, CONFLICT, ANNUAL_WET, SPI12, WHOLESALE_12M, RETAIL_12M, FLOOD_YEAR, FLOOD_12M]
    person_columns = [YEAR, CONFLICT, ANNUAL_WET, SPI12, WHOLESALE_12M, RETAIL_12M, FLOOD_YEAR, FLOOD_12M]
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    people = pd.read_parquet(EDUCATION_PATH, columns=person_columns)

    observed_waves = sorted(int(value) for value in households[YEAR].dropna().unique())
    assert observed_waves == EXPECTED_WAVES, (observed_waves, EXPECTED_WAVES)
    assert len(households) == 62_920
    assert len(people) == 268_485

    frame = build_table(households, people)
    write_workbook(frame)
    validate_output(frame)

    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Shock Support)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(frame.to_string(index=False))


if __name__ == "__main__":
    main()
