#!/usr/bin/env python3
"""Descriptive Outcomes by Historical Conflict Exposure.

Plan: Compare 18 confirmed outcomes, mechanisms, and controls across low,
middle, and high historical-conflict exposure.
Framework: AnaSOP Sections 5.1-5.2, 6.1, 6.3, 6.5, and the baseline-legacy
workflow step in Section 7. Statistics are survey weighted, use the primary
commune linkage, and are descriptive associations rather than causal effects.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT
    / "data/exp/internal_output_archive/tables/Table_descriptive_outcomes_by_historical_conflict_exposure.xlsx"
)

RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
AGE = "Age Years"
ATTENDANCE_ELIGIBLE = "School Attendance Outcome Eligible"
GROUP = "Historical conflict exposure"
GROUP_ORDER = ["Low", "Middle", "High"]

COLUMNS = [
    "Domain and measure",
    "Full sample N",
    "Full weighted mean",
    "Full weighted SD",
    "Low conflict mean",
    "Middle conflict mean",
    "High conflict mean",
    "Standardized difference (High − Low)",
]


@dataclass(frozen=True)
class MeasureSpec:
    domain: str
    variable: str
    label: str
    source: str
    scale: float = 1.0
    agriculture_only: bool = False
    attendance_only: bool = False


MEASURES = [
    MeasureSpec(
        "Agriculture",
        "Cultivated Crop Area m2",
        "Cultivated crop area (ha)",
        "household",
        1e-4,
        True,
    ),
    MeasureSpec(
        "Agriculture",
        "Crop Production Quantity kg",
        "Crop production quantity (1,000 kg)",
        "household",
        0.001,
        True,
    ),
    MeasureSpec(
        "Agriculture",
        "Crop Yield kg per ha",
        "Crop yield (kg/ha)",
        "household",
        agriculture_only=True,
    ),
    MeasureSpec(
        "Agriculture",
        "Post Harvest Loss Share",
        "Post-harvest loss share (%)",
        "household",
        100.0,
        True,
    ),
    MeasureSpec(
        "Agriculture",
        "Crop Diversity Count",
        "Crop diversity (number of crops)",
        "household",
        agriculture_only=True,
    ),
    MeasureSpec(
        "Agriculture",
        "Irrigable Parcel Share",
        "Irrigable parcel share (%)",
        "household",
        100.0,
        True,
    ),
    MeasureSpec(
        "Agriculture",
        "Real 2021 Crop Production Value Riels",
        "Crop production value (million 2021 riels)",
        "household",
        1e-6,
        True,
    ),
    MeasureSpec(
        "Agriculture",
        "Real 2021 Agricultural Input Cost Riels",
        "Agricultural input cost (million 2021 riels)",
        "household",
        1e-6,
        True,
    ),
    MeasureSpec(
        "Consumption and food security",
        "Real 2021 Food Consumption Value per Household Member Riels",
        "Food consumption per member (thousand 2021 riels)",
        "household",
        1e-3,
    ),
    MeasureSpec(
        "Consumption and food security",
        "Food Items with Positive Consumption Count",
        "Food items with positive consumption (count)",
        "household",
    ),
    MeasureSpec(
        "Consumption and food security",
        "Any Severe Food Insecurity Experience",
        "Any severe food insecurity experience (%)",
        "household",
        100.0,
    ),
    MeasureSpec(
        "Consumption and food security",
        "Food Insecurity Severity Sum",
        "Food insecurity severity (index)",
        "household",
    ),
    MeasureSpec(
        "Education",
        "Currently Attending School",
        "Currently attending school, ages 6–17 (%)",
        "education",
        100.0,
        attendance_only=True,
    ),
    MeasureSpec(
        "Education",
        "Years Attended School",
        "Years attended school",
        "education",
    ),
    MeasureSpec(
        "Education",
        "Real 2021 Education Expenditure Riels",
        "Education expenditure (thousand 2021 riels)",
        "education",
        1e-3,
    ),
    MeasureSpec("Controls", "Household Size", "Household size", "household"),
    MeasureSpec("Controls", AGE, "Age (years)", "education"),
    MeasureSpec("Controls", "Female", "Female (%)", "education", 100.0),
]


def assign_conflict_groups(
    values: pd.Series, lower_cut: float, upper_cut: float
) -> pd.Series:
    return pd.cut(
        values,
        bins=[float("-inf"), lower_cut, upper_cut, float("inf")],
        labels=GROUP_ORDER,
        include_lowest=True,
    )


def derive_conflict_cutpoints(households: pd.DataFrame) -> tuple[float, float]:
    communes = households.loc[households[RESOLUTION].eq("commune")]
    conflict_variation = communes.groupby(GEOGRAPHY)[CONFLICT].nunique(dropna=True)
    assert conflict_variation.max() == 1
    geography = (
        communes[[GEOGRAPHY, CONFLICT]]
        .drop_duplicates(GEOGRAPHY)
        .dropna(subset=[CONFLICT])
    )
    assert len(geography) == 1_490
    lower_cut = float(geography[CONFLICT].quantile(1 / 3))
    upper_cut = float(geography[CONFLICT].quantile(2 / 3))
    assert lower_cut < upper_cut
    return lower_cut, upper_cut


def weighted_mean_sd(values: pd.Series, weights: pd.Series) -> tuple[float, float]:
    y = values.to_numpy(dtype=float)
    w = weights.to_numpy(dtype=float)
    mean = float(np.average(y, weights=w))
    variance = float(np.average((y - mean) ** 2, weights=w))
    return mean, float(np.sqrt(variance))


def prepare_measure_sample(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: MeasureSpec,
    lower_cut: float,
    upper_cut: float,
) -> tuple[pd.DataFrame, str]:
    if spec.source == "household":
        source = households
        weight = HOUSEHOLD_WEIGHT
    else:
        source = education
        weight = PERSON_WEIGHT

    required = [RESOLUTION, CONFLICT, spec.variable, weight]
    if spec.agriculture_only and AGRICULTURAL_HOUSEHOLD not in required:
        required.append(AGRICULTURAL_HOUSEHOLD)
    if spec.attendance_only:
        required.extend([AGE, ATTENDANCE_ELIGIBLE])
    sample = source.loc[source[RESOLUTION].eq("commune"), required].copy()
    if spec.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)]
    if spec.attendance_only:
        sample = sample.loc[
            sample[AGE].between(6, 17) & sample[ATTENDANCE_ELIGIBLE].eq(1)
        ]
    sample = sample.dropna(subset=[CONFLICT, spec.variable, weight]).copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    sample[spec.variable] = pd.to_numeric(sample[spec.variable], errors="raise") * spec.scale
    sample[GROUP] = assign_conflict_groups(sample[CONFLICT], lower_cut, upper_cut)
    assert sample[GROUP].notna().all()
    assert sample.groupby(GROUP, observed=True).size().reindex(GROUP_ORDER).gt(0).all()
    return sample, weight


def build_table(
    households: pd.DataFrame, education: pd.DataFrame
) -> tuple[pd.DataFrame, tuple[float, float]]:
    lower_cut, upper_cut = derive_conflict_cutpoints(households)
    rows: list[dict[str, object]] = []
    for spec in MEASURES:
        sample, weight = prepare_measure_sample(
            households, education, spec, lower_cut, upper_cut
        )
        full_mean, full_sd = weighted_mean_sd(sample[spec.variable], sample[weight])
        group_stats: dict[str, tuple[float, float]] = {}
        for group in GROUP_ORDER:
            group_sample = sample.loc[sample[GROUP].eq(group)]
            group_stats[group] = weighted_mean_sd(
                group_sample[spec.variable], group_sample[weight]
            )
        pooled_sd = float(
            np.sqrt((group_stats["Low"][1] ** 2 + group_stats["High"][1] ** 2) / 2)
        )
        standardized_difference = (
            (group_stats["High"][0] - group_stats["Low"][0]) / pooled_sd
            if pooled_sd > 0
            else np.nan
        )
        rows.append(
            {
                "Domain and measure": f"{spec.domain} — {spec.label}",
                "Full sample N": int(len(sample)),
                "Full weighted mean": full_mean,
                "Full weighted SD": full_sd,
                "Low conflict mean": group_stats["Low"][0],
                "Middle conflict mean": group_stats["Middle"][0],
                "High conflict mean": group_stats["High"][0],
                "Standardized difference (High − Low)": standardized_difference,
            }
        )
    table = pd.DataFrame(rows, columns=COLUMNS)
    return table, (lower_cut, upper_cut)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Descriptive Outcomes"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(
        displayName="DescriptiveConflictExposureTable", ref=f"A1:H{last_row}"
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    section_fills = {
        "Agriculture": PatternFill("solid", fgColor="EAF2F8"),
        "Consumption and food security": PatternFill("solid", fgColor="E8F3EC"),
        "Education": PatternFill("solid", fgColor="F3EDF8"),
        "Controls": PatternFill("solid", fgColor="F2F2F2"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=9)
    navy_font = Font(color="1F4E78", bold=True, size=9)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 48

    section_start_rows = {2, 10, 14, 17}
    for row_index in range(2, last_row + 1):
        domain = str(sheet.cell(row=row_index, column=1).value).split(" — ", 1)[0]
        for column_index in range(1, 9):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="left" if column_index == 1 else "right",
                vertical="center",
                wrap_text=column_index == 1,
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = section_fills[domain]
        sheet.cell(row=row_index, column=1).font = navy_font
        sheet.row_dimensions[row_index].height = 34

    for cell in sheet["B"][1:]:
        cell.number_format = "#,##0"
    for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"
    for cell in sheet["H"][1:]:
        cell.number_format = "0.00"

    sheet.conditional_formatting.add(
        f"H2:H{last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="5B9BD5",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="max",
            end_color="ED7D31",
        ),
    )

    widths = [58, 17, 20, 18, 18, 20, 19, 29]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:H{last_row}"
    sheet.sheet_view.zoomScale = 75
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:H{last_row}"
    sheet.page_margins.left = 0.18
    sheet.page_margins.right = 0.18
    sheet.page_margins.top = 0.20
    sheet.page_margins.bottom = 0.20

    workbook.properties.title = "Descriptive Outcomes by Historical Conflict Exposure"
    workbook.properties.subject = (
        "Survey-weighted descriptive associations using primary commune linkage"
    )
    workbook.properties.creator = "Mike Li"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (18, 8), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["Domain and measure"].is_unique
    assert frame["Full sample N"].gt(0).all()
    assert frame.iloc[:, 2:].apply(lambda column: np.isfinite(column.astype(float))).all().all()
    assert frame["Domain and measure"].str.startswith(
        ("Agriculture", "Consumption and food security", "Education", "Controls")
    ).all()

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Descriptive Outcomes"]
    sheet = workbook["Descriptive Outcomes"]
    assert sheet.max_row == 19
    assert sheet.max_column == 8
    assert list(sheet.tables) == ["DescriptiveConflictExposureTable"]
    assert sheet.tables["DescriptiveConflictExposureTable"].ref == "A1:H19"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    household_columns = sorted(
        {
            RESOLUTION,
            GEOGRAPHY,
            CONFLICT,
            HOUSEHOLD_WEIGHT,
            AGRICULTURAL_HOUSEHOLD,
            *[spec.variable for spec in MEASURES if spec.source == "household"],
        }
    )
    education_columns = sorted(
        {
            RESOLUTION,
            CONFLICT,
            PERSON_WEIGHT,
            AGE,
            ATTENDANCE_ELIGIBLE,
            *[spec.variable for spec in MEASURES if spec.source == "education"],
        }
    )
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    frame, (lower_cut, upper_cut) = build_table(households, education)
    write_workbook(frame)
    validate_output(frame)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(f"Commune exposure tertile cuts: {lower_cut:.6f}, {upper_cut:.6f}")


if __name__ == "__main__":
    main()
