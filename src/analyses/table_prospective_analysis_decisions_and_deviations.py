#!/usr/bin/env python3
"""Prospective Analysis Decisions and Deviations.

Plan: Preserve the dated, outcome-blind choices governing the historical-boundary annual
land-NPP analysis and document whether the first activated outcome departed from them.
Framework: AnaSOP Sections 5.3-5.4, 6.8-6.12, and the contract-freezing workflow in
Section 7. The table is an audit record; it does not estimate a new outcome model.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
ANASOP = ROOT / "docs/AnaSOP.md"
FRAMELOG = ROOT / "docs/framelog/mili.md"
OUTPUT = (
    ROOT
    / "data/results/tables/Table_prospective_analysis_decisions_and_deviations.xlsx"
)

COLUMNS = [
    "Date",
    "Stage",
    "Decision",
    "Rationale",
    "Evidence unavailable at decision",
    "Gate affected",
    "Deviation status",
    "Consequence",
]


def build_table() -> pd.DataFrame:
    records = [
        {
            "Date": "2026-08-19",
            "Stage": "Evidence upgrade (MILI-D-20260819-033)",
            "Decision": (
                "Use a three-layer design: local historical-boundary identification, "
                "national bombing-based breadth, and independent validation."
            ),
            "Rationale": (
                "Separate local internal validity from national relevance and external "
                "validation instead of pooling unlike estimands."
            ),
            "Evidence unavailable at decision": (
                "Historical-boundary outcome effects; independent validation outcomes; "
                "mechanism-family results."
            ),
            "Gate affected": "Nature Communications evidence gate",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "A local result alone cannot support a broad national resilience claim."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Source and assignment freeze (MILI-D-20260819-033)",
            "Decision": (
                "Use the checksum-audited public Southwest–West zone geometry and "
                "independently reproduce treatment, distance, and boundary segments."
            ),
            "Rationale": (
                "Bind treatment assignment to a published historical contrast before "
                "examining contemporary outcomes."
            ),
            "Evidence unavailable at decision": (
                "Annual land-NPP interaction estimates and all validation estimates."
            ),
            "Gate affected": "Historical assignment reproduction",
            "Deviation status": "Activated—no deviation",
            "Consequence": (
                "The treatment contrast is higher versus lower repression, not conflict "
                "versus no conflict."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Climate linkage (MILI-D-20260819-034)",
            "Decision": (
                "Use exact historical commune codes first; allow only deterministic "
                "point-in-modern-commune fallback for unresolved legacy codes; never impute."
            ),
            "Rationale": (
                "Recover auditable rainfall linkage after administrative reorganization "
                "without assigning exposure values heuristically."
            ),
            "Evidence unavailable at decision": "Annual land-NPP interaction estimates.",
            "Gate affected": "Linkage and provenance",
            "Deviation status": "Activated—no deviation",
            "Consequence": (
                "Link method remains observable; unmatched or ambiguous cases remain missing."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Outcome definition (MILI-D-20260819-035)",
            "Decision": (
                "Treat MODIS annual land NPP as land vegetation production; retain quality "
                "fields; keep cropland productivity as a separate pending construct."
            ),
            "Rationale": (
                "The activated MODIS panel has no frozen cropland mask and therefore cannot "
                "measure crop yield or cropland-only productivity."
            ),
            "Evidence unavailable at decision": (
                "Land-NPP effects; cropland-masked productivity; independent product checks."
            ),
            "Gate affected": "Outcome validity and interpretation",
            "Deviation status": "Activated—no deviation",
            "Consequence": (
                "Claims are limited to local land NPP unless independent agricultural "
                "validation is later activated."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Blinded feasibility (MILI-D-20260819-036)",
            "Decision": (
                "Block effect estimation until outcome-blind support and clustered-power "
                "analyses are completed across prespecified bandwidths."
            ),
            "Rationale": (
                "Nominal village-year counts overstate information when treatment is spatial "
                "and rainfall and residual dependence are clustered."
            ),
            "Evidence unavailable at decision": (
                "NPP levels, anomalies, quality fields, and all treatment-effect estimates."
            ),
            "Gate affected": "Outcome blinding and effective power",
            "Deviation status": "Passed—no deviation",
            "Consequence": (
                "Bandwidth and precision criteria were selected without inspecting NPP effects."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Equivalence threshold (MILI-D-20260819-037)",
            "Decision": (
                "Set ±0.20 standardized outcome units per one-SD rainfall shock as the "
                "smallest effect of substantive interest."
            ),
            "Rationale": (
                "Use a substantively interpretable threshold that is detectable under the "
                "strong-dependence power design."
            ),
            "Evidence unavailable at decision": "All NPP interaction estimates and confidence intervals.",
            "Gate affected": "Power and equivalence",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "A standardized 95% CI wholly inside ±0.20 supports a substantively precise null."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Support and shock hierarchy (MILI-D-20260819-037)",
            "Decision": (
                "Use 5 km as primary support and May–October rainfall as the primary shock; "
                "retain 2, 10, 15, 20, and 30 km windows and annual rainfall as sensitivities."
            ),
            "Rationale": (
                "Prioritize a local comparison and the vegetation-relevant growing season "
                "while preserving transparent locality and shock-definition checks."
            ),
            "Evidence unavailable at decision": "All NPP interaction estimates by bandwidth and shock.",
            "Gate affected": "Primary specification and robustness",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Wider windows and annual rainfall cannot replace the frozen primary estimate."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Dual-specification freeze (MILI-D-20260819-038)",
            "Decision": (
                "Estimate the all-village 5 km model and require a nine-cross-side-commune "
                "confirmation model with climate-commune-by-year fixed effects."
            ),
            "Rationale": (
                "Address modern administrative and rainfall-geography confounding with a "
                "within-commune, within-year cross-side comparison."
            ),
            "Evidence unavailable at decision": "Primary and confirmation NPP estimates.",
            "Gate affected": "Modern-boundary safeguard",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Strong causal language requires directional and substantive compatibility "
                "across both models."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Covariate policy (MILI-D-20260819-038)",
            "Decision": (
                "Exclude timing-ambiguous 1975 settlement proxies from main controls and "
                "prespecify a hierarchical rainfall-by-river-distance sensitivity."
            ),
            "Rationale": (
                "Avoid conditioning on variables measured at the onset of Khmer Rouge rule "
                "while addressing the reviewed physical river-distance imbalance."
            ),
            "Evidence unavailable at decision": "All NPP interaction and robustness estimates.",
            "Gate affected": "Continuity and control validity",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Settlement differences remain timing diagnostics; river distance is a robustness check."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Inference freeze (MILI-D-20260819-038)",
            "Decision": (
                "Use equal village-year weights and two-way clustering by village and "
                "district-by-year; retain alternative dependence checks as sensitivities."
            ),
            "Rationale": (
                "Allow village serial dependence and contemporaneous spatial dependence "
                "within districts without overstating the five-segment bootstrap."
            ),
            "Evidence unavailable at decision": "NPP coefficients, standard errors, and confidence intervals.",
            "Gate affected": "Uncertainty and influence",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Segment- and commune-level resampling remains diagnostic rather than definitive."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Claim-promotion rule (MILI-D-20260819-033/038)",
            "Decision": (
                "Do not promote the result if specifications conflict, important effects of "
                "both signs remain plausible, one segment drives it, or validation is absent."
            ),
            "Rationale": (
                "Precommit interpretation to stability, equivalence, influence, and external "
                "validation rather than statistical significance alone."
            ),
            "Evidence unavailable at decision": (
                "Local outcome estimates; independent survey, crop, light, price, and mechanism evidence."
            ),
            "Gate affected": "Causal scope and transport",
            "Deviation status": "Open gate—no deviation",
            "Consequence": (
                "Current inference remains local land-NPP equivalence; broad validation gates stay pending."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "First-outcome contract audit (MILI-D-20260819-043/046)",
            "Decision": (
                "Audit the first NPP result against the frozen contract before accepting its "
                "figure and table; record any change as a dated deviation."
            ),
            "Rationale": (
                "Distinguish planned sensitivity analysis from outcome-driven specification changes."
            ),
            "Evidence unavailable at decision": "Not applicable—this is the post-estimation contract audit.",
            "Gate affected": "Prospective integrity",
            "Deviation status": "No deviation recorded",
            "Consequence": (
                "All frozen specifications were reported; the accepted result is a substantively precise local null."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Independent-validation source freeze (MILI-D-20260819-048)",
            "Decision": (
                "Sequence observed EOG VIIRS first, Landsat with an ESA cropland mask second, "
                "CSES 2023 as the socioeconomic holdout, and CAS 2024 as national transport evidence."
            ),
            "Rationale": (
                "Freeze provenance and intended estimands before validation outcomes are inspected."
            ),
            "Evidence unavailable at decision": (
                "All VIIRS, cropland, holdout-survey, and agricultural-validation estimates."
            ),
            "Gate affected": "Validation provenance and sequencing",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Local validation remains distinct from national transport, and unavailable sources stay pending."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "VIIRS outcome definition (MILI-D-20260819-049)",
            "Decision": (
                "Use asinh annual mean radiance over all pixels as primary; retain asinh median radiance, "
                "any nonzero radiance, and coverage restrictions as robustness checks; exclude 2012."
            ),
            "Rationale": (
                "Preserve rural zero-light observations and freeze transformations before boundary-side effects."
            ),
            "Evidence unavailable at decision": "All historical-boundary VIIRS interaction estimates.",
            "Gate affected": "Independent outcome validity",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "The comparable observed-sensor validation period is 2013-2021 and zero-light cells remain in scope."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "VIIRS quality-screen freeze (MILI-D-20260819-050)",
            "Decision": (
                "Retain all observed pixel-years in the main model; require at least 40 cloud-free "
                "observations only in the frozen quality sensitivity; keep the 30-observation flag diagnostic."
            ),
            "Rationale": (
                "Set a non-outcome-tuned quality check after a treatment-blinded coverage audit."
            ),
            "Evidence unavailable at decision": "All VIIRS treatment-by-rainfall coefficients.",
            "Gate affected": "Measurement quality and support",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "Coverage filtering cannot be selected after viewing the historical-boundary interaction."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "VIIRS contract audit (MILI-D-20260819-051)",
            "Decision": (
                "Audit the observed-VIIRS result against the frozen primary, mandatory confirmation, "
                "outcome, coverage, weighting, rainfall, bandwidth, and equivalence specifications."
            ),
            "Rationale": (
                "Apply the same prospective-integrity standard used for the first NPP outcome."
            ),
            "Evidence unavailable at decision": "Not applicable—this is the post-estimation contract audit.",
            "Gate affected": "Independent-validation integrity",
            "Deviation status": "Passed—no deviation",
            "Consequence": (
                "VIIRS supports a narrow precise-null shock-response conclusion, not absence of a level legacy."
            ),
        },
        {
            "Date": "2026-08-19",
            "Stage": "Validation-output separation (MILI-D-20260819-052)",
            "Decision": (
                "Report completed VIIRS evidence in separate appendix outputs while retaining the "
                "multi-source validation figure and main-text table as pending."
            ),
            "Rationale": (
                "Avoid presenting one spatial outcome as completion of the full validation program."
            ),
            "Evidence unavailable at decision": (
                "Holdout survey, independent agriculture, cropland-productivity, and external-price results."
            ),
            "Gate affected": "Validation completeness and placement",
            "Deviation status": "Activated—no deviation",
            "Consequence": (
                "Accepted VIIRS evidence is reportable without weakening the multi-source evidence gate."
            ),
        },
        {
            "Date": "2026-08-20",
            "Stage": "Reconstructed-light exclusion (MILI-D-20260820-001)",
            "Decision": (
                "Defer CNN-reconstructed long-run nighttime-light products and retain observed EOG VIIRS "
                "as the current nighttime-activity validation source."
            ),
            "Rationale": (
                "Climate-sensitive vegetation inputs in reconstructed products may compromise independence "
                "for rainfall-shock validation."
            ),
            "Evidence unavailable at decision": "No reconstructed-light estimates were downloaded or inspected.",
            "Gate affected": "Validation-source independence",
            "Deviation status": "Deferred—no deviation",
            "Consequence": (
                "The nighttime-activity conclusion remains based on a directly observed sensor series."
            ),
        },
        {
            "Date": "2026-08-20",
            "Stage": "Mechanism-variable activation (MILI-D-20260820-004/005)",
            "Decision": (
                "Activate village irrigated-land share as primary, retain permanent-market access as a "
                "two-wave appendix measure, and defer conditional road and sparse costly-coping variables."
            ),
            "Rationale": (
                "Use cross-wave comparable support and avoid interpreting routed missingness or rare outcomes."
            ),
            "Evidence unavailable at decision": "All activated village-mechanism regression estimates.",
            "Gate affected": "Mechanism measurement and support",
            "Deviation status": "Activated—no deviation",
            "Consequence": (
                "Mechanism evidence is limited to defensible infrastructure and agricultural-capacity measures."
            ),
        },
        {
            "Date": "2026-08-20",
            "Stage": "Mechanism estimation freeze (MILI-D-20260820-006)",
            "Decision": (
                "Use province-by-wave persistence gradients and a geography plus province-by-wave drought "
                "interaction; estimate villages unweighted, retain population exposure weighting as sensitivity, "
                "cluster by linked historical geography, and adjust six primary tests with Holm."
            ),
            "Rationale": (
                "Separate village observations from household weights and prevent selective mechanism-family definition."
            ),
            "Evidence unavailable at decision": "All six primary mechanism estimates and weighting sensitivities.",
            "Gate affected": "Mechanism estimand and multiplicity",
            "Deviation status": "Frozen—no deviation",
            "Consequence": (
                "No mechanism can be called robust unless it survives the expanded six-test family."
            ),
        },
        {
            "Date": "2026-08-20",
            "Stage": "Mechanism-family consistency audit (MILI-D-20260820-007/009)",
            "Decision": (
                "Use the six-test Holm family in every retained mechanism output and characterize parcel "
                "irrigation as directionally coherent but not multiplicity-robust."
            ),
            "Rationale": (
                "Eliminate the obsolete three-test adjusted p-value and maintain one inferential standard."
            ),
            "Evidence unavailable at decision": "Not applicable—this is a post-estimation cross-output audit.",
            "Gate affected": "Multiplicity and reporting consistency",
            "Deviation status": "Remediated—no design deviation",
            "Consequence": (
                "The authoritative parcel-irrigation Holm p-value is approximately 0.106 rather than 0.053."
            ),
        },
    ]
    return pd.DataFrame.from_records(records, columns=COLUMNS)


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Decision Log"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    last_row = len(frame) + 1
    excel_table = Table(displayName="ProspectiveDecisionAudit", ref=f"A1:H{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    bottom_rule = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=bottom_rule)
    worksheet.row_dimensions[1].height = 38

    status_fill = PatternFill("solid", fgColor="E2F0D9")
    open_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_index in range(2, last_row + 1):
        for column_index in range(1, 9):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(size=8.2)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        status_cell = worksheet.cell(row=row_index, column=7)
        status_cell.fill = (
            open_fill if str(status_cell.value).startswith("Open gate") else status_fill
        )
        worksheet.row_dimensions[row_index].height = 54

    widths = [14, 34, 52, 48, 45, 31, 24, 48]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:H{last_row}"
    worksheet.sheet_view.zoomScale = 75
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:H{last_row}"
    worksheet.page_margins.left = 0.18
    worksheet.page_margins.right = 0.18
    worksheet.page_margins.top = 0.20
    worksheet.page_margins.bottom = 0.20

    workbook.properties.title = "Prospective Analysis Decisions and Deviations"
    workbook.properties.subject = "Historical-boundary, validation, and mechanism contract audit"
    workbook.properties.creator = "Mike Li"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def validate_output(frame: pd.DataFrame) -> None:
    anasop = ANASOP.read_text(encoding="utf-8")
    framelog = FRAMELOG.read_text(encoding="utf-8")
    required_records = [
        "MILI-D-20260819-033",
        "MILI-D-20260819-034",
        "MILI-D-20260819-035",
        "MILI-D-20260819-036",
        "MILI-D-20260819-037",
        "MILI-D-20260819-038",
        "MILI-D-20260819-043",
        "MILI-D-20260819-046",
        "MILI-D-20260819-048",
        "MILI-D-20260819-049",
        "MILI-D-20260819-050",
        "MILI-D-20260819-051",
        "MILI-D-20260819-052",
        "MILI-D-20260820-001",
        "MILI-D-20260820-004",
        "MILI-D-20260820-005",
        "MILI-D-20260820-006",
        "MILI-D-20260820-007",
        "MILI-D-20260820-009",
    ]
    assert all(record in framelog for record in required_records)
    assert "0.20 standardized-outcome" in anasop
    assert "mandatory confirmation" in anasop.lower()
    assert frame.shape == (21, 8), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["Deviation status"].str.contains("deviation").all()
    assert frame["Deviation status"].eq("No deviation recorded").sum() == 1
    assert frame["Date"].eq("2026-08-20").sum() == 4
    assert frame["Decision"].str.contains("six-test Holm").any()
    assert frame["Consequence"].str.contains("0.106 rather than 0.053", regex=False).any()

    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == ["Decision Log"]
    worksheet = workbook["Decision Log"]
    assert worksheet.max_row == 22 and worksheet.max_column == 8
    assert set(worksheet.tables.keys()) == {"ProspectiveDecisionAudit"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?"))


def main() -> None:
    frame = build_table()
    write_workbook(frame)
    validate_output(frame)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print("Workbook sheets: 1 (Decision Log)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(frame[["Stage", "Gate affected", "Deviation status"]].to_string(index=False))


if __name__ == "__main__":
    main()
