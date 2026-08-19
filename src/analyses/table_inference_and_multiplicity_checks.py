#!/usr/bin/env python3
"""Inference and Multiplicity Checks.

Plan: Re-estimate the 24 central outcome-by-shock interactions and report
geography-clustered inference, Spatial-HAC inference at 50/100/200 km,
province-clustered restricted wild-bootstrap inference for price models, and
Holm-adjusted inference within outcome families.
Framework: AnaSOP Sections 5.1-5.2, 6.2-6.7, and the inference workflow step
in Section 7. Spatial-HAC combines serial clustering within historical
geography with Bartlett-weighted same-wave covariance across commune
centroids. Price wild-bootstrap p-values use 999 Rademacher draws.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy import sparse
from scipy.spatial import cKDTree
from scipy.stats import norm


ROOT = Path(__file__).resolve().parents[2]
HOUSEHOLD_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
BOUNDARY_PATH = ROOT / "data/raw/geography/odc_cambodia_communes_2014.gpkg"
OUTPUT_PATH = ROOT / "data/results/tables/Table_inference_and_multiplicity_checks.xlsx"

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
HOUSEHOLD_PROVINCE = "Province Code Component"
EDUCATION_PROVINCE = "Province Code"
HOUSEHOLD_WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
EXTREME_WET = "Annual Rainfall Extreme Wet Shock"
PRICE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"

CROP_YIELD = "Crop Yield kg per ha"
LOSS_SHARE = "Post Harvest Loss Share"
CROP_VALUE = "Real 2021 Crop Production Value Riels"
FOOD_CONSUMPTION = "Real 2021 Food Consumption Value per Household Member Riels"
SEVERE_FOOD_INSECURITY = "Any Severe Food Insecurity Experience"
FOOD_INSECURITY_SEVERITY = "Food Insecurity Severity Sum"
ATTENDANCE = "Currently Attending School"
EDUCATION_EXPENDITURE = "Real 2021 Education Expenditure Riels"
IRRIGATION = "Irrigable Parcel Share"
DIVERSITY = "Crop Diversity Count"
INPUT_COST = "Real 2021 Agricultural Input Cost Riels"

SHOCK_LABELS = {
    SPI12: "SPI-12 (1 SD; higher = wetter)",
    EXTREME_WET: "Annual extreme wet (0→1)",
    PRICE_12M: "Local rice-price change (1 SD)",
}
CUTOFFS_KM = (50, 100, 200)
BOOTSTRAP_DRAWS = 999

COLUMNS = [
    "Domain",
    "Outcome (model scale)",
    "Contemporary shock",
    "Estimate",
    "Geography-clustered p",
    "Spatial-HAC p (50 km)",
    "Spatial-HAC p (100 km)",
    "Spatial-HAC p (200 km)",
    "Province wild-bootstrap p",
    "Holm-adjusted p",
    "N",
    "Estimability status",
]


@dataclass(frozen=True)
class OutcomeSpec:
    domain: str
    outcome: str
    label: str
    source: str
    transform: str
    agriculture_only: bool = False
    school_age_only: bool = False


@dataclass(frozen=True)
class EstimateSpec:
    outcome: OutcomeSpec
    shock: str
    shock_multiplier: float = 1.0
    shock_label: str | None = None


OUTCOMES = [
    OutcomeSpec("Agriculture", CROP_YIELD, "Crop yield (asinh)", "household", "asinh", True),
    OutcomeSpec("Agriculture", LOSS_SHARE, "Post-harvest loss (percentage points)", "household", "percentage_points", True),
    OutcomeSpec("Agriculture", CROP_VALUE, "Crop production value (asinh, 2021 riels)", "household", "asinh", True),
    OutcomeSpec("Consumption", FOOD_CONSUMPTION, "Food consumption per member (asinh, 2021 riels)", "household", "asinh"),
    OutcomeSpec("Food security", SEVERE_FOOD_INSECURITY, "Severe food insecurity (percentage points)", "household", "percentage_points"),
    OutcomeSpec("Food security", FOOD_INSECURITY_SEVERITY, "Food insecurity severity (index points)", "household", "level"),
    OutcomeSpec("Education", ATTENDANCE, "School attendance, ages 6–17 (percentage points)", "education", "percentage_points", school_age_only=True),
    OutcomeSpec("Education", EDUCATION_EXPENDITURE, "Education expenditure (asinh, 2021 riels)", "education", "asinh"),
]
SPECS = [
    EstimateSpec(outcome, shock)
    for outcome in OUTCOMES
    for shock in (SPI12, EXTREME_WET, PRICE_12M)
]
SPECS.extend(
    [
        EstimateSpec(
            OutcomeSpec(
                "Mechanism",
                IRRIGATION,
                "Irrigable parcel share (percentage points)",
                "household",
                "percentage_points",
                True,
            ),
            SPI12,
            -1.0,
            "Drought severity (−SPI-12; 1 SD)",
        ),
        EstimateSpec(
            OutcomeSpec(
                "Mechanism",
                DIVERSITY,
                "Crop diversity (count)",
                "household",
                "level",
                True,
            ),
            SPI12,
            -1.0,
            "Drought severity (−SPI-12; 1 SD)",
        ),
        EstimateSpec(
            OutcomeSpec(
                "Mechanism",
                INPUT_COST,
                "Agricultural input cost (asinh, 2021 riels)",
                "household",
                "asinh",
                True,
            ),
            SPI12,
            -1.0,
            "Drought severity (−SPI-12; 1 SD)",
        ),
    ]
)


@dataclass(frozen=True)
class FittedInference:
    coefficient: float
    primary_p: float
    spatial_p: dict[int, float]
    wild_p: float | None
    sample_size: int
    geography_clusters: int
    province_clusters: int


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def transform_outcome(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "asinh":
        if np.nanmin(values) < 0:
            raise ValueError("asinh outcome unexpectedly contains negative values")
        return np.arcsinh(values)
    if transform == "percentage_points":
        if np.nanmin(values) < 0 or np.nanmax(values) > 1:
            raise ValueError("Share or binary outcome falls outside [0, 1]")
        return 100.0 * values
    if transform == "level":
        return values
    raise ValueError(f"Unknown outcome transform: {transform}")


def prepare_sample(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: EstimateSpec,
) -> tuple[pd.DataFrame, str, str, list[str]]:
    outcome = spec.outcome
    if outcome.source == "household":
        source = households
        weight = HOUSEHOLD_WEIGHT
        province = HOUSEHOLD_PROVINCE
        controls = [HOUSEHOLD_SIZE]
    else:
        source = education
        weight = PERSON_WEIGHT
        province = EDUCATION_PROVINCE
        controls = [HOUSEHOLD_SIZE, AGE, FEMALE]

    required = [
        outcome.outcome,
        spec.shock,
        CONFLICT,
        weight,
        GEOGRAPHY,
        province,
        YEAR,
        *controls,
    ]
    if outcome.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    required = list(dict.fromkeys(required))
    sample = source.loc[source[RESOLUTION].eq("commune"), required].dropna().copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    if outcome.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)].copy()
    if outcome.school_age_only:
        sample = sample.loc[sample[AGE].between(6, 17)].copy()
    sample[GEOGRAPHY] = sample[GEOGRAPHY].astype(str).str.zfill(6)
    if sample[GEOGRAPHY].nunique() < 50:
        raise ValueError(f"Too few geography clusters for {outcome.label}")
    return sample, weight, province, controls


def load_spatial_weights() -> tuple[dict[str, int], dict[int, sparse.csr_matrix]]:
    communes = gpd.read_file(BOUNDARY_PATH, columns=["com_code", "geometry"])
    communes[GEOGRAPHY] = communes["com_code"].astype(int).astype(str).str.zfill(6)
    if not communes[GEOGRAPHY].is_unique:
        raise ValueError("Commune boundary codes are not unique")
    projected = communes.to_crs(32648)
    centroids = projected.geometry.centroid
    coordinates = np.column_stack([centroids.x.to_numpy(), centroids.y.to_numpy()])
    geography_to_position = {
        code: position for position, code in enumerate(projected[GEOGRAPHY].tolist())
    }
    tree = cKDTree(coordinates)
    matrices: dict[int, sparse.csr_matrix] = {}
    for cutoff in CUTOFFS_KM:
        pairs = np.array(list(tree.query_pairs(cutoff * 1000.0)), dtype=int)
        if pairs.size == 0:
            raise ValueError(f"No commune pairs found within {cutoff} km")
        distances = np.linalg.norm(
            coordinates[pairs[:, 0]] - coordinates[pairs[:, 1]], axis=1
        )
        weights = 1.0 - distances / (cutoff * 1000.0)
        rows = np.concatenate([pairs[:, 0], pairs[:, 1]])
        columns = np.concatenate([pairs[:, 1], pairs[:, 0]])
        values = np.concatenate([weights, weights])
        matrices[cutoff] = sparse.csr_matrix(
            (values, (rows, columns)), shape=(len(projected), len(projected))
        )
    return geography_to_position, matrices


def cluster_meat(
    x: np.ndarray,
    residuals: np.ndarray,
    clusters: pd.Series,
) -> tuple[np.ndarray, int]:
    codes, inverse = np.unique(clusters.astype(str).to_numpy(), return_inverse=True)
    scores = np.zeros((len(codes), x.shape[1]), dtype=float)
    np.add.at(scores, inverse, x * residuals[:, None])
    return scores.T @ scores, len(codes)


def finite_sample_factor(n: int, k: int, clusters: int) -> float:
    if clusters <= 1 or n <= k:
        raise ValueError("Insufficient degrees of freedom for clustered inference")
    return (clusters / (clusters - 1.0)) * ((n - 1.0) / (n - k))


def spatial_hac_pvalues(
    x: np.ndarray,
    residuals: np.ndarray,
    coefficient: float,
    target_index: int,
    sample: pd.DataFrame,
    geography_to_position: dict[str, int],
    spatial_weights: dict[int, sparse.csr_matrix],
) -> dict[int, float]:
    n, k = x.shape
    inv_xx = np.linalg.inv(x.T @ x)
    geography = sample[GEOGRAPHY].astype(str).to_numpy()
    missing = sorted(set(geography) - set(geography_to_position))
    if missing:
        raise ValueError(f"Missing centroid coordinates for {len(missing)} geographies")
    geography_positions = np.array(
        [geography_to_position[value] for value in geography], dtype=int
    )
    observation_scores = x * residuals[:, None]
    geography_scores = np.zeros((len(geography_to_position), k), dtype=float)
    np.add.at(geography_scores, geography_positions, observation_scores)
    cluster_component = geography_scores.T @ geography_scores
    cluster_count = int(np.unique(geography).size)
    adjustment = finite_sample_factor(n, k, cluster_count)

    years = sample[YEAR].astype(int).to_numpy()
    results: dict[int, float] = {}
    for cutoff, weight_matrix in spatial_weights.items():
        meat = cluster_component.copy()
        for year in np.unique(years):
            mask = years == year
            year_scores = np.zeros((len(geography_to_position), k), dtype=float)
            np.add.at(
                year_scores,
                geography_positions[mask],
                observation_scores[mask],
            )
            meat += year_scores.T @ (weight_matrix @ year_scores)
        covariance = adjustment * inv_xx @ meat @ inv_xx
        variance = float(covariance[target_index, target_index])
        if not np.isfinite(variance) or variance <= 0:
            raise ValueError(f"Invalid Spatial-HAC variance at {cutoff} km")
        t_statistic = coefficient / np.sqrt(variance)
        results[cutoff] = float(2.0 * norm.sf(abs(t_statistic)))
    return results


def province_wild_bootstrap_pvalue(
    x: np.ndarray,
    y: np.ndarray,
    target_index: int,
    provinces: pd.Series,
    seed: int,
) -> tuple[float, int]:
    n, k = x.shape
    province_codes, inverse = np.unique(
        provinces.astype(str).to_numpy(), return_inverse=True
    )
    cluster_count = len(province_codes)
    inv_xx = np.linalg.inv(x.T @ x)
    beta = inv_xx @ (x.T @ y)
    residuals = y - x @ beta
    observed_meat, _ = cluster_meat(x, residuals, provinces)
    adjustment = finite_sample_factor(n, k, cluster_count)
    observed_covariance = adjustment * inv_xx @ observed_meat @ inv_xx
    observed_se = float(np.sqrt(observed_covariance[target_index, target_index]))
    observed_t = float(beta[target_index] / observed_se)

    restricted_x = np.delete(x, target_index, axis=1)
    restricted_beta = np.linalg.lstsq(restricted_x, y, rcond=None)[0]
    restricted_residuals = y - restricted_x @ restricted_beta
    restricted_scores = np.zeros((cluster_count, k), dtype=float)
    cluster_xx = np.zeros((cluster_count, k, k), dtype=float)
    for cluster in range(cluster_count):
        mask = inverse == cluster
        cluster_x = x[mask]
        restricted_scores[cluster] = cluster_x.T @ restricted_residuals[mask]
        cluster_xx[cluster] = cluster_x.T @ cluster_x

    rng = np.random.default_rng(seed)
    exceedances = 0
    valid_draws = 0
    for _ in range(BOOTSTRAP_DRAWS):
        multipliers = rng.choice(np.array([-1.0, 1.0]), size=cluster_count)
        delta = inv_xx @ (multipliers[:, None] * restricted_scores).sum(axis=0)
        bootstrap_scores = (
            multipliers[:, None] * restricted_scores
            - np.einsum("gij,j->gi", cluster_xx, delta)
        )
        bootstrap_meat = bootstrap_scores.T @ bootstrap_scores
        bootstrap_covariance = adjustment * inv_xx @ bootstrap_meat @ inv_xx
        variance = float(bootstrap_covariance[target_index, target_index])
        if not np.isfinite(variance) or variance <= 0:
            continue
        bootstrap_t = float(delta[target_index] / np.sqrt(variance))
        valid_draws += 1
        exceedances += int(abs(bootstrap_t) >= abs(observed_t))
    if valid_draws < int(0.95 * BOOTSTRAP_DRAWS):
        raise ValueError("Too few valid province wild-bootstrap draws")
    return (exceedances + 1.0) / (valid_draws + 1.0), cluster_count


def fit_inference(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: EstimateSpec,
    geography_to_position: dict[str, int],
    spatial_weights: dict[int, sparse.csr_matrix],
    seed: int,
) -> FittedInference:
    sample, weight_column, province_column, controls = prepare_sample(
        households, education, spec
    )
    weights = sample[weight_column].to_numpy(dtype=float)
    conflict_z = weighted_standardize(sample[CONFLICT].to_numpy(dtype=float), weights)
    raw_shock = sample[spec.shock].to_numpy(dtype=float)
    shock_model = (
        raw_shock
        if spec.shock == EXTREME_WET
        else weighted_standardize(raw_shock, weights)
    )
    shock_model = spec.shock_multiplier * shock_model
    outcome = transform_outcome(
        sample[spec.outcome.outcome].to_numpy(dtype=float), spec.outcome.transform
    )

    exogenous = pd.DataFrame(
        {
            "shock": shock_model,
            "conflict_x_shock": conflict_z * shock_model,
        },
        index=sample.index,
    )
    for control in controls:
        exogenous[control] = sample[control].to_numpy(dtype=float)
    province_wave = (
        sample[province_column].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {
            "geography": sample[GEOGRAPHY].astype("category"),
            "province_wave": province_wave.astype("category"),
        },
        index=sample.index,
    )
    model = AbsorbingLS(
        dependent=pd.Series(outcome, index=sample.index, name="outcome"),
        exog=exogenous,
        absorb=absorbed,
        weights=pd.Series(weights, index=sample.index),
        drop_absorbed=True,
    )
    fitted = model.fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": sample[GEOGRAPHY]}, index=sample.index),
        debiased=True,
    )
    target = "conflict_x_shock"
    if target not in fitted.params.index:
        raise ValueError(f"Interaction absorbed for {spec.outcome.label}")
    x = model.absorbed_exog.to_numpy(dtype=float)
    y = model.absorbed_dependent.to_numpy(dtype=float).reshape(-1)
    target_index = list(model.absorbed_exog.columns).index(target)
    coefficient = float(fitted.params[target])
    residuals = y - x @ fitted.params.to_numpy(dtype=float)
    spatial_p = spatial_hac_pvalues(
        x,
        residuals,
        coefficient,
        target_index,
        sample,
        geography_to_position,
        spatial_weights,
    )
    province_clusters = int(sample[province_column].astype(str).nunique())
    wild_p: float | None = None
    if spec.shock == PRICE_12M:
        wild_p, province_clusters = province_wild_bootstrap_pvalue(
            x,
            y,
            target_index,
            sample[province_column],
            seed,
        )
    return FittedInference(
        coefficient=coefficient,
        primary_p=float(fitted.pvalues[target]),
        spatial_p=spatial_p,
        wild_p=wild_p,
        sample_size=len(sample),
        geography_clusters=int(sample[GEOGRAPHY].nunique()),
        province_clusters=province_clusters,
    )


def multiplicity_family(domain: str) -> str:
    if domain in {"Consumption", "Food security"}:
        return "Consumption and food security"
    return domain


def holm_adjust(p_values: list[float]) -> list[float]:
    values = np.asarray(p_values, dtype=float)
    order = np.argsort(values)
    adjusted_sorted = np.empty_like(values)
    running_maximum = 0.0
    total = len(values)
    for rank, position in enumerate(order):
        candidate = min(1.0, (total - rank) * values[position])
        running_maximum = max(running_maximum, candidate)
        adjusted_sorted[position] = running_maximum
    return adjusted_sorted.tolist()


def build_table(
    households: pd.DataFrame,
    education: pd.DataFrame,
    geography_to_position: dict[str, int],
    spatial_weights: dict[int, sparse.csr_matrix],
) -> pd.DataFrame:
    fitted_results: list[FittedInference] = []
    for index, spec in enumerate(SPECS):
        fitted_results.append(
            fit_inference(
                households,
                education,
                spec,
                geography_to_position,
                spatial_weights,
                seed=20260819 + index,
            )
        )
        print(
            f"Estimated {index + 1:02d}/{len(SPECS)}: "
            f"{spec.outcome.label} × {SHOCK_LABELS[spec.shock]}"
        )

    adjusted = [np.nan] * len(SPECS)
    families = [multiplicity_family(spec.outcome.domain) for spec in SPECS]
    for family in sorted(set(families)):
        positions = [index for index, value in enumerate(families) if value == family]
        corrected = holm_adjust([fitted_results[index].primary_p for index in positions])
        for position, value in zip(positions, corrected):
            adjusted[position] = value

    rows: list[dict[str, object]] = []
    for spec, result, adjusted_p in zip(SPECS, fitted_results, adjusted):
        if spec.shock == PRICE_12M:
            status = (
                f"All estimable; {result.geography_clusters} geography and "
                f"{result.province_clusters} province clusters"
            )
        else:
            status = (
                f"Wild bootstrap not applicable; {result.geography_clusters} "
                "geography clusters"
            )
        rows.append(
            {
                "Domain": spec.outcome.domain,
                "Outcome (model scale)": spec.outcome.label,
                "Contemporary shock": spec.shock_label or SHOCK_LABELS[spec.shock],
                "Estimate": result.coefficient,
                "Geography-clustered p": result.primary_p,
                "Spatial-HAC p (50 km)": result.spatial_p[50],
                "Spatial-HAC p (100 km)": result.spatial_p[100],
                "Spatial-HAC p (200 km)": result.spatial_p[200],
                "Province wild-bootstrap p": result.wild_p,
                "Holm-adjusted p": adjusted_p,
                "N": result.sample_size,
                "Estimability status": status,
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inference and Multiplicity"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="InferenceMultiplicityTable", ref=f"A1:L{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    domain_fills = {
        "Agriculture": PatternFill("solid", fgColor="EAF2F8"),
        "Consumption": PatternFill("solid", fgColor="E2F0D9"),
        "Food security": PatternFill("solid", fgColor="DDEBF7"),
        "Education": PatternFill("solid", fgColor="FFF2CC"),
        "Mechanism": PatternFill("solid", fgColor="E4DFEC"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=8)
    navy_bold = Font(color="1F4E78", bold=True, size=8)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")
    section_start_rows = {2, 11, 14, 20, 26}

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 55

    for row_index in range(2, last_row + 1):
        domain = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 13):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {4, 5, 6, 7, 8, 9, 10, 11} else "left",
                vertical="center",
                wrap_text=column_index not in {4, 5, 6, 7, 8, 9, 10, 11},
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = domain_fills[domain]
        sheet.cell(row=row_index, column=1).font = navy_bold
        sheet.row_dimensions[row_index].height = 40

    for column in ("D", "E", "F", "G", "H", "I", "J"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.0000"
    for cell in sheet["K"][1:]:
        cell.number_format = "#,##0"

    widths = [18, 39, 28, 13, 18, 18, 18, 18, 23, 18, 12, 41]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:L{last_row}"
    sheet.sheet_view.zoomScale = 60
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:L{last_row}"
    sheet.page_margins.left = 0.10
    sheet.page_margins.right = 0.10
    sheet.page_margins.top = 0.13
    sheet.page_margins.bottom = 0.13

    workbook.properties.title = "Inference and Multiplicity Checks"
    workbook.properties.subject = "Spatial, coarse-cluster, and multiplicity sensitivity for central interaction estimates"
    workbook.properties.creator = "Mike Li"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (27, 12), frame.shape
    assert list(frame.columns) == COLUMNS
    assert frame["N"].gt(0).all()
    numeric_columns = [
        "Estimate",
        "Geography-clustered p",
        "Spatial-HAC p (50 km)",
        "Spatial-HAC p (100 km)",
        "Spatial-HAC p (200 km)",
        "Holm-adjusted p",
    ]
    for column in numeric_columns:
        assert np.isfinite(frame[column]).all(), column
    p_columns = numeric_columns[1:] + ["Province wild-bootstrap p"]
    for column in p_columns:
        observed = frame[column].dropna()
        assert observed.between(0, 1).all(), column
    price_rows = frame["Contemporary shock"].eq(SHOCK_LABELS[PRICE_12M])
    assert frame.loc[price_rows, "Province wild-bootstrap p"].notna().all()
    assert frame.loc[~price_rows, "Province wild-bootstrap p"].isna().all()
    assert frame.groupby("Domain").size().to_dict() == {
        "Agriculture": 9,
        "Consumption": 3,
        "Education": 6,
        "Food security": 6,
        "Mechanism": 3,
    }

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Inference and Multiplicity"]
    sheet = workbook["Inference and Multiplicity"]
    assert sheet.max_row == 28
    assert sheet.max_column == 12
    assert list(sheet.tables) == ["InferenceMultiplicityTable"]
    assert sheet.tables["InferenceMultiplicityTable"].ref == "A1:L28"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    household_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            HOUSEHOLD_PROVINCE,
            HOUSEHOLD_WEIGHT,
            HOUSEHOLD_SIZE,
            AGRICULTURAL_HOUSEHOLD,
            CONFLICT,
            SPI12,
            EXTREME_WET,
            PRICE_12M,
            *[spec.outcome.outcome for spec in SPECS if spec.outcome.source == "household"],
        }
    )
    education_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            EDUCATION_PROVINCE,
            PERSON_WEIGHT,
            HOUSEHOLD_SIZE,
            AGE,
            FEMALE,
            CONFLICT,
            SPI12,
            EXTREME_WET,
            PRICE_12M,
            *[spec.outcome.outcome for spec in SPECS if spec.outcome.source == "education"],
        }
    )
    households = pd.read_parquet(HOUSEHOLD_PATH, columns=household_columns)
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    geography_to_position, spatial_weights = load_spatial_weights()
    table = build_table(
        households,
        education,
        geography_to_position,
        spatial_weights,
    )
    write_workbook(table)
    validate_output(table)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {table.shape[0]} rows x {table.shape[1]} columns")
    print("Price wild-bootstrap rows:", int(table["Province wild-bootstrap p"].notna().sum()))


if __name__ == "__main__":
    main()
