#!/usr/bin/env python3
"""Alternative Conflict Measures and Shock Definitions.

Plan: Compare 20 alternative historical-conflict, climate, food-price,
linkage, weighting, and satellite specifications with their corresponding
core interaction estimates.
Framework: AnaSOP Sections 5.1-5.2, 6.2-6.7, and the robustness and satellite
workflow steps in Section 7. Definition comparisons use coverage-matched
samples unless the row explicitly tests linkage scope.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
INPUT_PATH = (
    ROOT / "data/processed/direction3_household_conflict_shock_preprocessed.parquet"
)
EDUCATION_PATH = (
    ROOT / "data/processed/direction3_education_conflict_shock_preprocessed.parquet"
)
OUTPUT_PATH = (
    ROOT
    / "data/results/tables/Table_alternative_conflict_measures_and_shock_definitions.xlsx"
)

YEAR = "Survey Year"
RESOLUTION = "Climate Geography Resolution"
GEOGRAPHY = "Climate Geography Code"
PROVINCE = "Province Code Component"
WEIGHT = "Household Survey Weight"
PERSON_WEIGHT = "Person Survey Weight"
HOUSEHOLD_SIZE = "Household Size"
AGE = "Age Years"
FEMALE = "Female"
AGRICULTURAL_HOUSEHOLD = "Agricultural Household"
EDUCATION_PROVINCE = "Province Code"

CORE_CONFLICT = "Log Bombing Unique Locations per 100 km2"
SPI12 = "Interview Month SPI 12 Month"
EXTREME_WET = "Annual Rainfall Extreme Wet Shock"
PRICE_12M = "12 Month Change in Local Relative Log Wholesale Rice Price"

CROP_YIELD = "Crop Yield kg per ha"
FOOD_CONSUMPTION = "Real 2021 Food Consumption Value per Household Member Riels"
ATTENDANCE = "Currently Attending School"

COLUMNS = [
    "Robustness family",
    "Outcome (model scale)",
    "Comparison (alternative vs core)",
    "Core estimate",
    "Alternative estimate",
    "Alternative 95% CI",
    "Alternative N",
    "Conflict definition",
    "Shock definition",
    "Specification indicators",
]


@dataclass(frozen=True)
class RobustnessSpec:
    family: str
    outcome: str
    outcome_label: str
    comparison: str
    core_shock: str
    core_shock_binary: bool
    alternative_conflict: str = CORE_CONFLICT
    alternative_conflict_label: str = "Log bombing-location density"
    alternative_conflict_transform: str = "identity"
    alternative_shock: str = SPI12
    alternative_shock_label: str = "SPI-12 (higher = wetter)"
    alternative_shock_binary: bool = False
    agriculture_only: bool = True
    alternative_commune_only: bool = True
    alternative_weighted: bool = True
    matched_support: bool = True
    source: str = "household"
    outcome_transform: str = "asinh"
    school_age_only: bool = False


SPECS = [
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Any bombing indicator vs log bombing density",
        SPI12,
        False,
        "Any US Bombing Record",
        "Any U.S. bombing record (standardized)",
    ),
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Bombing record intensity vs log bombing-location density",
        SPI12,
        False,
        "Bombing Record Count",
        "Log(1 + bombing record count), standardized",
        "log1p",
    ),
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Khmer Rouge prison count vs bombing density",
        SPI12,
        False,
        "Khmer Rouge Prison Count",
        "Khmer Rouge prison-site count (standardized)",
    ),
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Prison proximity vs bombing density",
        SPI12,
        False,
        "Distance to Nearest Khmer Rouge Prison km",
        "Proximity to nearest prison (−distance; standardized)",
        "negative",
    ),
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Khmer Rouge burial-site count vs bombing density",
        SPI12,
        False,
        "Khmer Rouge Burial Site Count",
        "Khmer Rouge burial-site count (standardized)",
    ),
    RobustnessSpec(
        "Historical conflict",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Burial-site proximity vs bombing density",
        SPI12,
        False,
        "Distance to Nearest Khmer Rouge Burial Site km",
        "Proximity to nearest burial site (−distance; standardized)",
        "negative",
    ),
    RobustnessSpec(
        "Drought window",
        CROP_YIELD,
        "Crop yield (asinh)",
        "SPI-3 vs SPI-12",
        SPI12,
        False,
        alternative_shock="Interview Month SPI 3 Month",
        alternative_shock_label="SPI-3 (higher = wetter; standardized)",
    ),
    RobustnessSpec(
        "Drought window",
        CROP_YIELD,
        "Crop yield (asinh)",
        "SPI-6 vs SPI-12",
        SPI12,
        False,
        alternative_shock="Interview Month SPI 6 Month",
        alternative_shock_label="SPI-6 (higher = wetter; standardized)",
    ),
    RobustnessSpec(
        "Drought window",
        CROP_YIELD,
        "Crop yield (asinh)",
        "SPI-3 drought indicator vs continuous SPI-12",
        SPI12,
        False,
        alternative_shock="Interview Month Drought Shock SPI 3",
        alternative_shock_label="SPI-3 drought indicator (0→1)",
        alternative_shock_binary=True,
    ),
    RobustnessSpec(
        "Drought window",
        CROP_YIELD,
        "Crop yield (asinh)",
        "SPI-6 drought indicator vs continuous SPI-12",
        SPI12,
        False,
        alternative_shock="Interview Month Drought Shock SPI 6",
        alternative_shock_label="SPI-6 drought indicator (0→1)",
        alternative_shock_binary=True,
    ),
    RobustnessSpec(
        "Wet-shock definition",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Annual rainfall anomaly vs annual extreme-wet indicator",
        EXTREME_WET,
        True,
        alternative_shock="Annual Rainfall Anomaly Z (1991-2020)",
        alternative_shock_label="Annual rainfall anomaly (standardized)",
    ),
    RobustnessSpec(
        "Wet-shock definition",
        CROP_YIELD,
        "Crop yield (asinh)",
        "May–October anomaly vs annual extreme-wet indicator",
        EXTREME_WET,
        True,
        alternative_shock="May October Rainfall Anomaly Z (1991-2020)",
        alternative_shock_label="May–October rainfall anomaly (standardized)",
    ),
    RobustnessSpec(
        "Wet-shock definition",
        CROP_YIELD,
        "Crop yield (asinh)",
        "May–October extreme wet vs annual extreme wet",
        EXTREME_WET,
        True,
        alternative_shock="May October Rainfall Extreme Wet Shock",
        alternative_shock_label="May–October extreme-wet indicator (0→1)",
        alternative_shock_binary=True,
    ),
    RobustnessSpec(
        "Food-price definition",
        FOOD_CONSUMPTION,
        "Food consumption per member (asinh)",
        "Wholesale price level vs 12-month wholesale change",
        PRICE_12M,
        False,
        alternative_shock="Local Relative Log Wholesale Rice Price",
        alternative_shock_label="Relative wholesale rice-price level (standardized)",
        agriculture_only=False,
    ),
    RobustnessSpec(
        "Food-price definition",
        FOOD_CONSUMPTION,
        "Food consumption per member (asinh)",
        "Broad retail price level vs 12-month wholesale change",
        PRICE_12M,
        False,
        alternative_shock="Broad Retail Food Local Relative Log Price",
        alternative_shock_label="Broad retail relative-price level (standardized)",
        agriculture_only=False,
    ),
    RobustnessSpec(
        "Food-price definition",
        FOOD_CONSUMPTION,
        "Food consumption per member (asinh)",
        "12-month broad retail change vs wholesale change",
        PRICE_12M,
        False,
        alternative_shock="12 Month Change in Broad Retail Food Local Relative Log Price",
        alternative_shock_label="12-month broad retail price change (standardized)",
        agriculture_only=False,
    ),
    RobustnessSpec(
        "Design sensitivity",
        CROP_YIELD,
        "Crop yield (asinh)",
        "All linkage levels vs commune-only linkage",
        SPI12,
        False,
        alternative_commune_only=False,
        matched_support=False,
    ),
    RobustnessSpec(
        "Design sensitivity",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Unweighted vs survey weighted",
        SPI12,
        False,
        alternative_weighted=False,
    ),
    RobustnessSpec(
        "Satellite validation",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Survey-year inundation vs annual extreme-wet rainfall",
        EXTREME_WET,
        True,
        alternative_shock="Survey Year Maximum Flooded Geography Share",
        alternative_shock_label="Survey-year maximum flooded share (standardized)",
    ),
    RobustnessSpec(
        "Satellite validation",
        CROP_YIELD,
        "Crop yield (asinh)",
        "Preceding-12-month inundation vs annual extreme-wet rainfall",
        EXTREME_WET,
        True,
        alternative_shock="Preceding 12 Month Maximum Flooded Geography Share",
        alternative_shock_label="Preceding-12-month maximum flooded share (standardized)",
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Any bombing indicator vs log bombing density",
        PRICE_12M,
        False,
        "Any US Bombing Record",
        "Any U.S. bombing record (standardized)",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Bombing record intensity vs log bombing-location density",
        PRICE_12M,
        False,
        "Bombing Record Count",
        "Log(1 + bombing record count), standardized",
        "log1p",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Khmer Rouge prison count vs bombing density",
        PRICE_12M,
        False,
        "Khmer Rouge Prison Count",
        "Khmer Rouge prison-site count (standardized)",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Prison proximity vs bombing density",
        PRICE_12M,
        False,
        "Distance to Nearest Khmer Rouge Prison km",
        "Proximity to nearest prison (−distance; standardized)",
        "negative",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Khmer Rouge burial-site count vs bombing density",
        PRICE_12M,
        False,
        "Khmer Rouge Burial Site Count",
        "Khmer Rouge burial-site count (standardized)",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result conflict definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Burial-site proximity vs bombing density",
        PRICE_12M,
        False,
        "Distance to Nearest Khmer Rouge Burial Site km",
        "Proximity to nearest burial site (−distance; standardized)",
        "negative",
        alternative_shock=PRICE_12M,
        alternative_shock_label="12-month local rice-price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result price definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Wholesale price level vs 12-month wholesale change",
        PRICE_12M,
        False,
        alternative_shock="Local Relative Log Wholesale Rice Price",
        alternative_shock_label="Relative wholesale rice-price level (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result price definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "Broad retail price level vs 12-month wholesale change",
        PRICE_12M,
        False,
        alternative_shock="Broad Retail Food Local Relative Log Price",
        alternative_shock_label="Broad retail relative-price level (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
    RobustnessSpec(
        "Key-result price definition",
        ATTENDANCE,
        "School attendance, ages 6–17 (percentage points)",
        "12-month broad retail change vs wholesale change",
        PRICE_12M,
        False,
        alternative_shock="12 Month Change in Broad Retail Food Local Relative Log Price",
        alternative_shock_label="12-month broad retail price change (standardized)",
        agriculture_only=False,
        source="education",
        outcome_transform="percentage_points",
        school_age_only=True,
    ),
]


@dataclass(frozen=True)
class ModelResult:
    estimate: float
    standard_error: float
    sample_size: int
    cluster_count: int

    @property
    def lower(self) -> float:
        return self.estimate - 1.96 * self.standard_error

    @property
    def upper(self) -> float:
        return self.estimate + 1.96 * self.standard_error


def weighted_standardize(values: np.ndarray, weights: np.ndarray) -> np.ndarray:
    mean = float(np.average(values, weights=weights))
    variance = float(np.average((values - mean) ** 2, weights=weights))
    standard_deviation = float(np.sqrt(variance))
    if not np.isfinite(standard_deviation) or standard_deviation <= 0:
        raise ValueError("Cannot standardize a variable with zero or invalid variance")
    return (values - mean) / standard_deviation


def transform_conflict(values: np.ndarray, transform: str) -> np.ndarray:
    if transform == "identity":
        return values
    if transform == "log1p":
        if np.nanmin(values) < 0:
            raise ValueError("log1p conflict measure contains negative values")
        return np.log1p(values)
    if transform == "negative":
        return -values
    raise ValueError(f"Unknown conflict transform: {transform}")


def prepare_sample(
    frame: pd.DataFrame,
    spec: RobustnessSpec,
    *,
    commune_only: bool,
    paired: bool,
) -> pd.DataFrame:
    weight = WEIGHT if spec.source == "household" else PERSON_WEIGHT
    province = PROVINCE if spec.source == "household" else EDUCATION_PROVINCE
    controls = (
        [HOUSEHOLD_SIZE]
        if spec.source == "household"
        else [HOUSEHOLD_SIZE, AGE, FEMALE]
    )
    required = [
        spec.outcome,
        CORE_CONFLICT,
        spec.core_shock,
        weight,
        RESOLUTION,
        GEOGRAPHY,
        province,
        YEAR,
        *controls,
    ]
    if paired:
        required.extend([spec.alternative_conflict, spec.alternative_shock])
    if spec.agriculture_only:
        required.append(AGRICULTURAL_HOUSEHOLD)
    required = list(dict.fromkeys(required))
    sample = frame[required].copy()
    if commune_only:
        sample = sample.loc[sample[RESOLUTION].eq("commune")]
    if spec.agriculture_only:
        sample = sample.loc[sample[AGRICULTURAL_HOUSEHOLD].eq(1)]
    sample = sample.dropna(subset=required).copy()
    sample = sample.loc[sample[weight].gt(0)].copy()
    if spec.school_age_only:
        sample = sample.loc[sample[AGE].between(6, 17)].copy()
    return sample


def fit_model(
    sample: pd.DataFrame,
    *,
    outcome: str,
    outcome_transform: str,
    conflict: str,
    conflict_transform: str,
    shock: str,
    shock_binary: bool,
    weighted: bool,
    weight_column: str,
    province_column: str,
    controls: list[str],
) -> ModelResult:
    model_weights = (
        sample[weight_column].to_numpy(dtype=float)
        if weighted
        else np.ones(len(sample), dtype=float)
    )
    conflict_values = transform_conflict(
        sample[conflict].to_numpy(dtype=float), conflict_transform
    )
    conflict_z = weighted_standardize(conflict_values, model_weights)
    shock_values = sample[shock].to_numpy(dtype=float)
    if shock_binary:
        if not pd.Series(shock_values).isin([0, 1]).all():
            raise ValueError(f"Binary shock contains values outside 0/1: {shock}")
        shock_model = shock_values
    else:
        shock_model = weighted_standardize(shock_values, model_weights)

    outcome_values = sample[outcome].to_numpy(dtype=float)
    if outcome_transform == "asinh":
        if np.nanmin(outcome_values) < 0:
            raise ValueError("Outcome unexpectedly contains negative values")
        model_outcome = np.arcsinh(outcome_values)
    elif outcome_transform == "percentage_points":
        if np.nanmin(outcome_values) < 0 or np.nanmax(outcome_values) > 1:
            raise ValueError("Binary outcome falls outside [0, 1]")
        model_outcome = 100.0 * outcome_values
    else:
        raise ValueError(f"Unknown outcome transform: {outcome_transform}")

    exogenous = pd.DataFrame(
        {
            "shock": shock_model,
            "conflict_x_shock": conflict_z * shock_model,
        },
        index=sample.index,
    )
    for control in controls:
        exogenous[control] = sample[control].to_numpy(dtype=float)
    geography = sample[RESOLUTION].astype(str) + ":" + sample[GEOGRAPHY].astype(str)
    province_wave = (
        sample[province_column].astype(str)
        + "_"
        + sample[YEAR].astype(int).astype(str)
    )
    absorbed = pd.DataFrame(
        {
            "geography": geography.astype("category"),
            "province_wave": province_wave.astype("category"),
        },
        index=sample.index,
    )
    model_arguments: dict[str, object] = {
        "dependent": pd.Series(model_outcome, index=sample.index, name="outcome"),
        "exog": exogenous,
        "absorb": absorbed,
        "drop_absorbed": True,
    }
    if weighted:
        model_arguments["weights"] = pd.Series(model_weights, index=sample.index)
    fitted = AbsorbingLS(**model_arguments).fit(
        cov_type="clustered",
        clusters=pd.DataFrame({"geography": geography}, index=sample.index),
        debiased=True,
    )
    if "conflict_x_shock" not in fitted.params.index:
        raise ValueError(f"Interaction was absorbed: {conflict} × {shock}")
    return ModelResult(
        estimate=float(fitted.params["conflict_x_shock"]),
        standard_error=float(fitted.std_errors["conflict_x_shock"]),
        sample_size=len(sample),
        cluster_count=int(geography.nunique()),
    )


def estimate_spec(
    households: pd.DataFrame,
    education: pd.DataFrame,
    spec: RobustnessSpec,
) -> tuple[ModelResult, ModelResult]:
    frame = households if spec.source == "household" else education
    weight = WEIGHT if spec.source == "household" else PERSON_WEIGHT
    province = PROVINCE if spec.source == "household" else EDUCATION_PROVINCE
    controls = (
        [HOUSEHOLD_SIZE]
        if spec.source == "household"
        else [HOUSEHOLD_SIZE, AGE, FEMALE]
    )
    if spec.matched_support:
        sample = prepare_sample(
            frame, spec, commune_only=True, paired=True
        )
        core_sample = sample
        alternative_sample = sample
    else:
        core_sample = prepare_sample(
            frame, spec, commune_only=True, paired=False
        )
        alternative_sample = prepare_sample(
            frame,
            spec,
            commune_only=spec.alternative_commune_only,
            paired=False,
        )

    core = fit_model(
        core_sample,
        outcome=spec.outcome,
        outcome_transform=spec.outcome_transform,
        conflict=CORE_CONFLICT,
        conflict_transform="identity",
        shock=spec.core_shock,
        shock_binary=spec.core_shock_binary,
        weighted=True,
        weight_column=weight,
        province_column=province,
        controls=controls,
    )
    alternative = fit_model(
        alternative_sample,
        outcome=spec.outcome,
        outcome_transform=spec.outcome_transform,
        conflict=spec.alternative_conflict,
        conflict_transform=spec.alternative_conflict_transform,
        shock=spec.alternative_shock,
        shock_binary=spec.alternative_shock_binary,
        weighted=spec.alternative_weighted,
        weight_column=weight,
        province_column=province,
        controls=controls,
    )
    return core, alternative


def build_table(
    households: pd.DataFrame,
    education: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for spec in SPECS:
        core, alternative = estimate_spec(households, education, spec)
        linkage = "Commune" if spec.alternative_commune_only else "All linkages"
        weighting = "weighted" if spec.alternative_weighted else "unweighted"
        support = "matched support" if spec.matched_support else "sample differs by design"
        rows.append(
            {
                "Robustness family": spec.family,
                "Outcome (model scale)": spec.outcome_label,
                "Comparison (alternative vs core)": spec.comparison,
                "Core estimate": core.estimate,
                "Alternative estimate": alternative.estimate,
                "Alternative 95% CI": f"[{alternative.lower:.3f}, {alternative.upper:.3f}]",
                "Alternative N": alternative.sample_size,
                "Conflict definition": spec.alternative_conflict_label,
                "Shock definition": spec.alternative_shock_label,
                "Specification indicators": (
                    f"{linkage}; {weighting}; {support}; geography + province×wave FE; "
                    f"geography-clustered SE ({alternative.cluster_count} clusters)"
                ),
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alternative Definitions"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D2"

    for column_index, column in enumerate(frame.columns, start=1):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(
        displayName="AlternativeDefinitionsTable", ref=f"A1:J{last_row}"
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    family_fills = {
        "Historical conflict": PatternFill("solid", fgColor="EAF2F8"),
        "Drought window": PatternFill("solid", fgColor="FFF2CC"),
        "Wet-shock definition": PatternFill("solid", fgColor="DDEBF7"),
        "Food-price definition": PatternFill("solid", fgColor="E2F0D9"),
        "Design sensitivity": PatternFill("solid", fgColor="F2F2F2"),
        "Satellite validation": PatternFill("solid", fgColor="E4DFEC"),
        "Key-result conflict definition": PatternFill("solid", fgColor="FCE4D6"),
        "Key-result price definition": PatternFill("solid", fgColor="DDEBF7"),
    }
    white_bold = Font(color="FFFFFF", bold=True, size=8.5)
    navy_bold = Font(color="1F4E78", bold=True, size=8.5)
    light_rule = Side(style="thin", color="C8D4DF")
    section_rule = Side(style="medium", color="7F9DB9")

    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=section_rule)
    sheet.row_dimensions[1].height = 50

    section_start_rows = {2, 8, 12, 15, 18, 20, 22, 28}
    for row_index in range(2, last_row + 1):
        family = str(sheet.cell(row=row_index, column=1).value)
        for column_index in range(1, 11):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                horizontal="right" if column_index in {4, 5, 7} else "left",
                vertical="center",
                wrap_text=column_index not in {4, 5, 7},
            )
            cell.border = Border(
                top=section_rule if row_index in section_start_rows else None,
                bottom=light_rule,
            )
        sheet.cell(row=row_index, column=1).fill = family_fills[family]
        sheet.cell(row=row_index, column=1).font = navy_bold
        sheet.row_dimensions[row_index].height = 42

    for column_letter in ("D", "E"):
        for cell in sheet[column_letter][1:]:
            cell.number_format = "0.000"
    for cell in sheet["G"][1:]:
        cell.number_format = "#,##0"

    widths = [24, 27, 45, 15, 18, 21, 15, 39, 41, 54]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = f"A1:J{last_row}"
    sheet.sheet_view.zoomScale = 65
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:J{last_row}"
    sheet.page_margins.left = 0.12
    sheet.page_margins.right = 0.12
    sheet.page_margins.top = 0.16
    sheet.page_margins.bottom = 0.16

    workbook.properties.title = "Alternative Conflict Measures and Shock Definitions"
    workbook.properties.subject = "Direction 3 coverage-matched robustness specifications"
    workbook.properties.creator = "Mike Li"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (29, 10), frame.shape
    assert list(frame.columns) == COLUMNS
    assert not frame[
        [
            "Robustness family",
            "Outcome (model scale)",
            "Comparison (alternative vs core)",
        ]
    ].duplicated().any()
    assert frame["Alternative N"].gt(0).all()
    assert np.isfinite(frame[["Core estimate", "Alternative estimate"]]).all().all()
    assert frame["Alternative 95% CI"].str.match(r"^\[-?\d+\.\d{3}, -?\d+\.\d{3}\]$").all()
    assert set(frame["Robustness family"]) == {
        "Historical conflict",
        "Drought window",
        "Wet-shock definition",
        "Food-price definition",
        "Design sensitivity",
        "Satellite validation",
        "Key-result conflict definition",
        "Key-result price definition",
    }

    workbook = load_workbook(OUTPUT_PATH, data_only=False)
    assert workbook.sheetnames == ["Alternative Definitions"]
    sheet = workbook["Alternative Definitions"]
    assert sheet.max_row == 30
    assert sheet.max_column == 10
    assert list(sheet.tables) == ["AlternativeDefinitionsTable"]
    assert sheet.tables["AlternativeDefinitionsTable"].ref == "A1:J30"
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                )


def main() -> None:
    household_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            PROVINCE,
            WEIGHT,
            HOUSEHOLD_SIZE,
            AGRICULTURAL_HOUSEHOLD,
            CORE_CONFLICT,
            *[spec.outcome for spec in SPECS if spec.source == "household"],
            *[spec.core_shock for spec in SPECS if spec.source == "household"],
            *[spec.alternative_conflict for spec in SPECS if spec.source == "household"],
            *[spec.alternative_shock for spec in SPECS if spec.source == "household"],
        }
    )
    education_columns = sorted(
        {
            YEAR,
            RESOLUTION,
            GEOGRAPHY,
            EDUCATION_PROVINCE,
            PERSON_WEIGHT,
            HOUSEHOLD_SIZE,
            AGE,
            FEMALE,
            CORE_CONFLICT,
            *[spec.outcome for spec in SPECS if spec.source == "education"],
            *[spec.core_shock for spec in SPECS if spec.source == "education"],
            *[spec.alternative_conflict for spec in SPECS if spec.source == "education"],
            *[spec.alternative_shock for spec in SPECS if spec.source == "education"],
        }
    )
    households = pd.read_parquet(INPUT_PATH, columns=household_columns)
    education = pd.read_parquet(EDUCATION_PATH, columns=education_columns)
    table = build_table(households, education)
    write_workbook(table)
    validate_output(table)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Dimensions: {table.shape[0]} rows x {table.shape[1]} columns")
    print(table.groupby("Robustness family").size().to_string())


if __name__ == "__main__":
    main()
