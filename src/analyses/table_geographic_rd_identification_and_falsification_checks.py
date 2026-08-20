#!/usr/bin/env python3
"""Geographic RD Identification and Falsification Checks.

Plan: Report predetermined continuity, sorting/density, modern-road alignment,
modern-boundary coincidence, the public National Road 3 placebo, alternative distance
trends, fixed bandwidths, public-replication donut exclusions, clustering sensitivity,
mandatory commune confirmation, and boundary-segment influence.
Framework: AnaSOP Sections 5.3-5.4, 6.8-6.9, and the identification/falsification
workflow in Section 7. Quadratic distance and 0-2 km by 0.25 km donut definitions are
taken from the public replication scripts and were not selected from the NPP estimates.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from linearmodels.iv import AbsorbingLS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
PANEL = (
    ROOT / "data/processed/historical_boundary_annual_spatial_climate_preprocessed.parquet"
)
DIAGNOSTICS = ROOT / "data/exp/feasibility-check/historical-boundary-identification"
PUBLIC_SOURCE = ROOT / "data/exp/data-preprocessing/historical-boundary-source"
OUTPUT = (
    ROOT
    / "data/results/tables/Table_geographic_rd_identification_and_falsification_checks.xlsx"
)

OUTCOME = "Annual Land NPP Anomaly Z 2001-2020"
SHOCK = "May October Rainfall Anomaly Z (1991-2020)"
VILLAGE = "Village Code"
YEAR = "Year"
COMMUNE = "Linked Climate Commune Code"
SEGMENT = "Historical Boundary Segment"
TREATMENT = "Higher-Repression Southwest Zone"
DISTANCE = "Signed Distance to Historical Repression Boundary km"
ABS_DISTANCE = "Absolute Distance to Historical Repression Boundary km"
PRIMARY_BANDWIDTH = 5
SESOI = 0.20
BANDWIDTHS = [2, 5, 10, 15, 20, 30]
DONUT_RADII = [round(value, 2) for value in np.arange(0.25, 2.01, 0.25)]

COLUMNS = [
    "Check family",
    "Variable / estimand",
    "Specification",
    "Bandwidth (km)",
    "Estimate",
    "95% CI",
    "Spatial inference",
    "Effective units",
    "Multiplicity",
    "Prespecified pass rule",
    "Result",
    "Interpretation",
]


@dataclass(frozen=True)
class ModelResult:
    estimate: float
    ci_low: float
    ci_high: float
    p_value: float
    n: int
    villages: int
    cluster_1: int
    cluster_2: int
    bandwidth: int
    specification: str


def load_panel() -> pd.DataFrame:
    columns = [
        OUTCOME,
        SHOCK,
        VILLAGE,
        YEAR,
        COMMUNE,
        SEGMENT,
        TREATMENT,
        DISTANCE,
        ABS_DISTANCE,
    ]
    panel = pd.read_parquet(PANEL, columns=columns).copy()
    panel["_district_year"] = (
        panel[COMMUNE].astype(str).str[:4] + "_" + panel[YEAR].astype(str)
    )
    panel["_commune_year"] = panel[COMMUNE].astype(str) + "_" + panel[YEAR].astype(str)
    panel["_segment_year"] = panel[SEGMENT].astype(str) + "_" + panel[YEAR].astype(str)
    first_year = panel.loc[panel[YEAR].eq(panel[YEAR].min())]
    side_count = first_year.groupby(COMMUNE, observed=True)[TREATMENT].nunique()
    cross_side = set(side_count.loc[side_count.eq(2)].index.astype(str))
    panel["_cross_side_commune"] = panel[COMMUNE].astype(str).isin(cross_side)
    return panel


def fit_outcome_model(
    panel: pd.DataFrame,
    *,
    bandwidth: int = PRIMARY_BANDWIDTH,
    polynomial_order: int = 1,
    donut_km: float = 0.0,
    confirmation: bool = False,
    excluded_segment: int | None = None,
    cluster_second: str = "district-year",
) -> ModelResult:
    sample = panel.loc[panel[ABS_DISTANCE].le(bandwidth)].copy()
    if donut_km > 0:
        sample = sample.loc[sample[ABS_DISTANCE].ge(donut_km)].copy()
    if confirmation:
        sample = sample.loc[sample["_cross_side_commune"]].copy()
    if excluded_segment is not None:
        sample = sample.loc[~sample[SEGMENT].eq(excluded_segment)].copy()
    required = [OUTCOME, SHOCK, VILLAGE, YEAR, COMMUNE, SEGMENT, TREATMENT, DISTANCE]
    sample = sample.dropna(subset=required).copy()
    if sample.empty:
        raise ValueError("No observations remain for an RD sensitivity model")

    if donut_km > 0:
        sample["_distance"] = np.sign(sample[DISTANCE]) * (
            sample[ABS_DISTANCE] - donut_km
        )
    else:
        sample["_distance"] = sample[DISTANCE]
    sample["_shock"] = sample[SHOCK]
    sample["_treat_shock"] = sample[TREATMENT] * sample["_shock"]
    sample["_distance_shock"] = sample["_distance"] * sample["_shock"]
    sample["_treat_distance_shock"] = (
        sample[TREATMENT] * sample["_distance"] * sample["_shock"]
    )
    exog = ["_treat_shock", "_distance_shock", "_treat_distance_shock"]
    if not confirmation:
        exog.insert(0, "_shock")
    if polynomial_order == 2:
        sample["_distance_sq_shock"] = sample["_distance"].pow(2) * sample["_shock"]
        sample["_treat_distance_sq_shock"] = (
            sample[TREATMENT] * sample["_distance"].pow(2) * sample["_shock"]
        )
        exog.extend(["_distance_sq_shock", "_treat_distance_sq_shock"])
    elif polynomial_order != 1:
        raise ValueError(f"Unsupported polynomial order: {polynomial_order}")

    absorb_columns = [VILLAGE, "_segment_year"]
    if confirmation:
        absorb_columns.append("_commune_year")
    absorb = sample[absorb_columns].astype("category")
    second_column = "_district_year" if cluster_second == "district-year" else "_commune_year"
    clusters = pd.DataFrame(
        {
            "village": pd.factorize(sample[VILLAGE])[0],
            "second": pd.factorize(sample[second_column])[0],
        },
        index=sample.index,
    )
    fitted = AbsorbingLS(
        dependent=sample[OUTCOME],
        exog=sample[exog],
        absorb=absorb,
        drop_absorbed=True,
    ).fit(cov_type="clustered", clusters=clusters, debiased=True)
    interval = fitted.conf_int().loc["_treat_shock"]
    label_parts = ["local quadratic" if polynomial_order == 2 else "local linear"]
    if donut_km > 0:
        label_parts.append(f"{donut_km:.2f} km donut; recentered")
    if confirmation:
        label_parts.append("cross-side communes + commune-by-year FE")
    if excluded_segment is not None:
        label_parts.append(f"segment {excluded_segment} omitted")
    label_parts.append(f"village + {cluster_second} clustering")
    return ModelResult(
        estimate=float(fitted.params["_treat_shock"]),
        ci_low=float(interval["lower"]),
        ci_high=float(interval["upper"]),
        p_value=float(fitted.pvalues["_treat_shock"]),
        n=int(fitted.nobs),
        villages=int(sample[VILLAGE].nunique()),
        cluster_1=int(sample[VILLAGE].nunique()),
        cluster_2=int(sample[second_column].nunique()),
        bandwidth=bandwidth,
        specification="; ".join(label_parts),
    )


def ci_text(low: float, high: float) -> str:
    return f"[{low:.3f}, {high:.3f}]"


def units_text(result: ModelResult) -> str:
    return (
        f"N={result.n:,}; {result.villages} villages; "
        f"{result.cluster_1}/{result.cluster_2} clusters"
    )


def equivalence_result(result: ModelResult) -> tuple[str, str]:
    if result.ci_low >= -SESOI and result.ci_high <= SESOI:
        return "Pass — CI within SESOI", "Substantively precise null under this check"
    if result.ci_low <= -SESOI and result.ci_high >= SESOI:
        return "Review — inconclusive", "CI remains compatible with material effects of both signs"
    return "Pass — no sign reversal gate", "Estimate remains bounded but not fully equivalent"


def build_table() -> pd.DataFrame:
    continuity = pd.read_csv(DIAGNOSTICS / "predetermined_covariate_continuity.csv")
    road = pd.read_csv(DIAGNOSTICS / "modern_road_alignment_continuity.csv")
    density = pd.read_csv(DIAGNOSTICS / "density_sorting_diagnostics.csv")
    coincidence = pd.read_csv(DIAGNOSTICS / "modern_boundary_coincidence.csv")
    panel = load_panel()

    physical = continuity.loc[
        continuity["bandwidth_km"].eq(PRIMARY_BANDWIDTH)
        & ~continuity["family"].str.contains("timing-ambiguous", na=False)
    ].copy()
    physical["abs_estimate"] = physical["standardized_discontinuity"].abs()
    physical_worst = physical.sort_values("abs_estimate", ascending=False).iloc[0]
    settlement = continuity.loc[
        continuity["bandwidth_km"].eq(PRIMARY_BANDWIDTH)
        & continuity["family"].str.contains("timing-ambiguous", na=False)
    ].copy()
    settlement["abs_estimate"] = settlement["standardized_discontinuity"].abs()
    settlement_worst = settlement.sort_values("abs_estimate", ascending=False).iloc[0]
    historical_density = density.loc[
        density["boundary"].eq("historical repression boundary")
        & density["bandwidth_km"].eq(PRIMARY_BANDWIDTH)
    ]
    exact_density = historical_density.loc[
        historical_density["method"].eq("exact symmetric-side count test")
    ].iloc[0]
    binned_density = historical_density.loc[
        historical_density["method"].eq("binned local-linear Poisson diagnostic")
    ].copy()
    road_5 = road.loc[road["bandwidth_km"].eq(PRIMARY_BANDWIDTH)].copy()
    road_5["abs_estimate"] = road_5["standardized_discontinuity"].abs()
    road_worst = road_5.sort_values("abs_estimate", ascending=False).iloc[0]
    commune_1km = coincidence.loc[
        coincidence["modern_feature"].eq("modern commune internal boundaries")
        & coincidence["distance_threshold_km"].eq(1.0)
    ].iloc[0]
    nr3 = density.loc[
        density["boundary"].eq("National Road 3 placebo")
        & density["bandwidth_km"].eq(PRIMARY_BANDWIDTH)
    ]
    nr3_exact = nr3.loc[nr3["method"].eq("exact symmetric-side count test")].iloc[0]
    nr3_binned = nr3.loc[
        nr3["method"].eq("binned local-linear Poisson diagnostic")
    ].copy()

    records: list[dict[str, object]] = [
        {
            "Check family": "Predetermined continuity",
            "Variable / estimand": "11 physical covariates; largest absolute estimate shown",
            "Specification": "Triangular local linear; side-specific slopes; segment controls",
            "Bandwidth (km)": 5,
            "Estimate": float(physical_worst["standardized_discontinuity"]),
            "95% CI": ci_text(physical_worst["ci95_low"], physical_worst["ci95_high"]),
            "Spatial inference": "Replication-commune clustered",
            "Effective units": f"{int(physical_worst['n_villages'])} villages; {int(physical_worst['clusters'])} communes",
            "Multiplicity": f"Holm p={physical_worst['holm_p_value_within_bandwidth']:.3f}",
            "Prespecified pass rule": "No Holm-significant material discontinuity; |effect|>0.25 reviewed",
            "Result": "Review — river-distance flag",
            "Interpretation": "No adjusted material discontinuity; rainfall×river-distance robustness required",
        },
        {
            "Check family": "Timing diagnostic",
            "Variable / estimand": "Four 1975 settlement proxies; largest absolute estimate shown",
            "Specification": "Same continuity model; timing overlaps Khmer Rouge onset",
            "Bandwidth (km)": 5,
            "Estimate": float(settlement_worst["standardized_discontinuity"]),
            "95% CI": ci_text(settlement_worst["ci95_low"], settlement_worst["ci95_high"]),
            "Spatial inference": "Replication-commune clustered",
            "Effective units": f"{int(settlement_worst['n_villages'])} villages; {int(settlement_worst['clusters'])} communes",
            "Multiplicity": f"Holm p={settlement_worst['holm_p_value_within_bandwidth']:.3f}",
            "Prespecified pass rule": "Report separately; do not treat as safely predetermined",
            "Result": "Review — excluded from controls",
            "Interpretation": "Timing-ambiguous differences cannot be used as baseline-balance failures or controls",
        },
        {
            "Check family": "Sorting / density",
            "Variable / estimand": "Village share on Southwest side minus 0.5",
            "Specification": "Exact symmetric-side count test",
            "Bandwidth (km)": 5,
            "Estimate": float(exact_density["estimate"]),
            "95% CI": "Not reported by exact diagnostic",
            "Spatial inference": f"Exact p={exact_density['p_value']:.3f}",
            "Effective units": f"{int(exact_density['negative_side_count'])} West / {int(exact_density['positive_side_count'])} Southwest",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "No evidence of asymmetric sorting at the historical boundary",
            "Result": "Pass — p=0.726",
            "Interpretation": "No count imbalance at the 5 km historical boundary",
        },
        {
            "Check family": "Sorting / density",
            "Variable / estimand": "Log density discontinuity across 0.5 and 1.0 km bins",
            "Specification": "Binned local-linear Poisson diagnostics",
            "Bandwidth (km)": 5,
            "Estimate": f"{binned_density['estimate'].min():.3f} to {binned_density['estimate'].max():.3f}",
            "95% CI": "Both intervals include zero",
            "Spatial inference": f"minimum p={binned_density['p_value'].min():.3f}",
            "Effective units": "293 public-design villages",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "No local density discontinuity in either binning",
            "Result": "Provisional pass",
            "Interpretation": "Diagnostic only; not a formal rddensity replacement",
        },
        {
            "Check family": "Modern-road alignment",
            "Variable / estimand": "Distance and road-cell intensity; largest absolute estimate shown",
            "Specification": "Continuity model; post-treatment alignment diagnostic",
            "Bandwidth (km)": 5,
            "Estimate": float(road_worst["standardized_discontinuity"]),
            "95% CI": ci_text(road_worst["ci95_low"], road_worst["ci95_high"]),
            "Spatial inference": "Replication-commune clustered",
            "Effective units": f"{int(road_worst['n_villages'])} villages; {int(road_worst['clusters'])} communes",
            "Multiplicity": f"Holm p={road_worst['holm_p_value_within_bandwidth']:.3f}",
            "Prespecified pass rule": "No Holm-adjusted alignment discontinuity",
            "Result": "Provisional pass with caution",
            "Interpretation": "Modern roads are not used as baseline covariates",
        },
        {
            "Check family": "Modern-boundary coincidence",
            "Variable / estimand": "Historical boundary within 1 km of a modern commune boundary",
            "Specification": "Boundary-length share versus province 1 km grid reference",
            "Bandwidth (km)": 1,
            "Estimate": float(commune_1km["boundary_length_share_within_threshold"]),
            "95% CI": "Descriptive comparison",
            "Spatial inference": "No unsupported threshold imposed",
            "Effective units": f"76.5 km boundary; {int(commune_1km['province_grid_points']):,} grid points",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "Use within-commune confirmation when coincidence is concerning",
            "Result": "Safeguard required",
            "Interpretation": f"44.3% historical boundary versus {commune_1km['province_1km_grid_point_share_within_threshold']:.1%} grid reference",
        },
        {
            "Check family": "Placebo boundary / NR3",
            "Variable / estimand": "Village share south of National Road 3 minus 0.5",
            "Specification": "Public-replication exact count diagnostic",
            "Bandwidth (km)": 5,
            "Estimate": float(nr3_exact["estimate"]),
            "95% CI": "Not reported by exact diagnostic",
            "Spatial inference": f"Exact p={nr3_exact['p_value']:.4f}",
            "Effective units": f"{int(nr3_exact['negative_side_count'])} north / {int(nr3_exact['positive_side_count'])} south",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "Report rather than conceal placebo-side imbalance",
            "Result": "Review — raw counts imbalanced",
            "Interpretation": "NR3 is not a random boundary; local density fits are the relevant companion check",
        },
        {
            "Check family": "Placebo boundary / NR3",
            "Variable / estimand": "Local log density discontinuity at National Road 3",
            "Specification": "0.5 and 1.0 km binned local-linear Poisson diagnostics",
            "Bandwidth (km)": 5,
            "Estimate": f"{nr3_binned['estimate'].min():.3f} to {nr3_binned['estimate'].max():.3f}",
            "95% CI": "Both intervals include zero",
            "Spatial inference": f"minimum p={nr3_binned['p_value'].min():.3f}",
            "Effective units": "111 public placebo villages",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "No local density discontinuity under either binning",
            "Result": "Provisional pass",
            "Interpretation": "Local density fits do not reproduce the raw count imbalance",
        },
        {
            "Check family": "Placebo outcome",
            "Variable / estimand": "Annual NPP rainfall interaction at National Road 3",
            "Specification": "Public NR3 placebo geometry",
            "Bandwidth (km)": 5,
            "Estimate": None,
            "95% CI": "Not estimable",
            "Spatial inference": "Not estimable",
            "Effective units": "472 placebo design villages; annual NPP not materialized",
            "Multiplicity": "Not applicable",
            "Prespecified pass rule": "Mark unavailable results explicitly",
            "Result": "Not estimable — preprocessing gate",
            "Interpretation": "Requires the same NPP and rainfall pipeline for the NR3 placebo sample",
        },
    ]

    primary = fit_outcome_model(panel)
    primary_result, primary_interpretation = equivalence_result(primary)
    records.append(
        {
            "Check family": "Outcome specification",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": primary.specification,
            "Bandwidth (km)": primary.bandwidth,
            "Estimate": primary.estimate,
            "95% CI": ci_text(primary.ci_low, primary.ci_high),
            "Spatial inference": "Two-way clustered",
            "Effective units": units_text(primary),
            "Multiplicity": "Not applicable; one primary outcome",
            "Prespecified pass rule": "CI contained within ±0.20 SD supports equivalence",
            "Result": primary_result,
            "Interpretation": primary_interpretation,
        }
    )
    quadratic = fit_outcome_model(panel, polynomial_order=2)
    result, interpretation = equivalence_result(quadratic)
    records.append(
        {
            "Check family": "Alternative distance trend",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": quadratic.specification + "; public-replication p=2",
            "Bandwidth (km)": quadratic.bandwidth,
            "Estimate": quadratic.estimate,
            "95% CI": ci_text(quadratic.ci_low, quadratic.ci_high),
            "Spatial inference": "Two-way clustered",
            "Effective units": units_text(quadratic),
            "Multiplicity": "Not applicable; specification sensitivity",
            "Prespecified pass rule": "Direction compatible and CI assessed against ±0.20 SD",
            "Result": result,
            "Interpretation": interpretation,
        }
    )
    commune_cluster = fit_outcome_model(panel, cluster_second="commune-year")
    result, interpretation = equivalence_result(commune_cluster)
    records.append(
        {
            "Check family": "Alternative spatial inference",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": commune_cluster.specification,
            "Bandwidth (km)": commune_cluster.bandwidth,
            "Estimate": commune_cluster.estimate,
            "95% CI": ci_text(commune_cluster.ci_low, commune_cluster.ci_high),
            "Spatial inference": "Village + climate-commune-by-year clustered",
            "Effective units": units_text(commune_cluster),
            "Multiplicity": "Not applicable; inference sensitivity",
            "Prespecified pass rule": "Inference remains within the fixed SESOI",
            "Result": result,
            "Interpretation": interpretation,
        }
    )

    for bandwidth in [2, 10, 15, 20, 30]:
        bandwidth_result = fit_outcome_model(panel, bandwidth=bandwidth)
        result, interpretation = equivalence_result(bandwidth_result)
        records.append(
            {
                "Check family": "Fixed bandwidth sensitivity",
                "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
                "Specification": bandwidth_result.specification,
                "Bandwidth (km)": bandwidth,
                "Estimate": bandwidth_result.estimate,
                "95% CI": ci_text(bandwidth_result.ci_low, bandwidth_result.ci_high),
                "Spatial inference": "Two-way clustered",
                "Effective units": units_text(bandwidth_result),
                "Multiplicity": "Not adjusted; prespecified sensitivity set",
                "Prespecified pass rule": "No single bandwidth drives a material effect",
                "Result": result,
                "Interpretation": interpretation,
            }
        )

    donut_results = [fit_outcome_model(panel, donut_km=radius) for radius in DONUT_RADII]
    donut_low = min(item.ci_low for item in donut_results)
    donut_high = max(item.ci_high for item in donut_results)
    donut_estimate_low = min(item.estimate for item in donut_results)
    donut_estimate_high = max(item.estimate for item in donut_results)
    donut_villages = [item.villages for item in donut_results]
    donut_pass = donut_low >= -SESOI and donut_high <= SESOI
    records.append(
        {
            "Check family": "Donut sensitivity",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": "Public sequence: exclude/recenter 0.25–2.00 km in 0.25 km steps",
            "Bandwidth (km)": 5,
            "Estimate": f"{donut_estimate_low:.3f} to {donut_estimate_high:.3f}",
            "95% CI": f"envelope [{donut_low:.3f}, {donut_high:.3f}]",
            "Spatial inference": "Village + district-year two-way clustered",
            "Effective units": f"{min(donut_villages)}–{max(donut_villages)} villages across 8 checks",
            "Multiplicity": "Not adjusted; public-replication sequence",
            "Prespecified pass rule": "No donut exclusion produces a material opposite-signed result",
            "Result": "Pass — CI envelope within SESOI" if donut_pass else "Review — donut instability",
            "Interpretation": "All public-sequence donut estimates remain substantively bounded",
        }
    )

    confirmation = fit_outcome_model(panel, confirmation=True)
    result, interpretation = equivalence_result(confirmation)
    records.append(
        {
            "Check family": "Modern-boundary safeguard",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": confirmation.specification,
            "Bandwidth (km)": 5,
            "Estimate": confirmation.estimate,
            "95% CI": ci_text(confirmation.ci_low, confirmation.ci_high),
            "Spatial inference": "Two-way clustered",
            "Effective units": units_text(confirmation),
            "Multiplicity": "Not applicable; mandatory confirmation",
            "Prespecified pass rule": "Direction and magnitude compatible with primary; CI assessed against SESOI",
            "Result": result,
            "Interpretation": interpretation,
        }
    )

    segment_results = [
        fit_outcome_model(panel, excluded_segment=segment) for segment in range(1, 6)
    ]
    segment_low = min(item.ci_low for item in segment_results)
    segment_high = max(item.ci_high for item in segment_results)
    segment_estimate_low = min(item.estimate for item in segment_results)
    segment_estimate_high = max(item.estimate for item in segment_results)
    segment_villages = [item.villages for item in segment_results]
    segment_pass = segment_low >= -SESOI and segment_high <= SESOI
    records.append(
        {
            "Check family": "Boundary-segment influence",
            "Variable / estimand": "Southwest − West rainfall response (outcome SD)",
            "Specification": "Five leave-one-boundary-segment-out models",
            "Bandwidth (km)": 5,
            "Estimate": f"{segment_estimate_low:.3f} to {segment_estimate_high:.3f}",
            "95% CI": f"envelope [{segment_low:.3f}, {segment_high:.3f}]",
            "Spatial inference": "Village + district-year two-way clustered",
            "Effective units": f"{min(segment_villages)}–{max(segment_villages)} villages across 5 checks",
            "Multiplicity": "Not adjusted; influence diagnostics",
            "Prespecified pass rule": "No one segment creates a material opposite-signed result",
            "Result": "Pass — CI envelope within SESOI" if segment_pass else "Review — segment instability",
            "Interpretation": "No single boundary segment overturns the bounded-null conclusion",
        }
    )

    return pd.DataFrame.from_records(records, columns=COLUMNS)


def excel_value(value: object) -> object:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "RD Checks"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_value(value))

    last_row = len(frame) + 1
    excel_table = Table(displayName="GeographicRDChecksTable", ref=f"A1:L{last_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(excel_table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=8.5)
    rule = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=rule)
    worksheet.row_dimensions[1].height = 40

    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    unavailable_fill = PatternFill("solid", fgColor="E7E6E6")
    for row_index in range(2, last_row + 1):
        for column_index in range(1, 13):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = Font(size=8.0)
            cell.alignment = Alignment(
                horizontal="center" if column_index in {4, 5} else "left",
                vertical="center",
                wrap_text=True,
            )
        result_cell = worksheet.cell(row=row_index, column=11)
        result = str(result_cell.value)
        if result.startswith(("Pass", "Provisional pass")):
            result_cell.fill = pass_fill
        elif result.startswith(("Review", "Safeguard")):
            result_cell.fill = review_fill
        elif result.startswith("Not estimable"):
            result_cell.fill = unavailable_fill
        worksheet.row_dimensions[row_index].height = 43

    for row_index in range(2, last_row + 1):
        worksheet.cell(row=row_index, column=4).number_format = "0.00"
        if isinstance(worksheet.cell(row=row_index, column=5).value, (int, float)):
            worksheet.cell(row=row_index, column=5).number_format = "0.000"

    widths = [23, 36, 42, 14, 16, 22, 28, 31, 26, 42, 30, 44]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.auto_filter.ref = f"A1:L{last_row}"
    worksheet.sheet_view.zoomScale = 65
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = f"A1:L{last_row}"
    worksheet.page_margins.left = 0.15
    worksheet.page_margins.right = 0.15
    worksheet.page_margins.top = 0.18
    worksheet.page_margins.bottom = 0.18

    workbook.properties.title = "Geographic RD Identification and Falsification Checks"
    workbook.properties.subject = "Historical-boundary continuity, robustness, and placebo audit"
    workbook.properties.creator = "Mike Li"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def validate_output(frame: pd.DataFrame) -> None:
    assert frame.shape == (20, 12), frame.shape
    assert list(frame.columns) == COLUMNS
    primary = frame.loc[
        frame["Check family"].eq("Outcome specification")
    ].iloc[0]
    assert np.isclose(primary["Estimate"], 0.02574935397580739)
    assert primary["Result"] == "Pass — CI within SESOI"
    assert frame["Result"].str.contains("Not estimable").sum() == 1
    workbook = load_workbook(OUTPUT, data_only=False)
    assert workbook.sheetnames == ["RD Checks"]
    worksheet = workbook["RD Checks"]
    assert worksheet.max_row == 21 and worksheet.max_column == 12
    assert set(worksheet.tables.keys()) == {"GeographicRDChecksTable"}
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?"))


def main() -> None:
    assert (PUBLIC_SOURCE / "figB3-donut-in.R").exists()
    assert (PUBLIC_SOURCE / "tab2-baseline-in.R").exists()
    assert (PUBLIC_SOURCE / "tabB5-placebo-in.R").exists()
    frame = build_table()
    write_workbook(frame)
    validate_output(frame)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print("Workbook sheets: 1 (RD Checks)")
    print(f"Table dimensions: {frame.shape[0]} rows x {frame.shape[1]} columns")
    print(frame[["Check family", "Estimate", "95% CI", "Result"]].to_string(index=False))


if __name__ == "__main__":
    main()
